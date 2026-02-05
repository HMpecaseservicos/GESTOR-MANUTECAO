"""
════════════════════════════════════════════════════════════════════════════════
                    SISTEMA DE GESTÃO DE MANUTENÇÃO DE FROTA
                              Versão Profissional 2.0
════════════════════════════════════════════════════════════════════════════════

Sistema completo para gerenciamento de manutenção de veículos e máquinas,
incluindo controle de estoque de peças, fornecedores, relatórios e dashboard
analítico com alertas inteligentes.

CARACTERÍSTICAS PRINCIPAIS:
✅ Gestão completa de veículos e manutenções
✅ Controle automático de estoque de peças
✅ Sistema de fornecedores e parceiros
✅ Relatórios em PDF profissionais
✅ Dashboard com métricas e alertas
✅ Interface responsiva e moderna
✅ Chatbot assistente inteligente

TECNOLOGIAS:
- Backend: Flask (Python)
- Frontend: Bootstrap 5 + JavaScript
- Banco: SQLite com triggers e índices
- Relatórios: ReportLab PDF
- UI/UX: Font Awesome + CSS3

DESENVOLVIDO: Outubro 2025
VERSÃO: 2.0.0 Professional
════════════════════════════════════════════════════════════════════════════════
"""

# =============================================
# IMPORTAÇÕES E CONFIGURAÇÕES
# =============================================

from flask import Flask, render_template, request, jsonify, redirect, url_for, send_file, Response, flash, session, g
from flask_wtf.csrf import CSRFProtect
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_login import login_required, current_user, login_user, logout_user
import sqlite3
from datetime import datetime
import os
import traceback
import logging
from logging.handlers import RotatingFileHandler
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import json
import csv
import codecs
from io import StringIO, BytesIO

# Importar módulos personalizados
from config import Config
from auth import login_manager, bcrypt, init_auth_tables, authenticate_user, admin_required, tecnico_required, log_action
from database_manager import db_manager
from empresa_helpers import super_admin_required

# =============================================
# CONFIGURAÇÃO DA APLICAÇÃO
# =============================================

app = Flask(__name__)
app.config.from_object(Config)

# Garantir que diretórios existem
Config.ensure_directories()

# Configurações do sistema (compatibilidade)
DATABASE = Config.DATABASE_PATH
REPORTS_DIR = Config.REPORTS_FOLDER
UPLOAD_FOLDER = Config.UPLOAD_FOLDER

# =============================================
# HELPER DE CONEXÃO COM BANCO DE DADOS
# =============================================

def get_db_connection():
    """
    Retorna conexão com o banco de dados correto (PostgreSQL ou SQLite).
    IMPORTANTE: Sempre usar esta função em vez de sqlite3.connect(DATABASE)
    """
    if Config.IS_POSTGRES:
        import psycopg2
        return psycopg2.connect(Config.DATABASE_URL)
    else:
        return sqlite3.connect(DATABASE)


def get_db_cursor(conn):
    """
    Retorna cursor do banco. Para PostgreSQL, retorna cursor com RealDictCursor.
    """
    if Config.IS_POSTGRES:
        from psycopg2.extras import RealDictCursor
        return conn.cursor(cursor_factory=RealDictCursor)
    else:
        return conn.cursor()


# Inicializar extensões de segurança
csrf = CSRFProtect(app)
login_manager.init_app(app)
bcrypt.init_app(app)

# Configurar Rate Limiting
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["200 per day", "50 per hour"],
    storage_uri=Config.RATELIMIT_STORAGE_URL,
    enabled=Config.RATELIMIT_ENABLED
)

# =============================================
# EXECUÇÃO AUTOMÁTICA DE MIGRAÇÕES (POSTGRESQL)
# =============================================

def run_migrations_on_startup():
    """Executa migrações pendentes no startup (PostgreSQL apenas)"""
    import sys as _sys
    if Config.IS_POSTGRES and Config.DATABASE_URL:
        try:
            print("🔧 PostgreSQL detectado - Executando migrações automaticamente...", flush=True)
            _sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'migrations'))
            from migration_manager import MigrationManager
            
            manager = MigrationManager(database_url=Config.DATABASE_URL)
            result = manager.run_pending_migrations()
            if result.get('success'):
                print(f"✅ Migrações concluídas! {result.get('migrations_run', 0)} aplicada(s)", flush=True)
            else:
                print(f"⚠️  Migrações com erro: {result.get('errors', [])}", flush=True)
        except Exception as e:
            print(f"⚠️  Erro ao executar migrações: {e}", flush=True)
            import traceback
            traceback.print_exc()
            _sys.stdout.flush()

# Executar migrações no startup
run_migrations_on_startup()

# Configurar página de login
login_manager.login_view = 'login'
login_manager.login_message = 'Por favor, faça login para acessar esta página.'
login_manager.login_message_category = 'info'

# Configurar logging
if not app.debug:
    if not os.path.exists(Config.LOG_FOLDER):
        os.mkdir(Config.LOG_FOLDER)
    file_handler = RotatingFileHandler(Config.LOG_FILE, maxBytes=10240000, backupCount=10)
    file_handler.setFormatter(logging.Formatter(
        '%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]'
    ))
    file_handler.setLevel(logging.INFO)
    app.logger.addHandler(file_handler)
    app.logger.setLevel(logging.INFO)
    app.logger.info('Sistema de Gestão de Frota iniciado')

# =============================================
# CRON JOBS EM BACKGROUND (ETAPA 14)
# =============================================

def iniciar_cron_jobs():
    """
    Inicia as tarefas automáticas em background.
    Só executa em produção (PostgreSQL) e apenas uma vez por container.
    """
    import threading
    import time
    from datetime import datetime
    
    # Flag para evitar múltiplas inicializações
    if hasattr(app, '_cron_iniciado') and app._cron_iniciado:
        return
    
    app._cron_iniciado = True
    
    print("🕐 Iniciando scheduler de cron jobs em background...", flush=True)
    
    def cron_loop():
        """Loop principal do cron"""
        # Aguardar 5 minutos após startup para primeira execução
        time.sleep(300)
        
        ultima_execucao_horaria = None
        ultima_execucao_diaria = None
        
        while True:
            try:
                agora = datetime.now()
                
                # Tarefa horária (a cada 1 hora)
                if ultima_execucao_horaria is None or (agora - ultima_execucao_horaria).seconds >= 3600:
                    print(f"[CRON {agora}] 🚀 Executando tarefas horárias...", flush=True)
                    try:
                        from cron_jobs import executar_tarefas_horarias
                        executar_tarefas_horarias()
                    except Exception as e:
                        print(f"[CRON] ❌ Erro nas tarefas horárias: {e}", flush=True)
                    ultima_execucao_horaria = agora
                
                # Tarefa diária (às 06:00)
                if agora.hour == 6:
                    if ultima_execucao_diaria is None or ultima_execucao_diaria.date() < agora.date():
                        print(f"[CRON {agora}] 🚀 Executando tarefas diárias...", flush=True)
                        try:
                            from cron_jobs import verificar_limites_proximos
                            verificar_limites_proximos()
                        except Exception as e:
                            print(f"[CRON] ❌ Erro nas tarefas diárias: {e}", flush=True)
                        ultima_execucao_diaria = agora
                
            except Exception as e:
                print(f"[CRON] ❌ Erro no loop de cron: {e}", flush=True)
            
            # Verificar a cada 5 minutos
            time.sleep(300)
    
    # Iniciar thread de cron
    cron_thread = threading.Thread(target=cron_loop, daemon=True, name="cron_jobs")
    cron_thread.start()
    print("✅ Scheduler de cron jobs iniciado com sucesso!", flush=True)

# Iniciar cron jobs apenas em produção (PostgreSQL)
if Config.IS_POSTGRES and os.environ.get('FLASK_ENV') != 'development':
    iniciar_cron_jobs()

# =============================================
# CONTEXTO GLOBAL - EMPRESA DO USUÁRIO LOGADO
# =============================================

@app.before_request
def load_empresa_context():
    """
    Carrega os dados da empresa do usuário logado no contexto global g.
    Disponibiliza g.empresa com tipo_operacao para uso em toda a aplicação.
    """
    g.empresa = None
    g.tipo_operacao = 'FROTA'  # Default seguro
    
    if current_user.is_authenticated and hasattr(current_user, 'empresa_id') and current_user.empresa_id:
        try:
            if Config.IS_POSTGRES:
                import psycopg2
                conn = psycopg2.connect(Config.DATABASE_URL)
                cursor = conn.cursor()
                cursor.execute(
                    "SELECT id, nome, tipo_operacao, plano FROM empresas WHERE id = %s",
                    (current_user.empresa_id,)
                )
            else:
                conn = sqlite3.connect(DATABASE)
                cursor = conn.cursor()
                cursor.execute(
                    "SELECT id, nome, tipo_operacao, plano FROM empresas WHERE id = ?",
                    (current_user.empresa_id,)
                )
            
            row = cursor.fetchone()
            if row:
                g.empresa = {
                    'id': row[0],
                    'nome': row[1],
                    'tipo_operacao': row[2] or 'FROTA',  # Fallback para empresas antigas
                    'plano': row[3]
                }
                g.tipo_operacao = g.empresa['tipo_operacao']
            
            cursor.close()
            conn.close()
        except Exception as e:
            app.logger.warning(f"Erro ao carregar contexto da empresa: {e}")


@app.context_processor
def inject_empresa_context():
    """
    Injeta variáveis de contexto da empresa em todos os templates.
    Permite usar {{ empresa }}, {{ is_frota }}, {{ is_servico }} nos templates.
    """
    from empresa_helpers import is_frota, is_servico
    return {
        'empresa': getattr(g, 'empresa', None),
        'tipo_operacao': getattr(g, 'tipo_operacao', 'FROTA'),
        'is_frota': is_frota(),
        'is_servico': is_servico()
    }

# Handler de erros globais
@app.errorhandler(404)
def not_found_error(error):
    return render_template('errors/404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    app.logger.error(f'Erro interno: {error}')
    return render_template('errors/500.html'), 500

@app.errorhandler(403)
def forbidden_error(error):
    return render_template('errors/403.html'), 403

# Filtro personalizado para formatação de moeda brasileira
@app.template_filter('moeda_br')
def moeda_br_filter(valor):
    """Formatar valor como moeda brasileira"""
    if valor is None:
        return 'R$ 0,00'
    try:
        return f"R$ {float(valor):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    except (ValueError, TypeError):
        return 'R$ 0,00'

# =============================================
# SISTEMA DE BANCO DE DADOS PROFISSIONAL
# =============================================

# NOTA: A função get_db_connection() principal está definida acima (linha ~80)
# com suporte a PostgreSQL e SQLite

def get_db_connection_optimized():
    """
    Cria uma conexão otimizada (apenas SQLite - para desenvolvimento)
    Em produção (PostgreSQL), usa conexão padrão
    """
    if Config.IS_POSTGRES:
        import psycopg2
        return psycopg2.connect(Config.DATABASE_URL)
    else:
        conn = sqlite3.connect(DATABASE, timeout=30.0)
        conn.execute('PRAGMA journal_mode=WAL;')
        conn.execute('PRAGMA synchronous=NORMAL;')
        conn.execute('PRAGMA temp_store=MEMORY;')
        conn.execute('PRAGMA mmap_size=268435456;')  # 256MB
        return conn

# Usar database_manager para inicialização
def init_db():
    """Inicializar banco de dados usando database_manager"""
    success = db_manager.init_database()
    if success:
        # Inicializar tabelas de autenticação
        conn = get_db_connection()
        init_auth_tables(conn)
        conn.close()
        # Inserir dados de exemplo se necessário
        db_manager.insert_sample_data()
    return success


# =============================================
# LANÇAMENTO FINANCEIRO AUTOMÁTICO
# =============================================

def lancar_financeiro_manutencao(manutencao_id, cursor, empresa_id, is_servico_mode):
    """
    Lança registro financeiro automaticamente para uma manutenção.
    
    REGRA DE OURO:
    - FROTA → DESPESA (custo)
    - SERVICO → ENTRADA (receita)
    
    Args:
        manutencao_id: ID da manutenção
        cursor: Cursor do banco de dados (já aberto)
        empresa_id: ID da empresa
        is_servico_mode: True se empresa é SERVICO, False se FROTA
    
    Returns:
        dict com success e message
    """
    try:
        # Verificar idempotência - já foi lançado?
        cursor.execute("""
            SELECT financeiro_lancado_em, financeiro_tipo, status, custo_total, valor_total_servicos
            FROM manutencoes WHERE id = %s
        """, (manutencao_id,))
        manut = cursor.fetchone()
        
        if not manut:
            return {'success': False, 'message': 'Manutenção não encontrada'}
        
        # Se já foi lançado, não lançar novamente (idempotência)
        if manut[0] is not None:
            print(f"   ⚠️ Financeiro já lançado em {manut[0]} como {manut[1]}")
            return {'success': True, 'message': 'Lançamento já realizado anteriormente', 'already_done': True}
        
        status = manut[2]
        custo_total = float(manut[3] or 0)
        valor_servicos = float(manut[4] or 0)
        
        # Determinar tipo de lançamento e valor
        if is_servico_mode:
            tipo_lancamento = 'ENTRADA'
            # Para SERVICO, calcular valor dos serviços prestados
            cursor.execute("""
                SELECT COALESCE(SUM(subtotal), 0) FROM manutencao_servicos WHERE manutencao_id = %s
            """, (manutencao_id,))
            valor = float(cursor.fetchone()[0] or 0)
            
            # Atualizar valor_total_servicos na manutenção
            cursor.execute("""
                UPDATE manutencoes SET valor_total_servicos = %s WHERE id = %s
            """, (valor, manutencao_id))
            
            # Atualizar OS com status FATURADO
            cursor.execute("""
                UPDATE ordens_servico 
                SET status = 'FATURADO', data_faturamento = CURRENT_TIMESTAMP, updated_at = CURRENT_TIMESTAMP
                WHERE manutencao_id = %s
            """, (manutencao_id,))
        else:
            tipo_lancamento = 'DESPESA'
            valor = custo_total
        
        if valor <= 0:
            print(f"   ⚠️ Valor zero, lançamento registrado mas sem valor")
        
        # Marcar como lançado (IDEMPOTÊNCIA)
        cursor.execute("""
            UPDATE manutencoes 
            SET financeiro_lancado_em = CURRENT_TIMESTAMP, 
                financeiro_tipo = %s
            WHERE id = %s
        """, (tipo_lancamento, manutencao_id))
        
        print(f"   ✅ Lançamento financeiro: {tipo_lancamento} de R$ {valor:.2f} para manutenção #{manutencao_id}")
        
        return {
            'success': True, 
            'message': f'Lançamento {tipo_lancamento} registrado',
            'tipo': tipo_lancamento,
            'valor': valor
        }
        
    except Exception as e:
        print(f"   ❌ Erro ao lançar financeiro: {e}")
        import traceback
        traceback.print_exc()
        return {'success': False, 'message': str(e)}

# =============================================
# ROTAS DE SISTEMA (SEM AUTENTICAÇÃO)
# =============================================

@app.route('/health')
@limiter.exempt
def health_check():
    """
    Endpoint de health check para Fly.io e monitoramento
    Verifica conectividade com banco de dados
    """
    try:
        if Config.IS_POSTGRES:
            # PostgreSQL
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            cursor.execute("SELECT 1")
            cursor.close()
            conn.close()
        else:
            # SQLite
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            cursor.execute("SELECT 1")
            cursor.close()
            conn.close()
        
        return jsonify({
            'status': 'healthy',
            'database': 'connected',
            'db_type': 'PostgreSQL' if Config.IS_POSTGRES else 'SQLite'
        }), 200
    except Exception as e:
        return jsonify({
            'status': 'unhealthy',
            'error': str(e),
            'database': 'disconnected'
        }), 503

# =============================================
# ROTAS DE AUTENTICAÇÃO
# =============================================

@app.route('/cadastro', methods=['GET', 'POST'])
@limiter.limit("3 per minute")
def cadastro():
    """Rota de auto-cadastro de empresa e usuário"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        # Dados da empresa
        empresa_nome = request.form.get('empresa_nome', '').strip()
        empresa_cnpj = request.form.get('empresa_cnpj', '').strip()
        empresa_telefone = request.form.get('empresa_telefone', '').strip()
        empresa_email = request.form.get('empresa_email', '').strip()
        plano = request.form.get('plano', 'basico')
        
        # Tipo de operação (ETAPA 1 HÍBRIDO)
        tipo_operacao = request.form.get('tipo_operacao', 'FROTA')
        if tipo_operacao not in ('FROTA', 'SERVICO'):
            tipo_operacao = 'FROTA'  # Fallback seguro
        
        # Dados do usuário
        nome = request.form.get('nome', '').strip()
        email = request.form.get('email', '').strip()
        username = request.form.get('username', '').strip().lower()
        telefone = request.form.get('telefone', '').strip()
        password = request.form.get('password', '')
        password_confirm = request.form.get('password_confirm', '')
        
        # Validações
        erros = []
        
        if not empresa_nome:
            erros.append('Nome da empresa é obrigatório')
        if not nome:
            erros.append('Seu nome é obrigatório')
        if not email:
            erros.append('Email é obrigatório')
        if not username:
            erros.append('Nome de usuário é obrigatório')
        if len(password) < 6:
            erros.append('Senha deve ter no mínimo 6 caracteres')
        if password != password_confirm:
            erros.append('As senhas não coincidem')
        
        # Verificar se username já existe
        conn = get_db_connection()
        cursor = conn.cursor()
        
        if Config.IS_POSTGRES:
            cursor.execute('SELECT id FROM usuarios WHERE username = %s', (username,))
        else:
            cursor.execute('SELECT id FROM usuarios WHERE username = ?', (username,))
        if cursor.fetchone():
            erros.append('Este nome de usuário já está em uso')
        
        if Config.IS_POSTGRES:
            cursor.execute('SELECT id FROM usuarios WHERE email = %s', (email,))
        else:
            cursor.execute('SELECT id FROM usuarios WHERE email = ?', (email,))
        if cursor.fetchone():
            erros.append('Este email já está cadastrado')
        
        if erros:
            for erro in erros:
                flash(erro, 'danger')
            conn.close()
            return render_template('cadastro.html')
        
        try:
            # Definir limites baseado no plano
            limites = {
                'basico': {'veiculos': 10, 'usuarios': 2},
                'profissional': {'veiculos': 50, 'usuarios': 10},
                'enterprise': {'veiculos': 999999, 'usuarios': 999999}
            }
            
            limite = limites.get(plano, limites['basico'])
            
            # Criar a empresa (com tipo_operacao - ETAPA 1 HÍBRIDO)
            if Config.IS_POSTGRES:
                cursor.execute('''
                    INSERT INTO empresas (nome, cnpj, telefone, email, plano, ativo, 
                                         limite_veiculos, limite_usuarios, tipo_operacao, data_criacao)
                    VALUES (%s, %s, %s, %s, %s, true, %s, %s, %s, NOW())
                    RETURNING id
                ''', (empresa_nome, empresa_cnpj, empresa_telefone, empresa_email, 
                      plano, limite['veiculos'], limite['usuarios'], tipo_operacao))
                empresa_id = cursor.fetchone()[0]
            else:
                cursor.execute('''
                    INSERT INTO empresas (nome, cnpj, telefone, email, plano, ativo, 
                                         limite_veiculos, limite_usuarios, tipo_operacao, data_criacao)
                    VALUES (?, ?, ?, ?, ?, 1, ?, ?, ?, datetime('now'))
                ''', (empresa_nome, empresa_cnpj, empresa_telefone, empresa_email, 
                      plano, limite['veiculos'], limite['usuarios'], tipo_operacao))
                empresa_id = cursor.lastrowid
            
            # Criar o usuário administrador
            password_hash = bcrypt.generate_password_hash(password).decode('utf-8')
            
            if Config.IS_POSTGRES:
                cursor.execute('''
                    INSERT INTO usuarios (username, password_hash, nome, email, telefone, 
                                         role, empresa_id, ativo, data_criacao)
                    VALUES (%s, %s, %s, %s, %s, 'Admin', %s, true, NOW())
                ''', (username, password_hash, nome, email, telefone, empresa_id))
            else:
                cursor.execute('''
                    INSERT INTO usuarios (username, password_hash, nome, email, telefone, 
                                         role, empresa_id, ativo, data_criacao)
                    VALUES (?, ?, ?, ?, ?, 'Admin', ?, 1, datetime('now'))
                ''', (username, password_hash, nome, email, telefone, empresa_id))
            
            conn.commit()
            conn.close()
            
            flash(f'Conta criada com sucesso! Empresa "{empresa_nome}" cadastrada. Faça login para continuar.', 'success')
            return redirect(url_for('login'))
            
        except Exception as e:
            conn.rollback()
            conn.close()
            print(f"Erro ao criar cadastro: {e}")
            flash('Erro ao criar cadastro. Tente novamente.', 'danger')
            return render_template('cadastro.html')
    
    return render_template('cadastro.html')


@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def login():
    """Rota de login"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        remember = request.form.get('remember', False)
        
        user, error = authenticate_user(username, password)
        
        if user:
            login_user(user, remember=remember)
            flash(f'Bem-vindo, {user.username}!', 'success')
            
            # Redirecionar para página solicitada ou dashboard
            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('dashboard'))
        else:
            flash(error or 'Usuário ou senha incorretos', 'danger')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    """Rota de logout"""
    if current_user.is_authenticated:
        log_action(current_user.id, 'LOGOUT', 'Logout realizado')
    logout_user()
    flash('Você saiu do sistema com sucesso.', 'info')
    return redirect(url_for('login'))


# =============================================
# ROTAS DE CONFIGURAÇÕES DO USUÁRIO
# =============================================

@app.route('/configuracoes')
@login_required
def configuracoes():
    """Página de configurações do usuário"""
    try:
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            
            # Buscar dados do usuário
            cursor.execute('SELECT id, username, email, nome, telefone, role, empresa_id FROM usuarios WHERE id = %s', 
                           (current_user.id,))
            row = cursor.fetchone()
            usuario = {
                'id': row[0],
                'username': row[1],
                'email': row[2],
                'nome': row[3],
                'telefone': row[4],
                'role': row[5],
                'empresa_id': row[6]
            }
            
            # Buscar dados da empresa
            empresa = None
            if usuario['empresa_id']:
                cursor.execute('SELECT id, nome, plano FROM empresas WHERE id = %s', (usuario['empresa_id'],))
                emp_row = cursor.fetchone()
                if emp_row:
                    empresa = {'id': emp_row[0], 'nome': emp_row[1], 'plano': emp_row[2]}
            
            cursor.close()
            conn.close()
        else:
            conn = get_db_connection()
            cursor = conn.cursor()
            
            # Buscar dados do usuário
            cursor.execute('SELECT id, username, email, nome, telefone, role, empresa_id FROM usuarios WHERE id = ?', 
                           (current_user.id,))
            row = cursor.fetchone()
            usuario = {
                'id': row[0],
                'username': row[1],
                'email': row[2],
                'nome': row[3],
                'telefone': row[4],
                'role': row[5],
                'empresa_id': row[6]
            }
            
            # Buscar dados da empresa
            empresa = None
            if usuario['empresa_id']:
                cursor.execute('SELECT id, nome, plano FROM empresas WHERE id = ?', (usuario['empresa_id'],))
                emp_row = cursor.fetchone()
                if emp_row:
                    empresa = {'id': emp_row[0], 'nome': emp_row[1], 'plano': emp_row[2]}
            
            conn.close()
        
        return render_template('configuracoes.html', usuario=usuario, empresa=empresa)
    except Exception as e:
        print(f"Erro em configuracoes: {e}")
        import traceback
        traceback.print_exc()
        flash('Erro ao carregar configurações.', 'danger')
        return redirect(url_for('dashboard'))


@app.route('/atualizar_perfil', methods=['POST'])
@login_required
def atualizar_perfil():
    """Atualizar dados do perfil do usuário"""
    nome = request.form.get('nome', '').strip()
    email = request.form.get('email', '').strip()
    telefone = request.form.get('telefone', '').strip()
    
    try:
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            
            # Verificar se email já está em uso por outro usuário
            cursor.execute('SELECT id FROM usuarios WHERE email = %s AND id != %s', (email, current_user.id))
            if cursor.fetchone():
                flash('Este email já está em uso por outro usuário.', 'danger')
                cursor.close()
                conn.close()
                return redirect(url_for('configuracoes'))
            
            cursor.execute('''
                UPDATE usuarios SET nome = %s, email = %s, telefone = %s WHERE id = %s
            ''', (nome, email, telefone, current_user.id))
            
            conn.commit()
            cursor.close()
            conn.close()
        else:
            conn = get_db_connection()
            cursor = conn.cursor()
            
            # Verificar se email já está em uso por outro usuário
            cursor.execute('SELECT id FROM usuarios WHERE email = ? AND id != ?', (email, current_user.id))
            if cursor.fetchone():
                flash('Este email já está em uso por outro usuário.', 'danger')
                conn.close()
                return redirect(url_for('configuracoes'))
            
            cursor.execute('''
                UPDATE usuarios SET nome = ?, email = ?, telefone = ? WHERE id = ?
            ''', (nome, email, telefone, current_user.id))
            
            conn.commit()
            conn.close()
        
        flash('Perfil atualizado com sucesso!', 'success')
    except Exception as e:
        print(f"Erro ao atualizar perfil: {e}")
        flash('Erro ao atualizar perfil.', 'danger')
    
    return redirect(url_for('configuracoes'))


@app.route('/alterar_minha_senha', methods=['POST'])
@login_required
def alterar_minha_senha():
    """Alterar senha do usuário via configurações"""
    senha_atual = request.form.get('senha_atual', '')
    nova_senha = request.form.get('nova_senha', '')
    confirmar_senha = request.form.get('confirmar_senha', '')
    
    if nova_senha != confirmar_senha:
        flash('As senhas não coincidem.', 'danger')
        return redirect(url_for('configuracoes'))
    
    if len(nova_senha) < 6:
        flash('A nova senha deve ter no mínimo 6 caracteres.', 'danger')
        return redirect(url_for('configuracoes'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Verificar senha atual
    cursor.execute('SELECT password_hash FROM usuarios WHERE id = ?', (current_user.id,))
    row = cursor.fetchone()
    
    if not row or not bcrypt.check_password_hash(row[0], senha_atual):
        flash('Senha atual incorreta.', 'danger')
        conn.close()
        return redirect(url_for('configuracoes'))
    
    # Atualizar senha
    nova_hash = bcrypt.generate_password_hash(nova_senha).decode('utf-8')
    cursor.execute('UPDATE usuarios SET password_hash = ? WHERE id = ?', (nova_hash, current_user.id))
    
    conn.commit()
    conn.close()
    
    flash('Senha alterada com sucesso!', 'success')
    return redirect(url_for('configuracoes'))


# =============================================
# ROTAS DE PLANOS E MONETIZAÇÃO (ETAPA 15)
# =============================================

def is_demo_user():
    """Verifica se o usuário atual é um usuário demo"""
    if not current_user.is_authenticated:
        return False
    return getattr(current_user, 'is_demo', False)


@app.route('/planos')
def planos():
    """Página pública de planos e preços"""
    from empresa_helpers import get_planos_disponiveis, get_plano, get_plano_info, get_plano_upgrade_sugerido
    
    # Bloquear usuários demo
    if is_demo_user():
        flash('Acesso restrito. Usuários de demonstração não podem acessar a área de planos.', 'warning')
        return redirect(url_for('dashboard'))
    
    planos_lista = get_planos_disponiveis()
    
    plano_atual = None
    plano_upgrade = None
    
    # Se usuário estiver logado, mostrar plano atual
    if current_user.is_authenticated:
        plano_id = get_plano()
        plano_atual = get_plano_info(plano_id)
        plano_upgrade = get_plano_upgrade_sugerido(plano_id)
    
    return render_template('planos.html',
                           planos=planos_lista,
                           plano_atual=plano_atual,
                           plano_upgrade=plano_upgrade)


@app.route('/api/solicitar-upgrade', methods=['POST'])
def api_solicitar_upgrade():
    """
    API para registrar solicitação de upgrade.
    Cria notificação interna e pode enviar email futuramente.
    """
    from empresa_helpers import create_notification, get_empresa_id
    
    # Bloquear usuários demo
    if is_demo_user():
        return jsonify({'success': False, 'message': 'Usuários de demonstração não podem solicitar upgrades'}), 403
    
    data = request.form
    nome = data.get('nome', '').strip()
    email = data.get('email', '').strip()
    telefone = data.get('telefone', '').strip()
    plano_solicitado = data.get('plano_solicitado', '').strip()
    mensagem = data.get('mensagem', '').strip()
    
    if not nome or not email:
        return jsonify({'success': False, 'message': 'Nome e email são obrigatórios'}), 400
    
    try:
        # Registrar no log
        app.logger.info(f"[UPGRADE] Solicitação recebida: {nome} ({email}) -> {plano_solicitado}")
        
        # Se usuário logado, criar notificação para admins
        if current_user.is_authenticated:
            empresa_id = get_empresa_id()
            
            # Criar notificação interna para admins
            titulo = f"📈 Solicitação de upgrade para {plano_solicitado}"
            msg = f"Usuário {nome} ({email}) solicitou upgrade.\n"
            if telefone:
                msg += f"Telefone: {telefone}\n"
            if mensagem:
                msg += f"Mensagem: {mensagem}"
            
            create_notification(empresa_id, 'SISTEMA', titulo, msg, link='/planos')
        
        # Registrar solicitação no banco (se quiser persistir)
        if Config.IS_POSTGRES:
            try:
                import psycopg2
                conn = psycopg2.connect(Config.DATABASE_URL)
                cursor = conn.cursor()
                
                # Criar tabela de solicitações se não existir
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS solicitacoes_upgrade (
                        id BIGSERIAL PRIMARY KEY,
                        empresa_id BIGINT,
                        nome VARCHAR(200),
                        email VARCHAR(200),
                        telefone VARCHAR(50),
                        plano_solicitado VARCHAR(50),
                        mensagem TEXT,
                        status VARCHAR(20) DEFAULT 'PENDENTE',
                        created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
                    )
                """)
                
                # Inserir solicitação
                cursor.execute("""
                    INSERT INTO solicitacoes_upgrade (empresa_id, nome, email, telefone, plano_solicitado, mensagem)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, (
                    get_empresa_id() if current_user.is_authenticated else None,
                    nome, email, telefone, plano_solicitado, mensagem
                ))
                
                conn.commit()
                cursor.close()
                conn.close()
            except Exception as e:
                app.logger.error(f"Erro ao salvar solicitação: {e}")
        
        return jsonify({
            'success': True,
            'message': 'Solicitação enviada com sucesso! Nossa equipe entrará em contato em breve.'
        })
        
    except Exception as e:
        app.logger.error(f"Erro na solicitação de upgrade: {e}")
        return jsonify({'success': False, 'message': 'Erro ao processar solicitação'}), 500


@app.route('/minha-empresa')
@login_required
def minha_empresa():
    """Página de configurações da empresa - Atualizada para PostgreSQL e planos SaaS"""
    from empresa_helpers import is_servico, get_empresa_id, get_plano, get_info_plano, contar_recursos_usados
    
    empresa_id = get_empresa_id()
    
    try:
        if Config.IS_POSTGRES:
            import psycopg2
            from psycopg2.extras import RealDictCursor
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            # Buscar empresa do usuário
            cursor.execute('''
                SELECT e.* FROM empresas e
                WHERE e.id = %s
            ''', (empresa_id,))
            
            empresa = cursor.fetchone()
            if not empresa:
                flash('Empresa não encontrada.', 'warning')
                cursor.close()
                conn.close()
                return redirect(url_for('dashboard'))
            
            # Contar recursos usados
            recursos = contar_recursos_usados(cursor, empresa_id)
            
            # Estatísticas adicionais
            cursor.execute('SELECT COUNT(*) FROM manutencoes WHERE empresa_id = %s', (empresa_id,))
            manutencoes = cursor.fetchone()[0]
            
            cursor.execute('SELECT COUNT(*) FROM pecas WHERE empresa_id = %s', (empresa_id,))
            pecas = cursor.fetchone()[0]
            
            cursor.execute('SELECT COUNT(*) FROM fornecedores WHERE empresa_id = %s', (empresa_id,))
            fornecedores = cursor.fetchone()[0]
            
            cursor.execute('SELECT COUNT(*) FROM tecnicos WHERE empresa_id = %s', (empresa_id,))
            tecnicos = cursor.fetchone()[0]
            
            cursor.close()
            conn.close()
            
        else:
            conn = get_db_connection()
            cursor = conn.cursor()
            
            # Buscar empresa do usuário
            cursor.execute('''
                SELECT e.* FROM empresas e
                JOIN usuarios u ON u.empresa_id = e.id
                WHERE u.id = ?
            ''', (current_user.id,))
            
            row = cursor.fetchone()
            if not row:
                flash('Empresa não encontrada.', 'warning')
                conn.close()
                return redirect(url_for('dashboard'))
            
            cols = [desc[0] for desc in cursor.description]
            empresa = dict(zip(cols, row))
            
            # Contar recursos
            cursor.execute('SELECT COUNT(*) FROM veiculos WHERE empresa_id = ? AND ativo = 1', (empresa_id,))
            recursos = {
                'veiculos': cursor.fetchone()[0],
                'clientes': 0,
                'usuarios': 0
            }
            
            cursor.execute('SELECT COUNT(*) FROM usuarios WHERE empresa_id = ? AND ativo = 1', (empresa_id,))
            recursos['usuarios'] = cursor.fetchone()[0]
            
            cursor.execute('SELECT COUNT(*) FROM manutencoes WHERE empresa_id = ?', (empresa_id,))
            manutencoes = cursor.fetchone()[0]
            
            cursor.execute('SELECT COUNT(*) FROM pecas WHERE empresa_id = ?', (empresa_id,))
            pecas = cursor.fetchone()[0]
            
            cursor.execute('SELECT COUNT(*) FROM fornecedores WHERE empresa_id = ?', (empresa_id,))
            fornecedores = cursor.fetchone()[0]
            
            cursor.execute('SELECT COUNT(*) FROM tecnicos WHERE empresa_id = ?', (empresa_id,))
            tecnicos = cursor.fetchone()[0]
            
            conn.close()
        
        # Montar stats
        stats = {
            'veiculos': recursos.get('veiculos', 0),
            'clientes': recursos.get('clientes', 0),
            'usuarios': recursos.get('usuarios', 0),
            'manutencoes': manutencoes,
            'pecas': pecas,
            'fornecedores': fornecedores,
            'tecnicos': tecnicos
        }
        
        # Informações do plano (ETAPA 15)
        from empresa_helpers import get_plano_upgrade_sugerido
        info_plano = get_info_plano()
        plano_upgrade = get_plano_upgrade_sugerido(get_plano())
        
        return render_template('minha_empresa.html', 
                               empresa=empresa, 
                               stats=stats, 
                               info_plano=info_plano,
                               plano_upgrade=plano_upgrade,
                               is_servico=is_servico())
        
    except Exception as e:
        print(f"Erro ao carregar minha empresa: {e}")
        traceback.print_exc()
        flash('Erro ao carregar informações da empresa.', 'danger')
        return redirect(url_for('dashboard'))


@app.route('/atualizar_minha_empresa', methods=['POST'])
@login_required
def atualizar_minha_empresa():
    """Atualizar dados da empresa do usuário logado"""
    from empresa_helpers import get_empresa_id, is_admin
    
    # Apenas ADMIN pode editar dados da empresa
    if not is_admin():
        flash('Apenas administradores podem editar dados da empresa.', 'warning')
        return redirect(url_for('minha_empresa'))
    
    empresa_id = get_empresa_id()
    
    nome = request.form.get('nome', '').strip()
    cnpj = request.form.get('cnpj', '').strip()
    telefone = request.form.get('telefone', '').strip()
    email = request.form.get('email', '').strip()
    endereco = request.form.get('endereco', '').strip()
    cidade = request.form.get('cidade', '').strip()
    estado = request.form.get('estado', '').strip()
    cep = request.form.get('cep', '').strip()
    
    try:
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            
            cursor.execute('''
                UPDATE empresas 
                SET nome = %s, cnpj = %s, telefone = %s, email = %s, 
                    endereco = %s, cidade = %s, estado = %s, cep = %s
                WHERE id = %s
            ''', (nome, cnpj, telefone, email, endereco, cidade, estado, cep, empresa_id))
            
            conn.commit()
            cursor.close()
            conn.close()
        else:
            conn = get_db_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                UPDATE empresas 
                SET nome = ?, cnpj = ?, telefone = ?, email = ?, endereco = ?, cidade = ?, estado = ?, cep = ?
                WHERE id = ?
            ''', (nome, cnpj, telefone, email, endereco, cidade, estado, cep, empresa_id))
            
            conn.commit()
            conn.close()
        
        flash('Dados da empresa atualizados com sucesso!', 'success')
    except Exception as e:
        print(f"Erro ao atualizar empresa: {e}")
        traceback.print_exc()
        flash('Erro ao atualizar dados da empresa.', 'danger')
    
    return redirect(url_for('minha_empresa'))


# =============================================
# ROTAS DE NOTIFICAÇÕES (ETAPA 12)
# =============================================

@app.route('/notificacoes')
@login_required
def notificacoes():
    """Página de histórico de notificações"""
    from empresa_helpers import get_empresa_id, is_admin, get_recent_notifications, TIPOS_NOTIFICACAO
    
    empresa_id = get_empresa_id()
    usuario_id = current_user.id
    admin = is_admin()
    
    # Buscar todas as notificações (limite maior para histórico)
    notificacoes_lista = get_recent_notifications(empresa_id, usuario_id, admin, limit=100)
    
    return render_template('notificacoes.html', 
                           notificacoes=notificacoes_lista,
                           tipos=TIPOS_NOTIFICACAO)


@app.route('/api/notificacoes')
@login_required
def api_notificacoes():
    """API para buscar notificações recentes (dropdown)"""
    from empresa_helpers import get_empresa_id, is_admin, get_recent_notifications, get_unread_count
    
    empresa_id = get_empresa_id()
    usuario_id = current_user.id
    admin = is_admin()
    
    notificacoes = get_recent_notifications(empresa_id, usuario_id, admin, limit=5)
    unread_count = get_unread_count(empresa_id, usuario_id, admin)
    
    # Formatar para JSON
    result = []
    for n in notificacoes:
        result.append({
            'id': n['id'],
            'tipo': n['tipo'],
            'titulo': n['titulo'],
            'mensagem': n['mensagem'],
            'lida': n['lida'],
            'link': n['link'],
            'created_at': str(n['created_at'])[:16] if n['created_at'] else None
        })
    
    return jsonify({
        'success': True,
        'notificacoes': result,
        'unread_count': unread_count
    })


@app.route('/api/notificacoes/marcar-lida', methods=['POST'])
@login_required
def api_marcar_notificacao_lida():
    """Marcar notificação como lida"""
    from empresa_helpers import get_empresa_id, is_admin, mark_notification_read
    
    empresa_id = get_empresa_id()
    usuario_id = current_user.id
    admin = is_admin()
    
    data = request.json
    notificacao_id = data.get('notificacao_id')
    
    if not notificacao_id:
        return jsonify({'success': False, 'message': 'ID da notificação não informado'}), 400
    
    success = mark_notification_read(notificacao_id, empresa_id, usuario_id, admin)
    
    return jsonify({'success': success})


@app.route('/api/notificacoes/marcar-todas-lidas', methods=['POST'])
@login_required
def api_marcar_todas_lidas():
    """Marcar todas as notificações como lidas"""
    from empresa_helpers import get_empresa_id, is_admin, mark_all_notifications_read
    
    empresa_id = get_empresa_id()
    usuario_id = current_user.id
    admin = is_admin()
    
    success = mark_all_notifications_read(empresa_id, usuario_id, admin)
    
    return jsonify({'success': success})


# =============================================
# API DE CRON JOBS (ETAPA 14 - APENAS ADMIN)
# =============================================

@app.route('/api/cron/executar', methods=['POST'])
@login_required
def api_executar_cron():
    """
    Executa cron jobs manualmente - APENAS ADMIN.
    Útil para testes e diagnóstico.
    
    Body JSON:
    - tarefa: 'manutencoes', 'faturamento', 'limites', 'horaria', 'diaria', 'todas'
    """
    from empresa_helpers import is_admin
    
    if not is_admin():
        return jsonify({'success': False, 'message': 'Acesso restrito a administradores'}), 403
    
    data = request.get_json() or {}
    tarefa = data.get('tarefa', 'todas')
    
    try:
        from cron_jobs import (
            verificar_manutencoes_atrasadas,
            verificar_servicos_sem_faturamento,
            verificar_limites_proximos,
            executar_tarefas_horarias,
            executar_tarefas_diarias,
            executar_todas_tarefas
        )
        
        resultados = {}
        
        if tarefa == 'manutencoes':
            resultados['manutencoes_atrasadas'] = verificar_manutencoes_atrasadas()
        elif tarefa == 'faturamento':
            resultados['servicos_sem_faturamento'] = verificar_servicos_sem_faturamento()
        elif tarefa == 'limites':
            resultados['limites_proximos'] = verificar_limites_proximos()
        elif tarefa == 'horaria':
            resultados = executar_tarefas_horarias()
        elif tarefa == 'diaria':
            resultados = executar_tarefas_diarias()
        else:  # todas
            resultados = executar_todas_tarefas()
        
        return jsonify({
            'success': True, 
            'message': f'Tarefa(s) executada(s) com sucesso',
            'tarefa': tarefa,
            'resultados': resultados
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/cron/status')
@login_required
def api_cron_status():
    """Retorna status do sistema de cron jobs - APENAS ADMIN"""
    from empresa_helpers import is_admin
    
    if not is_admin():
        return jsonify({'success': False, 'message': 'Acesso restrito a administradores'}), 403
    
    cron_ativo = hasattr(app, '_cron_iniciado') and app._cron_iniciado
    
    return jsonify({
        'success': True,
        'cron_ativo': cron_ativo,
        'ambiente': 'produção' if Config.IS_POSTGRES else 'desenvolvimento',
        'tarefas_disponiveis': [
            {'id': 'manutencoes', 'nome': 'Manutenções Atrasadas', 'frequencia': 'Horária'},
            {'id': 'faturamento', 'nome': 'Serviços sem Faturamento', 'frequencia': 'Horária'},
            {'id': 'limites', 'nome': 'Verificar Limites 80%', 'frequencia': 'Diária'},
        ]
    })


# =============================================
# ROTAS DE GESTÃO DE USUÁRIOS (ETAPA 11)
# =============================================

@app.route('/usuarios')
@login_required
def usuarios():
    """Lista de usuários da empresa - APENAS ADMIN"""
    from empresa_helpers import is_admin, get_empresa_id, get_limite_usuarios
    
    if not is_admin():
        flash('Acesso restrito. Apenas administradores podem gerenciar usuários.', 'warning')
        return redirect(url_for('dashboard'))
    
    empresa_id = get_empresa_id()
    usuarios_lista = []
    stats = {'usuarios': 0, 'admins': 0, 'operadores': 0}
    
    try:
        if Config.IS_POSTGRES:
            import psycopg2
            from psycopg2.extras import RealDictCursor
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            cursor.execute("""
                SELECT id, username, nome, email, role, ativo, ultimo_login, data_criacao
                FROM usuarios 
                WHERE empresa_id = %s
                ORDER BY role DESC, nome
            """, (empresa_id,))
            usuarios_lista = cursor.fetchall()
            
            # Estatísticas
            cursor.execute("SELECT COUNT(*) as cnt FROM usuarios WHERE empresa_id = %s AND ativo = true", (empresa_id,))
            result = cursor.fetchone()
            stats['usuarios'] = result['cnt'] if result else 0
            
            cursor.execute("SELECT COUNT(*) as cnt FROM usuarios WHERE empresa_id = %s AND ativo = true AND role = 'ADMIN'", (empresa_id,))
            result = cursor.fetchone()
            stats['admins'] = result['cnt'] if result else 0
            
            cursor.execute("SELECT COUNT(*) as cnt FROM usuarios WHERE empresa_id = %s AND ativo = true AND role = 'OPERADOR'", (empresa_id,))
            result = cursor.fetchone()
            stats['operadores'] = result['cnt'] if result else 0
            
            cursor.close()
            conn.close()
        else:
            conn = sqlite3.connect(DATABASE)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT id, username, nome, email, role, ativo, ultimo_login, data_criacao
                FROM usuarios 
                WHERE empresa_id = ?
                ORDER BY role DESC, nome
            """, (empresa_id,))
            usuarios_lista = [dict(row) for row in cursor.fetchall()]
            
            cursor.execute("SELECT COUNT(*) FROM usuarios WHERE empresa_id = ? AND ativo = 1", (empresa_id,))
            stats['usuarios'] = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM usuarios WHERE empresa_id = ? AND ativo = 1 AND role = 'ADMIN'", (empresa_id,))
            stats['admins'] = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM usuarios WHERE empresa_id = ? AND ativo = 1 AND role = 'OPERADOR'", (empresa_id,))
            stats['operadores'] = cursor.fetchone()[0]
            
            conn.close()
            
    except Exception as e:
        print(f"Erro ao listar usuários: {e}")
        traceback.print_exc()
        flash('Erro ao carregar lista de usuários.', 'danger')
    
    limite_usuarios = get_limite_usuarios()
    
    return render_template('usuarios.html', 
                           usuarios=usuarios_lista, 
                           stats=stats,
                           limite_usuarios=limite_usuarios)


@app.route('/usuarios/criar', methods=['POST'])
@login_required
def criar_usuario():
    """Criar novo usuário na empresa - APENAS ADMIN"""
    from empresa_helpers import is_admin, get_empresa_id, verificar_limite_usuarios
    
    if not is_admin():
        flash('Acesso restrito. Apenas administradores podem criar usuários.', 'danger')
        return redirect(url_for('usuarios'))
    
    empresa_id = get_empresa_id()
    
    nome = request.form.get('nome', '').strip()
    email = request.form.get('email', '').strip()
    username = request.form.get('username', '').strip().lower()
    password = request.form.get('password', '')
    role = request.form.get('role', 'OPERADOR')
    
    # Validações
    if not nome:
        flash('Nome é obrigatório.', 'danger')
        return redirect(url_for('usuarios'))
    
    if not email:
        flash('Email é obrigatório.', 'danger')
        return redirect(url_for('usuarios'))
    
    if not username:
        flash('Nome de usuário é obrigatório.', 'danger')
        return redirect(url_for('usuarios'))
    
    if len(password) < 6:
        flash('Senha deve ter no mínimo 6 caracteres.', 'danger')
        return redirect(url_for('usuarios'))
    
    if role not in ('ADMIN', 'OPERADOR'):
        role = 'OPERADOR'
    
    try:
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            
            # Verificar limite de usuários
            pode_criar, msg_limite = verificar_limite_usuarios(cursor, empresa_id)
            if not pode_criar:
                cursor.close()
                conn.close()
                # Notificar sobre bloqueio de limite (ETAPA 12)
                from empresa_helpers import create_notification
                create_notification(empresa_id, 'LIMITE_BLOQUEIO', 
                    'Limite de usuários atingido!', 
                    msg_limite, 
                    link='/minha-empresa')
                flash(msg_limite, 'warning')
                return redirect(url_for('usuarios'))
            
            # Verificar se username já existe
            cursor.execute("SELECT id FROM usuarios WHERE username = %s", (username,))
            if cursor.fetchone():
                cursor.close()
                conn.close()
                flash('Este nome de usuário já está em uso.', 'danger')
                return redirect(url_for('usuarios'))
            
            # Verificar se email já existe na empresa
            cursor.execute("SELECT id FROM usuarios WHERE email = %s AND empresa_id = %s", (email, empresa_id))
            if cursor.fetchone():
                cursor.close()
                conn.close()
                flash('Este email já está cadastrado na empresa.', 'danger')
                return redirect(url_for('usuarios'))
            
            # Criar usuário
            password_hash = bcrypt.generate_password_hash(password).decode('utf-8')
            
            cursor.execute("""
                INSERT INTO usuarios (empresa_id, username, password_hash, nome, email, role, ativo, data_criacao)
                VALUES (%s, %s, %s, %s, %s, %s, true, CURRENT_TIMESTAMP)
            """, (empresa_id, username, password_hash, nome, email, role))
            
            conn.commit()
            cursor.close()
            conn.close()
            
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            
            cursor.execute("SELECT id FROM usuarios WHERE username = ?", (username,))
            if cursor.fetchone():
                conn.close()
                flash('Este nome de usuário já está em uso.', 'danger')
                return redirect(url_for('usuarios'))
            
            password_hash = bcrypt.generate_password_hash(password).decode('utf-8')
            
            cursor.execute("""
                INSERT INTO usuarios (empresa_id, username, password_hash, nome, email, role, ativo, data_criacao)
                VALUES (?, ?, ?, ?, ?, ?, 1, datetime('now'))
            """, (empresa_id, username, password_hash, nome, email, role))
            
            conn.commit()
            conn.close()
        
        flash(f'Usuário {username} criado com sucesso!', 'success')
        
        # Notificar sobre criação de usuário (ETAPA 12)
        from empresa_helpers import notify_user_created
        notify_user_created(empresa_id, current_user.id, nome, role)
        
    except Exception as e:
        print(f"Erro ao criar usuário: {e}")
        traceback.print_exc()
        flash('Erro ao criar usuário.', 'danger')
    
    return redirect(url_for('usuarios'))


@app.route('/usuarios/editar', methods=['POST'])
@login_required
def editar_usuario():
    """Editar usuário existente - APENAS ADMIN"""
    from empresa_helpers import is_admin, get_empresa_id
    
    if not is_admin():
        flash('Acesso restrito. Apenas administradores podem editar usuários.', 'danger')
        return redirect(url_for('usuarios'))
    
    empresa_id = get_empresa_id()
    
    usuario_id = request.form.get('usuario_id')
    nome = request.form.get('nome', '').strip()
    email = request.form.get('email', '').strip()
    role = request.form.get('role', 'OPERADOR')
    
    if not usuario_id:
        flash('Usuário não encontrado.', 'danger')
        return redirect(url_for('usuarios'))
    
    if not nome:
        flash('Nome é obrigatório.', 'danger')
        return redirect(url_for('usuarios'))
    
    if role not in ('ADMIN', 'OPERADOR'):
        role = 'OPERADOR'
    
    try:
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            
            # Verificar se usuário pertence à empresa
            cursor.execute("SELECT id, role FROM usuarios WHERE id = %s AND empresa_id = %s", (usuario_id, empresa_id))
            usuario_atual = cursor.fetchone()
            
            if not usuario_atual:
                cursor.close()
                conn.close()
                flash('Usuário não encontrado.', 'danger')
                return redirect(url_for('usuarios'))
            
            # Se está removendo permissão de ADMIN, verificar se não é o último
            if usuario_atual[1] == 'ADMIN' and role == 'OPERADOR':
                cursor.execute("SELECT COUNT(*) FROM usuarios WHERE empresa_id = %s AND role = 'ADMIN' AND ativo = true", (empresa_id,))
                total_admins = cursor.fetchone()[0]
                
                if total_admins <= 1:
                    cursor.close()
                    conn.close()
                    flash('Não é possível remover o último administrador da empresa.', 'danger')
                    return redirect(url_for('usuarios'))
            
            # Atualizar usuário
            cursor.execute("""
                UPDATE usuarios 
                SET nome = %s, email = %s, role = %s, updated_at = CURRENT_TIMESTAMP
                WHERE id = %s AND empresa_id = %s
            """, (nome, email, role, usuario_id, empresa_id))
            
            conn.commit()
            cursor.close()
            conn.close()
            
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            
            cursor.execute("SELECT id, role FROM usuarios WHERE id = ? AND empresa_id = ?", (usuario_id, empresa_id))
            usuario_atual = cursor.fetchone()
            
            if not usuario_atual:
                conn.close()
                flash('Usuário não encontrado.', 'danger')
                return redirect(url_for('usuarios'))
            
            if usuario_atual[1] == 'ADMIN' and role == 'OPERADOR':
                cursor.execute("SELECT COUNT(*) FROM usuarios WHERE empresa_id = ? AND role = 'ADMIN' AND ativo = 1", (empresa_id,))
                if cursor.fetchone()[0] <= 1:
                    conn.close()
                    flash('Não é possível remover o último administrador da empresa.', 'danger')
                    return redirect(url_for('usuarios'))
            
            cursor.execute("""
                UPDATE usuarios 
                SET nome = ?, email = ?, role = ?
                WHERE id = ? AND empresa_id = ?
            """, (nome, email, role, usuario_id, empresa_id))
            
            conn.commit()
            conn.close()
        
        flash('Usuário atualizado com sucesso!', 'success')
        
    except Exception as e:
        print(f"Erro ao editar usuário: {e}")
        traceback.print_exc()
        flash('Erro ao editar usuário.', 'danger')
    
    return redirect(url_for('usuarios'))


@app.route('/usuarios/toggle-status', methods=['POST'])
@login_required
def toggle_usuario_status():
    """Ativar/Desativar usuário - APENAS ADMIN"""
    from empresa_helpers import is_admin, get_empresa_id
    
    if not is_admin():
        return jsonify({'success': False, 'message': 'Acesso restrito a administradores'}), 403
    
    empresa_id = get_empresa_id()
    data = request.json
    usuario_id = data.get('usuario_id')
    
    if not usuario_id:
        return jsonify({'success': False, 'message': 'Usuário não informado'}), 400
    
    # Não permitir alterar próprio status
    if int(usuario_id) == current_user.id:
        return jsonify({'success': False, 'message': 'Você não pode alterar seu próprio status'}), 400
    
    try:
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            
            # Verificar se usuário pertence à empresa
            cursor.execute("SELECT id, ativo, role FROM usuarios WHERE id = %s AND empresa_id = %s", (usuario_id, empresa_id))
            usuario = cursor.fetchone()
            
            if not usuario:
                cursor.close()
                conn.close()
                return jsonify({'success': False, 'message': 'Usuário não encontrado'}), 404
            
            novo_status = not usuario[1]
            
            # Se está desativando um ADMIN, verificar se não é o último
            if not novo_status and usuario[2] == 'ADMIN':
                cursor.execute("SELECT COUNT(*) FROM usuarios WHERE empresa_id = %s AND role = 'ADMIN' AND ativo = true", (empresa_id,))
                total_admins = cursor.fetchone()[0]
                
                if total_admins <= 1:
                    cursor.close()
                    conn.close()
                    return jsonify({'success': False, 'message': 'Não é possível desativar o último administrador da empresa'}), 400
            
            cursor.execute("""
                UPDATE usuarios SET ativo = %s, updated_at = CURRENT_TIMESTAMP
                WHERE id = %s AND empresa_id = %s
            """, (novo_status, usuario_id, empresa_id))
            
            conn.commit()
            cursor.close()
            conn.close()
            
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            
            cursor.execute("SELECT id, ativo, role FROM usuarios WHERE id = ? AND empresa_id = ?", (usuario_id, empresa_id))
            usuario = cursor.fetchone()
            
            if not usuario:
                conn.close()
                return jsonify({'success': False, 'message': 'Usuário não encontrado'}), 404
            
            novo_status = 0 if usuario[1] else 1
            
            if not novo_status and usuario[2] == 'ADMIN':
                cursor.execute("SELECT COUNT(*) FROM usuarios WHERE empresa_id = ? AND role = 'ADMIN' AND ativo = 1", (empresa_id,))
                if cursor.fetchone()[0] <= 1:
                    conn.close()
                    return jsonify({'success': False, 'message': 'Não é possível desativar o último administrador da empresa'}), 400
            
            cursor.execute("UPDATE usuarios SET ativo = ? WHERE id = ? AND empresa_id = ?", (novo_status, usuario_id, empresa_id))
            
            conn.commit()
            conn.close()
        
        return jsonify({'success': True, 'message': 'Status alterado com sucesso'})
        
    except Exception as e:
        print(f"Erro ao alterar status do usuário: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'message': 'Erro ao alterar status'}), 500


# =============================================
# ROTAS PRINCIPAIS (COM PROTEÇÃO)
# =============================================

@app.route('/')
@login_required
def dashboard():
    from empresa_helpers import get_empresa_id
    
    empresa_id = get_empresa_id()
    
    try:
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            
            # Estatísticas gerais (filtrado por empresa)
            cursor.execute("SELECT COUNT(*) FROM veiculos WHERE empresa_id = %s", (empresa_id,))
            total_veiculos = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM manutencoes WHERE empresa_id = %s AND status IN ('Agendada', 'ORCAMENTO', 'APROVADO', 'EM_EXECUCAO')", (empresa_id,))
            manutencoes_pendentes = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM manutencoes WHERE empresa_id = %s AND status IN ('Concluída', 'FINALIZADO', 'FATURADO') AND data_realizada >= CURRENT_DATE - INTERVAL '30 days'", (empresa_id,))
            manutencoes_30_dias = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM tecnicos WHERE empresa_id = %s AND status = 'Ativo'", (empresa_id,))
            tecnicos_ativos = cursor.fetchone()[0]
            
            # Próximas manutenções
            cursor.execute('''
                SELECT v.placa, v.modelo, m.tipo, m.data_agendada 
                FROM manutencoes m 
                JOIN veiculos v ON m.veiculo_id = v.id 
                WHERE m.empresa_id = %s AND m.status IN ('Agendada', 'ORCAMENTO', 'APROVADO', 'EM_EXECUCAO')
                ORDER BY m.data_agendada 
                LIMIT 5
            ''', (empresa_id,))
            proximas_manutencoes = cursor.fetchall()
            
            # Alertas de estoque baixo
            cursor.execute("SELECT nome, quantidade_estoque FROM pecas WHERE empresa_id = %s AND quantidade_estoque < 3", (empresa_id,))
            alertas_estoque = cursor.fetchall()
            
            cursor.close()
            conn.close()
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            
            # Estatísticas gerais
            cursor.execute("SELECT COUNT(*) FROM veiculos WHERE empresa_id = ?", (empresa_id,))
            total_veiculos = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM manutencoes WHERE empresa_id = ? AND status = 'Agendada'", (empresa_id,))
            manutencoes_pendentes = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM manutencoes WHERE empresa_id = ? AND status = 'Concluída' AND data_realizada >= date('now', '-30 days')", (empresa_id,))
            manutencoes_30_dias = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM tecnicos WHERE empresa_id = ? AND status = 'Ativo'", (empresa_id,))
            tecnicos_ativos = cursor.fetchone()[0]
            
            # Próximas manutenções
            cursor.execute('''
                SELECT v.placa, v.modelo, m.tipo, m.data_agendada 
                FROM manutencoes m 
                JOIN veiculos v ON m.veiculo_id = v.id 
                WHERE m.empresa_id = ? AND m.status = 'Agendada' 
                ORDER BY m.data_agendada 
                LIMIT 5
            ''', (empresa_id,))
            proximas_manutencoes = cursor.fetchall()
            
            # Alertas de estoque baixo
            cursor.execute("SELECT nome, quantidade_estoque FROM pecas WHERE empresa_id = ? AND quantidade_estoque < 3", (empresa_id,))
            alertas_estoque = cursor.fetchall()
            
            conn.close()
        
        return render_template('dashboard.html', 
                             total_veiculos=total_veiculos,
                             manutencoes_pendentes=manutencoes_pendentes,
                             manutencoes_30_dias=manutencoes_30_dias,
                             tecnicos_ativos=tecnicos_ativos,
                             proximas_manutencoes=proximas_manutencoes,
                             alertas_estoque=alertas_estoque)
    except Exception as e:
        print(f"Erro no dashboard: {e}")
        import traceback
        traceback.print_exc()
        return render_template('dashboard.html', 
                             total_veiculos=0,
                             manutencoes_pendentes=0,
                             manutencoes_30_dias=0,
                             tecnicos_ativos=0,
                             proximas_manutencoes=[],
                             alertas_estoque=[])

@app.route('/veiculos')
@login_required
def veiculos():
    from empresa_helpers import is_servico, get_empresa_id
    
    empresa_id = get_empresa_id()
    veiculos_lista = []
    clientes_lista = []
    
    try:
        if Config.IS_POSTGRES:
            import psycopg2
            from psycopg2.extras import RealDictCursor
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            # Buscar veículos com JOIN para cliente (modo SERVICO)
            cursor.execute('''
                SELECT v.*, COUNT(m.id) as total_manutencoes,
                       c.nome as cliente_nome
                FROM veiculos v 
                LEFT JOIN manutencoes m ON v.id = m.veiculo_id 
                LEFT JOIN clientes c ON v.cliente_id = c.id
                WHERE v.empresa_id = %s
                GROUP BY v.id, c.nome
                ORDER BY v.placa
            ''', (empresa_id,))
            veiculos_lista = cursor.fetchall()
            
            # Buscar clientes ativos para o formulário (apenas SERVICO)
            if is_servico():
                cursor.execute('''
                    SELECT id, nome FROM clientes 
                    WHERE empresa_id = %s AND status = 'ATIVO'
                    ORDER BY nome
                ''', (empresa_id,))
                clientes_lista = cursor.fetchall()
            
            cursor.close()
            conn.close()
        else:
            conn = sqlite3.connect(DATABASE)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT v.*, COUNT(m.id) as total_manutencoes,
                       c.nome as cliente_nome
                FROM veiculos v 
                LEFT JOIN manutencoes m ON v.id = m.veiculo_id 
                LEFT JOIN clientes c ON v.cliente_id = c.id
                WHERE v.empresa_id = ?
                GROUP BY v.id
                ORDER BY v.placa
            ''', (empresa_id,))
            veiculos_lista = [dict(row) for row in cursor.fetchall()]
            
            if is_servico():
                cursor.execute('''
                    SELECT id, nome FROM clientes 
                    WHERE empresa_id = ? AND status = 'ATIVO'
                    ORDER BY nome
                ''', (empresa_id,))
                clientes_lista = [dict(row) for row in cursor.fetchall()]
            
            conn.close()
            
    except Exception as e:
        print(f"Erro ao listar veículos: {e}")
        traceback.print_exc()
        flash('Erro ao carregar lista de veículos.', 'danger')
    
    return render_template('veiculos.html', veiculos=veiculos_lista, clientes=clientes_lista)

@app.route('/api/veiculos', methods=['GET'])
@login_required
def api_veiculos():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT id, tipo, marca, modelo, placa, ano, quilometragem, proxima_manutencao, status
            FROM veiculos 
            ORDER BY placa
        ''')
        
        veiculos = []
        for row in cursor.fetchall():
            veiculos.append({
                'id': row[0],
                'tipo': row[1],
                'marca': row[2],
                'modelo': row[3],
                'placa': row[4],
                'ano': row[5],
                'quilometragem': row[6],
                'proxima_manutencao': row[7],
                'status': row[8]
            })
        
        conn.close()
        return jsonify(veiculos)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/veiculo/<int:veiculo_id>')
@login_required
def detalhes_veiculo(veiculo_id):
    from empresa_helpers import get_empresa_id
    
    empresa_id = get_empresa_id()
    
    if Config.IS_POSTGRES:
        import psycopg2
        conn = psycopg2.connect(Config.DATABASE_URL)
        cursor = conn.cursor()
        placeholder = '%s'
        date_func = "CURRENT_DATE"
    else:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        placeholder = '?'
        date_func = "date('now')"
    
    # Buscar veículo com filtro por empresa_id (isolamento de dados)
    cursor.execute(f"SELECT * FROM veiculos WHERE id = {placeholder} AND empresa_id = {placeholder}", 
                   (veiculo_id, empresa_id))
    veiculo = cursor.fetchone()
    
    # Se veículo não encontrado ou não pertence à empresa, retornar 404
    if not veiculo:
        conn.close()
        abort(404)
    
    cursor.execute(f"SELECT * FROM manutencoes WHERE veiculo_id = {placeholder} ORDER BY data_agendada DESC", 
                   (veiculo_id,))
    manutencoes = cursor.fetchall()
    
    # Busca a próxima manutenção agendada (status='Agendada' e data >= hoje)
    cursor.execute(f"""
        SELECT data_agendada 
        FROM manutencoes 
        WHERE veiculo_id = {placeholder}
        AND status = 'Agendada' 
        AND date(data_agendada) >= {date_func}
        ORDER BY data_agendada ASC 
        LIMIT 1
    """, (veiculo_id,))
    proxima_manutencao = cursor.fetchone()
    proxima_data = proxima_manutencao[0] if proxima_manutencao else None
    
    cursor.execute(f'''
        SELECT DISTINCT p.*, f.nome as fornecedor_nome 
        FROM pecas p 
        JOIN fornecedores f ON p.fornecedor_id = f.id 
        JOIN manutencao_pecas mp ON p.id = mp.peca_id
        JOIN manutencoes m ON mp.manutencao_id = m.id
        WHERE m.veiculo_id = {placeholder}
        ORDER BY p.nome
    ''', (veiculo_id,))
    pecas_compativeis = cursor.fetchall()
    
    # Busca técnicos e fornecedores para o modal de agendamento (filtrado por empresa)
    cursor.execute(f"SELECT * FROM tecnicos WHERE empresa_id = {placeholder} ORDER BY nome", 
                   (empresa_id,))
    tecnicos = cursor.fetchall()
    
    cursor.execute(f"SELECT * FROM fornecedores WHERE empresa_id = {placeholder} AND (especialidade LIKE '%serviço%' OR especialidade LIKE '%Serviço%') ORDER BY nome", 
                   (empresa_id,))
    fornecedores_servicos = cursor.fetchall()
    
    conn.close()
    
    return render_template('detalhes_veiculo.html', 
                         veiculo=veiculo, 
                         manutencoes=manutencoes, 
                         pecas_compativeis=pecas_compativeis,
                         tecnicos=tecnicos,
                         fornecedores_servicos=fornecedores_servicos,
                         proxima_manutencao=proxima_data)

@app.route('/api/veiculo', methods=['POST'])
@login_required
def criar_veiculo():
    """Criar novo veículo"""
    from empresa_helpers import is_servico, get_empresa_id, verificar_limite_veiculos
    
    try:
        data = request.json
        empresa_id = get_empresa_id()
        cliente_id = data.get('cliente_id') or None
        
        # Validação para modo SERVICO: cliente_id obrigatório
        if is_servico():
            if not cliente_id:
                return jsonify({
                    'success': False,
                    'message': 'Para empresas de serviço, é obrigatório selecionar um cliente!'
                }), 400
        else:
            # Modo FROTA: cliente_id deve ser NULL
            cliente_id = None
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            
            # Verificar limite de veículos do plano
            pode_criar, msg_limite = verificar_limite_veiculos(cursor, empresa_id)
            if not pode_criar:
                cursor.close()
                conn.close()
                # Notificar sobre bloqueio de limite (ETAPA 12)
                from empresa_helpers import create_notification
                create_notification(empresa_id, 'LIMITE_BLOQUEIO', 
                    'Limite de veículos atingido!', 
                    msg_limite, 
                    link='/minha-empresa')
                return jsonify({
                    'success': False,
                    'message': msg_limite
                }), 403
            
            # Validar placa única na empresa
            cursor.execute("SELECT id FROM veiculos WHERE placa = %s AND empresa_id = %s", 
                          (data.get('placa'), empresa_id))
            if cursor.fetchone():
                cursor.close()
                conn.close()
                return jsonify({
                    'success': False,
                    'message': 'Placa já cadastrada no sistema!'
                }), 400
            
            # Validar que cliente pertence à empresa (modo SERVICO)
            if cliente_id:
                cursor.execute("SELECT id FROM clientes WHERE id = %s AND empresa_id = %s", 
                              (cliente_id, empresa_id))
                if not cursor.fetchone():
                    cursor.close()
                    conn.close()
                    return jsonify({
                        'success': False,
                        'message': 'Cliente inválido ou não pertence à sua empresa!'
                    }), 400
            
            cursor.execute('''
                INSERT INTO veiculos (empresa_id, tipo, marca, modelo, placa, ano, quilometragem, 
                                      proxima_manutencao, status, cliente_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            ''', (
                empresa_id,
                data.get('tipo'),
                data.get('marca'), 
                data.get('modelo'),
                data.get('placa'),
                data.get('ano'),
                data.get('quilometragem', 0),
                data.get('proxima_manutencao') or None,
                'Operacional',
                cliente_id
            ))
            
            veiculo_id = cursor.fetchone()[0]
            conn.commit()
            cursor.close()
            conn.close()
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            
            # Validar placa única na empresa
            cursor.execute("SELECT id FROM veiculos WHERE placa = ? AND empresa_id = ?", 
                          (data.get('placa'), empresa_id))
            if cursor.fetchone():
                conn.close()
                return jsonify({
                    'success': False,
                    'message': 'Placa já cadastrada no sistema!'
                }), 400
            
            # Validar cliente (modo SERVICO)
            if cliente_id:
                cursor.execute("SELECT id FROM clientes WHERE id = ? AND empresa_id = ?", 
                              (cliente_id, empresa_id))
                if not cursor.fetchone():
                    conn.close()
                    return jsonify({
                        'success': False,
                        'message': 'Cliente inválido ou não pertence à sua empresa!'
                    }), 400
            
            cursor.execute('''
                INSERT INTO veiculos (empresa_id, tipo, marca, modelo, placa, ano, quilometragem, 
                                      proxima_manutencao, status, cliente_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                empresa_id,
                data.get('tipo'),
                data.get('marca'), 
                data.get('modelo'),
                data.get('placa'),
                data.get('ano'),
                data.get('quilometragem', 0),
                data.get('proxima_manutencao') or None,
                'Operacional',
                cliente_id
            ))
            
            veiculo_id = cursor.lastrowid
            conn.commit()
            conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Veículo cadastrado com sucesso!',
            'id': veiculo_id
        })
        
    except Exception as e:
        print(f"Erro ao criar veículo: {e}")
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Erro ao criar veículo: {str(e)}'
        }), 500

@app.route('/api/veiculo/<int:veiculo_id>', methods=['PUT'])
@login_required
def editar_veiculo(veiculo_id):
    """Editar veículo existente"""
    from empresa_helpers import is_servico, get_empresa_id
    
    try:
        data = request.json
        empresa_id = get_empresa_id()
        cliente_id = data.get('cliente_id') or None
        
        # Validação para modo SERVICO: cliente_id obrigatório
        if is_servico():
            if not cliente_id:
                return jsonify({
                    'success': False,
                    'message': 'Para empresas de serviço, é obrigatório selecionar um cliente!'
                }), 400
        else:
            # Modo FROTA: cliente_id deve ser NULL
            cliente_id = None
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            
            # Verificar se veículo existe e pertence à empresa
            cursor.execute("SELECT id FROM veiculos WHERE id = %s AND empresa_id = %s", 
                          (veiculo_id, empresa_id))
            if not cursor.fetchone():
                cursor.close()
                conn.close()
                return jsonify({
                    'success': False,
                    'message': 'Veículo não encontrado!'
                }), 404
            
            # Validar placa única (exceto o próprio veículo)
            cursor.execute("SELECT id FROM veiculos WHERE placa = %s AND id != %s AND empresa_id = %s", 
                          (data.get('placa'), veiculo_id, empresa_id))
            if cursor.fetchone():
                cursor.close()
                conn.close()
                return jsonify({
                    'success': False,
                    'message': 'Placa já cadastrada para outro veículo!'
                }), 400
            
            # Validar cliente (modo SERVICO)
            if cliente_id:
                cursor.execute("SELECT id FROM clientes WHERE id = %s AND empresa_id = %s", 
                              (cliente_id, empresa_id))
                if not cursor.fetchone():
                    cursor.close()
                    conn.close()
                    return jsonify({
                        'success': False,
                        'message': 'Cliente inválido ou não pertence à sua empresa!'
                    }), 400
            
            cursor.execute('''
                UPDATE veiculos 
                SET tipo=%s, marca=%s, modelo=%s, placa=%s, ano=%s, quilometragem=%s, 
                    proxima_manutencao=%s, cliente_id=%s, updated_at=CURRENT_TIMESTAMP
                WHERE id=%s AND empresa_id=%s
            ''', (
                data.get('tipo'),
                data.get('marca'), 
                data.get('modelo'),
                data.get('placa'),
                data.get('ano'),
                data.get('quilometragem', 0),
                data.get('proxima_manutencao') or None,
                cliente_id,
                veiculo_id,
                empresa_id
            ))
            
            conn.commit()
            cursor.close()
            conn.close()
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            
            # Verificar se veículo existe e pertence à empresa
            cursor.execute("SELECT id FROM veiculos WHERE id = ? AND empresa_id = ?", 
                          (veiculo_id, empresa_id))
            if not cursor.fetchone():
                conn.close()
                return jsonify({
                    'success': False,
                    'message': 'Veículo não encontrado!'
                }), 404
            
            # Validar placa única (exceto o próprio veículo)
            cursor.execute("SELECT id FROM veiculos WHERE placa = ? AND id != ? AND empresa_id = ?", 
                          (data.get('placa'), veiculo_id, empresa_id))
            if cursor.fetchone():
                conn.close()
                return jsonify({
                    'success': False,
                    'message': 'Placa já cadastrada para outro veículo!'
                }), 400
            
            # Validar cliente (modo SERVICO)
            if cliente_id:
                cursor.execute("SELECT id FROM clientes WHERE id = ? AND empresa_id = ?", 
                              (cliente_id, empresa_id))
                if not cursor.fetchone():
                    conn.close()
                    return jsonify({
                        'success': False,
                        'message': 'Cliente inválido ou não pertence à sua empresa!'
                    }), 400
            
            cursor.execute('''
                UPDATE veiculos 
                SET tipo=?, marca=?, modelo=?, placa=?, ano=?, quilometragem=?, 
                    proxima_manutencao=?, cliente_id=?
                WHERE id=? AND empresa_id=?
            ''', (
                data.get('tipo'),
                data.get('marca'), 
                data.get('modelo'),
                data.get('placa'),
                data.get('ano'),
                data.get('quilometragem', 0),
                data.get('proxima_manutencao') or None,
                cliente_id,
                veiculo_id,
                empresa_id
            ))
            
            conn.commit()
            conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Veículo atualizado com sucesso!'
        })
        
    except Exception as e:
        print(f"Erro ao editar veículo: {e}")
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Erro ao editar veículo: {str(e)}'
        }), 500

@app.route('/api/veiculo/<int:veiculo_id>', methods=['GET'])
@login_required
def obter_veiculo(veiculo_id):
    """Obter dados de um veículo específico"""
    from empresa_helpers import get_empresa_id
    
    try:
        empresa_id = get_empresa_id()
        
        if Config.IS_POSTGRES:
            import psycopg2
            from psycopg2.extras import RealDictCursor
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            cursor.execute("""
                SELECT v.*, c.nome as cliente_nome
                FROM veiculos v
                LEFT JOIN clientes c ON v.cliente_id = c.id
                WHERE v.id = %s AND v.empresa_id = %s
            """, (veiculo_id, empresa_id))
            veiculo = cursor.fetchone()
            
            cursor.close()
            conn.close()
            
            if veiculo:
                return jsonify({
                    'success': True,
                    'veiculo': {
                        'id': veiculo['id'],
                        'tipo': veiculo['tipo'],
                        'marca': veiculo['marca'],
                        'modelo': veiculo['modelo'],
                        'placa': veiculo['placa'],
                        'ano': veiculo['ano'],
                        'quilometragem': veiculo['quilometragem'],
                        'proxima_manutencao': str(veiculo['proxima_manutencao']) if veiculo['proxima_manutencao'] else None,
                        'status': veiculo['status'],
                        'cliente_id': veiculo['cliente_id'],
                        'cliente_nome': veiculo['cliente_nome']
                    }
                })
            else:
                return jsonify({
                    'success': False,
                    'message': 'Veículo não encontrado!'
                }), 404
        else:
            conn = sqlite3.connect(DATABASE)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT v.*, c.nome as cliente_nome
                FROM veiculos v
                LEFT JOIN clientes c ON v.cliente_id = c.id
                WHERE v.id = ? AND v.empresa_id = ?
            """, (veiculo_id, empresa_id))
            row = cursor.fetchone()
            conn.close()
            
            if row:
                veiculo = dict(row)
                return jsonify({
                    'success': True,
                    'veiculo': {
                        'id': veiculo['id'],
                        'tipo': veiculo['tipo'],
                        'marca': veiculo['marca'],
                        'modelo': veiculo['modelo'],
                        'placa': veiculo['placa'],
                        'ano': veiculo['ano'],
                        'quilometragem': veiculo['quilometragem'],
                        'proxima_manutencao': veiculo['proxima_manutencao'],
                        'status': veiculo['status'],
                        'cliente_id': veiculo.get('cliente_id'),
                        'cliente_nome': veiculo.get('cliente_nome')
                    }
                })
            else:
                return jsonify({
                    'success': False,
                    'message': 'Veículo não encontrado!'
                }), 404
            
    except Exception as e:
        print(f"Erro ao obter veículo: {e}")
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Erro ao obter veículo: {str(e)}'
        }), 500

@app.route('/api/veiculo/<int:veiculo_id>', methods=['DELETE'])
@login_required
def excluir_veiculo(veiculo_id):
    """Excluir um veículo"""
    from empresa_helpers import get_empresa_id
    
    try:
        empresa_id = get_empresa_id()
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            placeholder = '%s'
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            placeholder = '?'
        
        # Verificar se o veículo pertence à empresa
        cursor.execute(f"SELECT id FROM veiculos WHERE id = {placeholder} AND empresa_id = {placeholder}", 
                       (veiculo_id, empresa_id))
        if not cursor.fetchone():
            conn.close()
            return jsonify({
                'success': False,
                'message': 'Veículo não encontrado ou sem permissão!'
            }), 404
        
        # Verificar se o veículo tem manutenções associadas
        cursor.execute(f"SELECT COUNT(*) FROM manutencoes WHERE veiculo_id = {placeholder}", (veiculo_id,))
        count = cursor.fetchone()[0]
        
        if count > 0:
            conn.close()
            return jsonify({
                'success': False,
                'message': f'Não é possível excluir este veículo pois possui {count} manutenção(ões) registrada(s)!'
            }), 400
        
        # Excluir o veículo
        cursor.execute(f"DELETE FROM veiculos WHERE id = {placeholder} AND empresa_id = {placeholder}", 
                       (veiculo_id, empresa_id))
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Veículo excluído com sucesso!'
        })
        
    except Exception as e:
        print(f"Erro ao excluir veículo: {e}")
        return jsonify({
            'success': False,
            'message': 'Erro ao excluir veículo!'
        }), 500

@app.route('/manutencao')
@login_required
def manutencao():
    from empresa_helpers import get_empresa_id
    
    empresa_id = get_empresa_id()
    conn = get_db_connection()
    cursor = conn.cursor()
    
    if Config.IS_POSTGRES:
        placeholder = '%s'
    else:
        placeholder = '?'
    
    # Buscar manutenções com telefone de técnicos OU fornecedores (FILTRADO POR EMPRESA)
    cursor.execute(f'''
        SELECT 
            m.*, 
            v.placa, 
            v.modelo, 
            COALESCE(t.telefone, f.telefone) as telefone
        FROM manutencoes m 
        JOIN veiculos v ON m.veiculo_id = v.id 
        LEFT JOIN tecnicos t ON m.tecnico = t.nome AND t.empresa_id = {placeholder}
        LEFT JOIN fornecedores f ON m.tecnico = f.nome AND f.empresa_id = {placeholder}
        WHERE m.empresa_id = {placeholder}
        ORDER BY m.id DESC
    ''', (empresa_id, empresa_id, empresa_id))
    manutencoes = cursor.fetchall()
    
    # Buscar veículos (FILTRADO POR EMPRESA)
    cursor.execute(f"SELECT id, placa, modelo FROM veiculos WHERE empresa_id = {placeholder}", (empresa_id,))
    veiculos = cursor.fetchall()
    
    # Buscar técnicos (FILTRADO POR EMPRESA)
    cursor.execute(f"SELECT id, nome FROM tecnicos WHERE status='Ativo' AND empresa_id = {placeholder} ORDER BY nome", (empresa_id,))
    tecnicos = cursor.fetchall()
    
    # Buscar fornecedores que prestam serviços (FILTRADO POR EMPRESA)
    cursor.execute(f"SELECT id, nome, telefone FROM fornecedores WHERE empresa_id = {placeholder} AND (especialidade LIKE '%%serviço%%' OR especialidade LIKE '%%Serviço%%') ORDER BY nome", (empresa_id,))
    fornecedores_servicos = cursor.fetchall()
    
    conn.close()
    return render_template('manutencao.html', 
                         manutencoes=manutencoes, 
                         veiculos=veiculos, 
                         tecnicos=tecnicos,
                         fornecedores_servicos=fornecedores_servicos)

@app.route('/api/manutencao', methods=['POST'])
@login_required
def criar_manutencao():
    from empresa_helpers import is_servico, get_empresa_id
    
    data = request.json
    servicos_items = data.get('servicos', [])
    empresa_id = get_empresa_id()
    
    # Status inicial baseado no modo da empresa
    status_inicial = 'RASCUNHO' if is_servico() else 'Agendada'
    
    try:
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            
            # Inserir manutenção (inclui empresa_id para isolamento de dados)
            cursor.execute('''
                INSERT INTO manutencoes (empresa_id, veiculo_id, tipo, descricao, data_agendada, tecnico, status)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            ''', (empresa_id, data['veiculo_id'], data['tipo'], data['descricao'], data['data_agendada'], 
                  data.get('tecnico'), status_inicial))
            
            manutencao_id = cursor.fetchone()[0]
            
            # Se é empresa SERVICO, processar itens de serviço
            if is_servico() and servicos_items:
                for item in servicos_items:
                    nome_servico = item.get('nome_servico', '').strip()
                    quantidade = float(item.get('quantidade', 1))
                    valor_unitario = float(item.get('valor_unitario', 0))
                    
                    if not nome_servico:
                        continue
                    
                    # Catálogo Passivo: verificar se serviço existe, criar se não
                    cursor.execute("""
                        SELECT id FROM servicos 
                        WHERE empresa_id = %s AND LOWER(nome) = LOWER(%s)
                    """, (empresa_id, nome_servico))
                    servico_row = cursor.fetchone()
                    
                    if servico_row:
                        servico_id = servico_row[0]
                    else:
                        # Criar serviço no catálogo passivo
                        cursor.execute("""
                            INSERT INTO servicos (empresa_id, nome, preco_base, categoria, ativo)
                            VALUES (%s, %s, %s, %s, true)
                            RETURNING id
                        """, (empresa_id, nome_servico, valor_unitario, 'Geral'))
                        servico_id = cursor.fetchone()[0]
                    
                    # Inserir item de serviço na manutenção
                    cursor.execute("""
                        INSERT INTO manutencao_servicos (manutencao_id, servico_id, nome_servico, quantidade, valor_unitario)
                        VALUES (%s, %s, %s, %s, %s)
                    """, (manutencao_id, servico_id, nome_servico, quantidade, valor_unitario))
            
            conn.commit()
            cursor.close()
            conn.close()
        else:
            # SQLite (desenvolvimento local)
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            
            cursor.execute('''
                INSERT INTO manutencoes (veiculo_id, tipo, descricao, data_agendada, tecnico, status)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (data['veiculo_id'], data['tipo'], data['descricao'], data['data_agendada'], 
                  data.get('tecnico'), 'Agendada'))
            
            conn.commit()
            conn.close()
        
        return jsonify({'success': True})
    except Exception as e:
        print(f"Erro ao criar manutenção: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

# Rota para obter dados de uma manutenção específica
@app.route('/manutencao/get/<int:manutencao_id>')
@login_required
def get_manutencao(manutencao_id):
    from empresa_helpers import get_empresa_id
    
    try:
        empresa_id = get_empresa_id()
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            placeholder = '%s'
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            placeholder = '?'
        
        # Buscar manutenção com filtro por empresa_id (isolamento de dados)
        cursor.execute(f'''
            SELECT m.id, m.veiculo_id, m.tipo, m.descricao, m.data_agendada, m.data_realizada, 
                   m.custo_mao_obra, m.status, m.tecnico, v.placa, v.modelo
            FROM manutencoes m 
            JOIN veiculos v ON m.veiculo_id = v.id
            WHERE m.id = {placeholder} AND m.empresa_id = {placeholder}
        ''', (manutencao_id, empresa_id))
        manutencao = cursor.fetchone()
        
        # Buscar peças utilizadas na manutenção
        cursor.execute(f'''
            SELECT mp.id, mp.manutencao_id, mp.peca_id, mp.quantidade, mp.preco_unitario,
                   p.nome, p.codigo, (mp.quantidade * mp.preco_unitario) as subtotal
            FROM manutencao_pecas mp
            JOIN pecas p ON mp.peca_id = p.id
            WHERE mp.manutencao_id = {placeholder}
            ORDER BY mp.id DESC
        ''', (manutencao_id,))
        pecas_utilizadas = cursor.fetchall()
        
        conn.close()
        
        if manutencao:
            # Calcular total das peças
            total_pecas = sum(float(peca[7]) for peca in pecas_utilizadas) if pecas_utilizadas else 0
            
            return jsonify({
                'success': True,
                'manutencao': {
                    'id': manutencao[0],
                    'veiculo_id': manutencao[1],
                    'tipo': manutencao[2],
                    'descricao': manutencao[3],
                    'data_agendada': manutencao[4],
                    'data_realizada': manutencao[5],
                    'custo': manutencao[6],
                    'status': manutencao[7],
                    'tecnico': manutencao[8],
                    'veiculo_placa': manutencao[9],
                    'veiculo_modelo': manutencao[10]
                },
                'pecas_utilizadas': [
                    {
                        'id': peca[0],
                        'quantidade': peca[3],
                        'preco_unitario': peca[4],
                        'nome': peca[5],
                        'codigo': peca[6],
                        'subtotal': peca[7]
                    } for peca in pecas_utilizadas
                ],
                'total_pecas': total_pecas
            })
        else:
            return jsonify({'success': False, 'message': 'Manutenção não encontrada'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

# Rota para obter serviços de uma manutenção (MODO SERVICO)
@app.route('/manutencao/<int:manutencao_id>/servicos')
@login_required
def get_manutencao_servicos(manutencao_id):
    """Retorna os itens de serviço de uma manutenção - apenas para empresas SERVICO"""
    from empresa_helpers import is_servico
    
    if not is_servico():
        return jsonify({'success': False, 'message': 'Recurso não disponível'}), 403
    
    try:
        servicos_lista = []
        
        if Config.IS_POSTGRES:
            import psycopg2
            from psycopg2.extras import RealDictCursor
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            cursor.execute("""
                SELECT id, servico_id, nome_servico, quantidade, valor_unitario, subtotal
                FROM manutencao_servicos
                WHERE manutencao_id = %s
                ORDER BY id
            """, (manutencao_id,))
            servicos_lista = [dict(row) for row in cursor.fetchall()]
            
            cursor.close()
            conn.close()
        else:
            # SQLite - retornar lista vazia pois tabela pode não existir
            pass
        
        return jsonify({'success': True, 'servicos': servicos_lista})
    except Exception as e:
        print(f"Erro ao carregar serviços da manutenção: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

# =============================================
# ROTAS DE ORÇAMENTO E ORDEM DE SERVIÇO
# =============================================

@app.route('/api/manutencao/<int:manutencao_id>/gerar-orcamento', methods=['POST'])
@login_required
def gerar_orcamento(manutencao_id):
    """Gera orçamento - muda status para ORCAMENTO"""
    from empresa_helpers import is_servico, get_empresa_id
    
    if not is_servico():
        return jsonify({'success': False, 'message': 'Recurso disponível apenas para modo Serviço'}), 403
    
    try:
        empresa_id = get_empresa_id()
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            
            # Verificar se manutenção pertence à empresa
            cursor.execute("SELECT id FROM manutencoes WHERE id = %s AND empresa_id = %s", (manutencao_id, empresa_id))
            if not cursor.fetchone():
                return jsonify({'success': False, 'message': 'Manutenção não encontrada'}), 404
            
            cursor.execute("""
                UPDATE manutencoes SET status = 'ORCAMENTO', updated_at = CURRENT_TIMESTAMP
                WHERE id = %s AND empresa_id = %s
            """, (manutencao_id, empresa_id))
            
            conn.commit()
            cursor.close()
            conn.close()
        
        return jsonify({'success': True, 'message': 'Orçamento gerado com sucesso'})
    except Exception as e:
        print(f"Erro ao gerar orçamento: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/manutencao/<int:manutencao_id>/aprovar-orcamento', methods=['POST'])
@login_required
def aprovar_orcamento(manutencao_id):
    """Aprova orçamento - muda status para APROVADO e registra data/hora"""
    from empresa_helpers import is_servico, get_empresa_id
    
    if not is_servico():
        return jsonify({'success': False, 'message': 'Recurso disponível apenas para modo Serviço'}), 403
    
    try:
        empresa_id = get_empresa_id()
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            
            # Verificar se manutenção pertence à empresa e se status é ORCAMENTO
            cursor.execute("SELECT status FROM manutencoes WHERE id = %s AND empresa_id = %s", (manutencao_id, empresa_id))
            row = cursor.fetchone()
            if not row:
                return jsonify({'success': False, 'message': 'Manutenção não encontrada'}), 404
            if row[0] != 'ORCAMENTO':
                return jsonify({'success': False, 'message': 'Apenas orçamentos pendentes podem ser aprovados'}), 400
            
            # Atualizar status (com filtro empresa_id para segurança)
            cursor.execute("""
                UPDATE manutencoes SET status = 'APROVADO', updated_at = CURRENT_TIMESTAMP
                WHERE id = %s AND empresa_id = %s
            """, (manutencao_id, empresa_id))
            
            # Buscar dados para criar registro em ordens_servico
            cursor.execute("""
                SELECT m.id, v.cliente_id, 
                       COALESCE(SUM(ms.subtotal), 0) as valor_servicos
                FROM manutencoes m
                JOIN veiculos v ON m.veiculo_id = v.id
                LEFT JOIN manutencao_servicos ms ON ms.manutencao_id = m.id
                WHERE m.id = %s
                GROUP BY m.id, v.cliente_id
            """, (manutencao_id,))
            manut_data = cursor.fetchone()
            
            if manut_data and manut_data[1]:  # Se tem cliente_id
                cliente_id = manut_data[1]
                valor_servicos = manut_data[2] or 0
                
                # Gerar número da OS
                cursor.execute("""
                    SELECT COALESCE(MAX(CAST(SUBSTRING(numero_os FROM 'OS-([0-9]+)') AS INTEGER)), 0) + 1
                    FROM ordens_servico WHERE empresa_id = %s
                """, (empresa_id,))
                prox_num = cursor.fetchone()[0]
                numero_os = f"OS-{prox_num:06d}"
                
                # Verificar se já existe OS para esta manutenção
                cursor.execute("SELECT id FROM ordens_servico WHERE manutencao_id = %s", (manutencao_id,))
                if not cursor.fetchone():
                    cursor.execute("""
                        INSERT INTO ordens_servico (empresa_id, manutencao_id, cliente_id, numero_os, status, valor_servicos, data_aprovacao)
                        VALUES (%s, %s, %s, %s, 'APROVADO', %s, CURRENT_TIMESTAMP)
                    """, (empresa_id, manutencao_id, cliente_id, numero_os, valor_servicos))
            
            conn.commit()
            cursor.close()
            conn.close()
        
        return jsonify({'success': True, 'message': 'Orçamento aprovado com sucesso'})
    except Exception as e:
        print(f"Erro ao aprovar orçamento: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/manutencao/<int:manutencao_id>/iniciar-execucao', methods=['POST'])
@login_required
def iniciar_execucao(manutencao_id):
    """Inicia execução - muda status para EM_EXECUCAO"""
    from empresa_helpers import is_servico, get_empresa_id
    
    if not is_servico():
        return jsonify({'success': False, 'message': 'Recurso disponível apenas para modo Serviço'}), 403
    
    try:
        empresa_id = get_empresa_id()
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            
            # Verificar se manutenção pertence à empresa
            cursor.execute("SELECT id FROM manutencoes WHERE id = %s AND empresa_id = %s", (manutencao_id, empresa_id))
            if not cursor.fetchone():
                return jsonify({'success': False, 'message': 'Manutenção não encontrada'}), 404
            
            cursor.execute("""
                UPDATE manutencoes SET status = 'EM_EXECUCAO', updated_at = CURRENT_TIMESTAMP
                WHERE id = %s AND empresa_id = %s
            """, (manutencao_id, empresa_id))
            
            # Atualizar OS também (com filtro empresa_id)
            cursor.execute("""
                UPDATE ordens_servico SET status = 'EM_EXECUCAO', updated_at = CURRENT_TIMESTAMP
                WHERE manutencao_id = %s AND empresa_id = %s
            """, (manutencao_id, empresa_id))
            
            conn.commit()
            cursor.close()
            conn.close()
        
        return jsonify({'success': True, 'message': 'Execução iniciada'})
    except Exception as e:
        print(f"Erro ao iniciar execução: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/manutencao/<int:manutencao_id>/finalizar-servico', methods=['POST'])
@login_required
def finalizar_servico(manutencao_id):
    """Finaliza serviço - muda status para FINALIZADO e lança financeiro (ENTRADA)"""
    from empresa_helpers import is_servico, get_empresa_id
    
    if not is_servico():
        return jsonify({'success': False, 'message': 'Recurso disponível apenas para modo Serviço'}), 403
    
    try:
        empresa_id = get_empresa_id()
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            
            # Verificar se manutenção pertence à empresa
            cursor.execute("SELECT id FROM manutencoes WHERE id = %s AND empresa_id = %s", (manutencao_id, empresa_id))
            if not cursor.fetchone():
                return jsonify({'success': False, 'message': 'Manutenção não encontrada'}), 404
            
            cursor.execute("""
                UPDATE manutencoes SET status = 'FINALIZADO', data_realizada = CURRENT_DATE, updated_at = CURRENT_TIMESTAMP
                WHERE id = %s AND empresa_id = %s
            """, (manutencao_id, empresa_id))
            
            # Atualizar OS também (com filtro empresa_id)
            cursor.execute("""
                UPDATE ordens_servico SET status = 'FINALIZADO', data_conclusao = CURRENT_TIMESTAMP, updated_at = CURRENT_TIMESTAMP
                WHERE manutencao_id = %s AND empresa_id = %s
            """, (manutencao_id, empresa_id))
            
            # LANÇAMENTO FINANCEIRO AUTOMÁTICO (ENTRADA para SERVICO)
            resultado_financeiro = lancar_financeiro_manutencao(manutencao_id, cursor, empresa_id, is_servico_mode=True)
            
            conn.commit()
            cursor.close()
            conn.close()
        
        return jsonify({'success': True, 'message': 'Serviço finalizado'})
    except Exception as e:
        print(f"Erro ao finalizar serviço: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/manutencao/<int:manutencao_id>/pdf-orcamento')
@login_required
def pdf_orcamento(manutencao_id):
    """Gera PDF de Orçamento"""
    from empresa_helpers import is_servico, get_empresa_id
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from datetime import datetime
    import io
    
    if not is_servico():
        return jsonify({'success': False, 'message': 'Recurso disponível apenas para modo Serviço'}), 403
    
    try:
        empresa_id = get_empresa_id()
        dados = {}
        
        if Config.IS_POSTGRES:
            import psycopg2
            from psycopg2.extras import RealDictCursor
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            # Dados da empresa
            cursor.execute("SELECT nome, cnpj, telefone, email, endereco FROM empresas WHERE id = %s", (empresa_id,))
            dados['empresa'] = cursor.fetchone()
            
            # Dados da manutenção (com filtro empresa_id para segurança)
            cursor.execute("""
                SELECT m.id, m.descricao, m.data_agendada, m.status,
                       v.placa, v.modelo, v.marca, v.ano,
                       c.nome as cliente_nome, c.documento, c.telefone as cliente_telefone, c.email as cliente_email, c.endereco as cliente_endereco
                FROM manutencoes m
                JOIN veiculos v ON m.veiculo_id = v.id
                LEFT JOIN clientes c ON v.cliente_id = c.id
                WHERE m.id = %s AND m.empresa_id = %s
            """, (manutencao_id, empresa_id))
            dados['manutencao'] = cursor.fetchone()
            
            if not dados['manutencao']:
                return jsonify({'success': False, 'message': 'Manutenção não encontrada'}), 404
            
            # Serviços
            cursor.execute("""
                SELECT nome_servico, quantidade, valor_unitario, subtotal
                FROM manutencao_servicos
                WHERE manutencao_id = %s
            """, (manutencao_id,))
            dados['servicos'] = cursor.fetchall()
            
            cursor.close()
            conn.close()
        
        # Gerar PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=20*mm, rightMargin=20*mm, topMargin=20*mm, bottomMargin=20*mm)
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=1, spaceAfter=20)
        elements.append(Paragraph("ORÇAMENTO", title_style))
        
        # Dados da empresa
        emp = dados.get('empresa', {})
        if emp:
            elements.append(Paragraph(f"<b>{emp.get('nome', 'Empresa')}</b>", styles['Normal']))
            if emp.get('cnpj'):
                elements.append(Paragraph(f"CNPJ: {emp.get('cnpj')}", styles['Normal']))
            if emp.get('telefone'):
                elements.append(Paragraph(f"Tel: {emp.get('telefone')}", styles['Normal']))
            if emp.get('email'):
                elements.append(Paragraph(f"Email: {emp.get('email')}", styles['Normal']))
        elements.append(Spacer(1, 10))
        
        # Linha divisória
        elements.append(Paragraph("_" * 80, styles['Normal']))
        elements.append(Spacer(1, 10))
        
        # Dados do cliente
        m = dados.get('manutencao', {})
        if m:
            elements.append(Paragraph("<b>CLIENTE</b>", styles['Heading3']))
            elements.append(Paragraph(f"Nome: {m.get('cliente_nome', 'N/A')}", styles['Normal']))
            if m.get('documento'):
                elements.append(Paragraph(f"CPF/CNPJ: {m.get('documento')}", styles['Normal']))
            if m.get('cliente_telefone'):
                elements.append(Paragraph(f"Telefone: {m.get('cliente_telefone')}", styles['Normal']))
            elements.append(Spacer(1, 10))
            
            # Dados do veículo
            elements.append(Paragraph("<b>VEÍCULO</b>", styles['Heading3']))
            elements.append(Paragraph(f"Placa: {m.get('placa', 'N/A')} | Modelo: {m.get('modelo', 'N/A')} | Marca: {m.get('marca', 'N/A')} | Ano: {m.get('ano', 'N/A')}", styles['Normal']))
            elements.append(Spacer(1, 10))
            
            # Descrição
            elements.append(Paragraph("<b>DESCRIÇÃO DO SERVIÇO</b>", styles['Heading3']))
            elements.append(Paragraph(m.get('descricao', ''), styles['Normal']))
        
        elements.append(Spacer(1, 15))
        
        # Tabela de serviços
        elements.append(Paragraph("<b>SERVIÇOS</b>", styles['Heading3']))
        servicos = dados.get('servicos', [])
        if servicos:
            table_data = [['Serviço', 'Qtd', 'Valor Unit.', 'Subtotal']]
            total = 0
            for s in servicos:
                subtotal = float(s.get('subtotal', 0))
                total += subtotal
                table_data.append([
                    s.get('nome_servico', ''),
                    str(s.get('quantidade', 1)),
                    f"R$ {float(s.get('valor_unitario', 0)):.2f}",
                    f"R$ {subtotal:.2f}"
                ])
            table_data.append(['', '', 'TOTAL:', f"R$ {total:.2f}"])
            
            table = Table(table_data, colWidths=[250, 40, 80, 80])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
                ('FONTNAME', (2, -1), (-1, -1), 'Helvetica-Bold'),
            ]))
            elements.append(table)
        else:
            elements.append(Paragraph("Nenhum serviço adicionado.", styles['Normal']))
        
        elements.append(Spacer(1, 30))
        
        # Data e validade
        elements.append(Paragraph(f"Data: {datetime.now().strftime('%d/%m/%Y')}", styles['Normal']))
        elements.append(Paragraph("Validade: 15 dias", styles['Normal']))
        
        elements.append(Spacer(1, 40))
        
        # Assinatura
        elements.append(Paragraph("_" * 40, styles['Normal']))
        elements.append(Paragraph("Assinatura do Cliente", styles['Normal']))
        
        doc.build(elements)
        
        buffer.seek(0)
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'orcamento_{manutencao_id}.pdf'
        )
    except Exception as e:
        print(f"Erro ao gerar PDF orçamento: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/manutencao/<int:manutencao_id>/pdf-ordem-servico')
@login_required
def pdf_ordem_servico(manutencao_id):
    """Gera PDF de Ordem de Serviço"""
    from empresa_helpers import is_servico, get_empresa_id
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from datetime import datetime
    import io
    
    if not is_servico():
        return jsonify({'success': False, 'message': 'Recurso disponível apenas para modo Serviço'}), 403
    
    try:
        empresa_id = get_empresa_id()
        dados = {}
        
        if Config.IS_POSTGRES:
            import psycopg2
            from psycopg2.extras import RealDictCursor
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            # Dados da empresa
            cursor.execute("SELECT nome, cnpj, telefone, email, endereco FROM empresas WHERE id = %s", (empresa_id,))
            dados['empresa'] = cursor.fetchone()
            
            # Dados da OS
            cursor.execute("""
                SELECT os.numero_os, os.status as os_status, os.data_aprovacao, os.valor_servicos,
                       m.id, m.descricao, m.data_agendada, m.status, m.tecnico,
                       v.placa, v.modelo, v.marca, v.ano,
                       c.nome as cliente_nome, c.documento, c.telefone as cliente_telefone, c.email as cliente_email
                FROM ordens_servico os
                JOIN manutencoes m ON os.manutencao_id = m.id
                JOIN veiculos v ON m.veiculo_id = v.id
                LEFT JOIN clientes c ON os.cliente_id = c.id
                WHERE os.manutencao_id = %s AND os.empresa_id = %s
            """, (manutencao_id, empresa_id))
            dados['os'] = cursor.fetchone()
            
            # Serviços
            cursor.execute("""
                SELECT nome_servico, quantidade, valor_unitario, subtotal
                FROM manutencao_servicos
                WHERE manutencao_id = %s
            """, (manutencao_id,))
            dados['servicos'] = cursor.fetchall()
            
            cursor.close()
            conn.close()
        
        if not dados.get('os'):
            return jsonify({'success': False, 'message': 'Ordem de serviço não encontrada. Aprove o orçamento primeiro.'}), 404
        
        # Gerar PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=20*mm, rightMargin=20*mm, topMargin=20*mm, bottomMargin=20*mm)
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=1, spaceAfter=10)
        elements.append(Paragraph("ORDEM DE SERVIÇO", title_style))
        
        os_data = dados.get('os', {})
        elements.append(Paragraph(f"<b>Nº {os_data.get('numero_os', 'N/A')}</b>", ParagraphStyle('Center', parent=styles['Normal'], alignment=1)))
        elements.append(Spacer(1, 15))
        
        # Dados da empresa
        emp = dados.get('empresa', {})
        if emp:
            elements.append(Paragraph(f"<b>{emp.get('nome', 'Empresa')}</b>", styles['Normal']))
            if emp.get('cnpj'):
                elements.append(Paragraph(f"CNPJ: {emp.get('cnpj')}", styles['Normal']))
            if emp.get('telefone'):
                elements.append(Paragraph(f"Tel: {emp.get('telefone')}", styles['Normal']))
        elements.append(Spacer(1, 10))
        
        elements.append(Paragraph("_" * 80, styles['Normal']))
        elements.append(Spacer(1, 10))
        
        # Dados do cliente
        if os_data:
            elements.append(Paragraph("<b>CLIENTE</b>", styles['Heading3']))
            elements.append(Paragraph(f"Nome: {os_data.get('cliente_nome', 'N/A')}", styles['Normal']))
            if os_data.get('documento'):
                elements.append(Paragraph(f"CPF/CNPJ: {os_data.get('documento')}", styles['Normal']))
            if os_data.get('cliente_telefone'):
                elements.append(Paragraph(f"Telefone: {os_data.get('cliente_telefone')}", styles['Normal']))
            elements.append(Spacer(1, 10))
            
            # Dados do veículo
            elements.append(Paragraph("<b>VEÍCULO</b>", styles['Heading3']))
            elements.append(Paragraph(f"Placa: {os_data.get('placa', 'N/A')} | Modelo: {os_data.get('modelo', 'N/A')} | Marca: {os_data.get('marca', 'N/A')} | Ano: {os_data.get('ano', 'N/A')}", styles['Normal']))
            elements.append(Spacer(1, 10))
            
            # Status e Técnico
            elements.append(Paragraph("<b>INFORMAÇÕES DO SERVIÇO</b>", styles['Heading3']))
            elements.append(Paragraph(f"Status: {os_data.get('os_status', 'N/A')}", styles['Normal']))
            if os_data.get('tecnico'):
                elements.append(Paragraph(f"Técnico Responsável: {os_data.get('tecnico')}", styles['Normal']))
            if os_data.get('data_aprovacao'):
                data_aprov = os_data.get('data_aprovacao')
                if hasattr(data_aprov, 'strftime'):
                    data_aprov = data_aprov.strftime('%d/%m/%Y %H:%M')
                elements.append(Paragraph(f"Data de Aprovação: {data_aprov}", styles['Normal']))
            elements.append(Spacer(1, 10))
            
            # Descrição
            elements.append(Paragraph("<b>DESCRIÇÃO</b>", styles['Heading3']))
            elements.append(Paragraph(os_data.get('descricao', ''), styles['Normal']))
        
        elements.append(Spacer(1, 15))
        
        # Tabela de serviços
        elements.append(Paragraph("<b>SERVIÇOS A EXECUTAR</b>", styles['Heading3']))
        servicos = dados.get('servicos', [])
        if servicos:
            table_data = [['Serviço', 'Qtd', 'Valor Unit.', 'Subtotal']]
            total = 0
            for s in servicos:
                subtotal = float(s.get('subtotal', 0))
                total += subtotal
                table_data.append([
                    s.get('nome_servico', ''),
                    str(s.get('quantidade', 1)),
                    f"R$ {float(s.get('valor_unitario', 0)):.2f}",
                    f"R$ {subtotal:.2f}"
                ])
            table_data.append(['', '', 'TOTAL:', f"R$ {total:.2f}"])
            
            table = Table(table_data, colWidths=[250, 40, 80, 80])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#198754')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
                ('FONTNAME', (2, -1), (-1, -1), 'Helvetica-Bold'),
            ]))
            elements.append(table)
        
        elements.append(Spacer(1, 30))
        
        # Data de emissão
        elements.append(Paragraph(f"Data de Emissão: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
        
        elements.append(Spacer(1, 40))
        
        # Assinaturas
        sig_table = Table([
            ['_' * 30, '_' * 30],
            ['Cliente', 'Responsável Técnico']
        ], colWidths=[200, 200])
        sig_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]))
        elements.append(sig_table)
        
        doc.build(elements)
        
        buffer.seek(0)
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'ordem_servico_{manutencao_id}.pdf'
        )
    except Exception as e:
        print(f"Erro ao gerar PDF OS: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

# Rota para atualizar status da manutenção
@app.route('/manutencao/status/<int:manutencao_id>', methods=['PUT'])
@login_required
def atualizar_status_manutencao(manutencao_id):
    try:
        data = request.json
        novo_status = data.get('status')
        
        if not novo_status:
            return jsonify({'success': False, 'message': 'Status não fornecido'})
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            UPDATE manutencoes 
            SET status = ?
            WHERE id = ?
        ''', (novo_status, manutencao_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Status atualizado com sucesso'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

# Rota para editar manutenção
@app.route('/manutencao/edit/<int:manutencao_id>', methods=['POST'])
@login_required
def edit_manutencao(manutencao_id):
    from empresa_helpers import is_servico, get_empresa_id
    
    try:
        # Aceitar JSON ou FormData
        if request.is_json:
            data = request.json
            veiculo_id = data.get('veiculo_id')
            tipo = data.get('tipo')
            descricao = data.get('descricao')
            data_agendada = data.get('data_agendada')
            data_realizada = data.get('data_realizada') or None
            custo = data.get('custo')
            status = data.get('status')
            tecnico = data.get('tecnico')
            servicos_items = data.get('servicos', [])
        else:
            veiculo_id = request.form['veiculo_id']
            tipo = request.form['tipo']
            descricao = request.form['descricao']
            data_agendada = request.form['data_agendada']
            data_realizada = request.form.get('data_realizada', None)
            custo = request.form.get('custo', None)
            status = request.form['status']
            tecnico = request.form['tecnico']
            servicos_items = []
        
        # Converter campos vazios para None
        if data_realizada == '':
            data_realizada = None
        if custo == '' or custo is None:
            custo = None
        else:
            custo = float(custo) if custo else None
        
        empresa_id = get_empresa_id()
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            
            # Buscar status anterior e verificar se pertence à empresa
            cursor.execute("SELECT status, financeiro_lancado_em FROM manutencoes WHERE id = %s AND empresa_id = %s", (manutencao_id, empresa_id))
            status_anterior_row = cursor.fetchone()
            if not status_anterior_row:
                return jsonify({'success': False, 'message': 'Manutenção não encontrada'}), 404
            status_anterior = status_anterior_row[0] if status_anterior_row else None
            ja_lancado = status_anterior_row[1] if status_anterior_row else None
            
            # Atualizar custo_total se informado
            custo_update = custo if custo is not None else None
            
            # Atualizar manutenção (com filtro empresa_id para segurança)
            cursor.execute('''
                UPDATE manutencoes 
                SET veiculo_id=%s, tipo=%s, descricao=%s, data_agendada=%s, data_realizada=%s, 
                    custo_total=%s, status=%s, tecnico=%s, updated_at=CURRENT_TIMESTAMP
                WHERE id=%s AND empresa_id=%s
            ''', (veiculo_id, tipo, descricao, data_agendada, data_realizada, custo_update, status, tecnico, manutencao_id, empresa_id))
            
            # Se é empresa SERVICO, processar itens de serviço
            if is_servico():
                # Remover itens anteriores
                cursor.execute("DELETE FROM manutencao_servicos WHERE manutencao_id = %s", (manutencao_id,))
                
                # Inserir novos itens
                for item in servicos_items:
                    nome_servico = item.get('nome_servico', '').strip()
                    quantidade = float(item.get('quantidade', 1))
                    valor_unitario = float(item.get('valor_unitario', 0))
                    
                    if not nome_servico:
                        continue
                    
                    # Catálogo Passivo: verificar se serviço existe, criar se não
                    cursor.execute("""
                        SELECT id FROM servicos 
                        WHERE empresa_id = %s AND LOWER(nome) = LOWER(%s)
                    """, (empresa_id, nome_servico))
                    servico_row = cursor.fetchone()
                    
                    if servico_row:
                        servico_id = servico_row[0]
                    else:
                        # Criar serviço no catálogo passivo
                        cursor.execute("""
                            INSERT INTO servicos (empresa_id, nome, preco_base, categoria, ativo)
                            VALUES (%s, %s, %s, %s, true)
                            RETURNING id
                        """, (empresa_id, nome_servico, valor_unitario, 'Geral'))
                        servico_id = cursor.fetchone()[0]
                    
                    # Inserir item de serviço na manutenção
                    cursor.execute("""
                        INSERT INTO manutencao_servicos (manutencao_id, servico_id, nome_servico, quantidade, valor_unitario)
                        VALUES (%s, %s, %s, %s, %s)
                    """, (manutencao_id, servico_id, nome_servico, quantidade, valor_unitario))
            
            # LANÇAMENTO FINANCEIRO AUTOMÁTICO (quando status muda para finalizado)
            status_finalizacao_servico = ['FINALIZADO', 'FATURADO']
            status_finalizacao_frota = ['Concluída']
            
            deve_lancar = False
            if is_servico() and status in status_finalizacao_servico and not ja_lancado:
                deve_lancar = True
                is_servico_mode = True
            elif not is_servico() and status in status_finalizacao_frota and not ja_lancado:
                deve_lancar = True
                is_servico_mode = False
            
            if deve_lancar:
                resultado_financeiro = lancar_financeiro_manutencao(manutencao_id, cursor, empresa_id, is_servico_mode)
                print(f"   Resultado financeiro: {resultado_financeiro}")
            
            conn.commit()
            cursor.close()
            conn.close()
        else:
            # SQLite (desenvolvimento local)
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            
            # Verificar se a manutenção foi concluída e tem custo
            cursor.execute('SELECT status, custo FROM manutencoes WHERE id = ?', (manutencao_id,))
            manutencao_atual = cursor.fetchone()
            
            cursor.execute('''
                UPDATE manutencoes 
                SET veiculo_id=?, tipo=?, descricao=?, data_agendada=?, data_realizada=?, custo=?, status=?, tecnico=?
                WHERE id=?
            ''', (veiculo_id, tipo, descricao, data_agendada, data_realizada, custo, status, tecnico, manutencao_id))
            
            # Gerar despesa automática se manutenção foi concluída com custo
            if status == 'Concluída' and custo and custo > 0:
                # Verificar se já existe despesa para esta manutenção
                cursor.execute('SELECT id FROM despesas WHERE manutencao_id = ?', (manutencao_id,))
                despesa_existente = cursor.fetchone()
                
                if not despesa_existente:
                    # Buscar categoria "Manutenção"
                    cursor.execute('SELECT id FROM categorias_despesa WHERE nome = ?', ('Manutenção',))
                    categoria_manutencao = cursor.fetchone()
                    categoria_id = categoria_manutencao[0] if categoria_manutencao else None
                    
                    # Criar despesa automática
                    data_despesa = data_realizada if data_realizada else data_agendada
                    descricao_despesa = f"Manutenção {tipo}: {descricao}"
                    
                    cursor.execute('''
                        INSERT INTO despesas (descricao, valor, data_despesa, categoria_id, veiculo_id, tipo, manutencao_id)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (descricao_despesa, custo, data_despesa, categoria_id, veiculo_id, 'Automática', manutencao_id))
            
            conn.commit()
            conn.close()
        
        return jsonify({'success': True, 'message': 'Manutenção atualizada com sucesso!'})
    except Exception as e:
        print(f"Erro ao editar manutenção: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})

# Rota para deletar manutenção
@app.route('/manutencao/delete/<int:manutencao_id>', methods=['DELETE'])
@login_required
def delete_manutencao(manutencao_id):
    from empresa_helpers import get_empresa_id
    
    try:
        empresa_id = get_empresa_id()
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            cursor.execute('DELETE FROM manutencoes WHERE id = %s AND empresa_id = %s', 
                           (manutencao_id, empresa_id))
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            cursor.execute('DELETE FROM manutencoes WHERE id = ? AND empresa_id = ?', 
                           (manutencao_id, empresa_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Manutenção excluída com sucesso!'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

# Rota para atualizar status da manutenção
@app.route('/manutencao/status/<int:manutencao_id>', methods=['POST'])
@login_required
def update_status_manutencao(manutencao_id):
    from empresa_helpers import get_empresa_id
    
    try:
        empresa_id = get_empresa_id()
        novo_status = request.form['status']
        data_realizada = None
        
        # Se está finalizando, definir data de realização
        if novo_status == 'Concluída':
            from datetime import datetime
            data_realizada = datetime.now().strftime('%Y-%m-%d')
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            cursor.execute('''
                UPDATE manutencoes 
                SET status = %s, data_realizada = %s
                WHERE id = %s AND empresa_id = %s
            ''', (novo_status, data_realizada, manutencao_id, empresa_id))
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            cursor.execute('''
                UPDATE manutencoes 
                SET status = ?, data_realizada = ?
                WHERE id = ? AND empresa_id = ?
            ''', (novo_status, data_realizada, manutencao_id, empresa_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': f'Status atualizado para {novo_status}!'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

# Rotas para gerenciar peças nas manutenções
@app.route('/manutencao/<int:manutencao_id>/pecas')
@login_required
def get_pecas_manutencao(manutencao_id):
    if Config.IS_POSTGRES:
        import psycopg2
        from psycopg2.extras import RealDictCursor
        conn = psycopg2.connect(Config.DATABASE_URL)
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute('''
            SELECT mp.id, mp.manutencao_id, mp.peca_id, mp.quantidade, mp.preco_unitario, 
                   p.nome, p.codigo, p.preco,
                   (mp.quantidade * mp.preco_unitario) as subtotal
            FROM manutencao_pecas mp
            JOIN pecas p ON mp.peca_id = p.id
            WHERE mp.manutencao_id = %s
            ORDER BY mp.id DESC
        ''', (manutencao_id,))
        pecas_utilizadas = [dict(row) for row in cursor.fetchall()]
    else:
        conn = sqlite3.connect(DATABASE)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute('''
            SELECT mp.id, mp.manutencao_id, mp.peca_id, mp.quantidade, mp.preco_unitario, 
                   p.nome, p.codigo, p.preco,
                   (mp.quantidade * mp.preco_unitario) as subtotal
            FROM manutencao_pecas mp
            JOIN pecas p ON mp.peca_id = p.id
            WHERE mp.manutencao_id = ?
            ORDER BY mp.id DESC
        ''', (manutencao_id,))
        pecas_utilizadas = [dict(row) for row in cursor.fetchall()]
    
    conn.close()
    return jsonify(pecas_utilizadas)

@app.route('/manutencao/<int:manutencao_id>/pecas/disponiveis')
@login_required
def get_pecas_disponiveis(manutencao_id):
    from empresa_helpers import get_empresa_id
    
    empresa_id = get_empresa_id()
    
    if Config.IS_POSTGRES:
        import psycopg2
        conn = psycopg2.connect(Config.DATABASE_URL)
        cursor = conn.cursor()
        placeholder = '%s'
    else:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        placeholder = '?'
    
    # Buscar informações do veículo da manutenção
    cursor.execute(f'''
        SELECT v.modelo, v.marca, v.tipo
        FROM manutencoes m
        JOIN veiculos v ON m.veiculo_id = v.id
        WHERE m.id = {placeholder}
    ''', (manutencao_id,))
    veiculo = cursor.fetchone()
    
    # Buscar TODAS as peças com estoque disponível (filtrado por empresa)
    # Priorizar peças compatíveis, mas mostrar todas
    if veiculo:
        marca = veiculo[1].lower()
        modelo = veiculo[0].lower()
        tipo = veiculo[2].lower()
        
        cursor.execute(f'''
            SELECT p.id, p.nome, p.codigo, p.veiculo_compativel, p.preco, p.fornecedor_id, p.quantidade_estoque, f.nome as fornecedor_nome,
                   CASE 
                       WHEN LOWER(p.veiculo_compativel) = 'universal' THEN 1
                       WHEN LOWER(p.veiculo_compativel) LIKE {placeholder} THEN 2
                       WHEN LOWER(p.veiculo_compativel) LIKE {placeholder} THEN 3
                       WHEN LOWER(p.veiculo_compativel) LIKE {placeholder} THEN 4
                       ELSE 5
                   END as prioridade
            FROM pecas p
            JOIN fornecedores f ON p.fornecedor_id = f.id
            WHERE p.quantidade_estoque > 0 AND p.empresa_id = {placeholder}
            ORDER BY prioridade, p.nome
        ''', (f'%{marca}%', f'%{modelo}%', f'%{tipo}%', empresa_id))
    else:
        # Se não conseguir identificar veículo, mostrar todas com estoque
        cursor.execute(f'''
            SELECT p.id, p.nome, p.codigo, p.veiculo_compativel, p.preco, p.fornecedor_id, p.quantidade_estoque, f.nome as fornecedor_nome, 1 as prioridade
            FROM pecas p
            JOIN fornecedores f ON p.fornecedor_id = f.id
            WHERE p.quantidade_estoque > 0 AND p.empresa_id = {placeholder}
            ORDER BY p.nome
        ''', (empresa_id,))
    
    pecas_disponiveis = cursor.fetchall()
    conn.close()
    
    # Remover a coluna de prioridade do resultado (última coluna)
    pecas_clean = [peca[:-1] for peca in pecas_disponiveis]
    
    return jsonify(pecas_clean)

@app.route('/manutencao/<int:manutencao_id>/pecas/buscar')
@login_required
def buscar_pecas_manutencao(manutencao_id):
    """Buscar peças por nome ou código para adicionar em manutenção"""
    from empresa_helpers import get_empresa_id
    
    search_term = request.args.get('q', '').strip()
    empresa_id = get_empresa_id()
    
    if Config.IS_POSTGRES:
        import psycopg2
        conn = psycopg2.connect(Config.DATABASE_URL)
        cursor = conn.cursor()
        placeholder = '%s'
    else:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        placeholder = '?'
    
    if search_term:
        # Buscar por nome, código ou compatibilidade (filtrado por empresa)
        cursor.execute(f'''
            SELECT p.id, p.nome, p.codigo, p.veiculo_compativel, p.preco, p.fornecedor_id, p.quantidade_estoque, f.nome as fornecedor_nome
            FROM pecas p
            JOIN fornecedores f ON p.fornecedor_id = f.id
            WHERE p.quantidade_estoque > 0
            AND p.empresa_id = {placeholder}
            AND (
                LOWER(p.nome) LIKE {placeholder} OR 
                LOWER(p.codigo) LIKE {placeholder} OR
                LOWER(p.veiculo_compativel) LIKE {placeholder}
            )
            ORDER BY 
                CASE 
                    WHEN LOWER(p.nome) LIKE {placeholder} THEN 1
                    WHEN LOWER(p.codigo) LIKE {placeholder} THEN 2
                    ELSE 3
                END,
                p.nome
        ''', (empresa_id, f'%{search_term.lower()}%', f'%{search_term.lower()}%', f'%{search_term.lower()}%',
              f'%{search_term.lower()}%', f'%{search_term.lower()}%'))
    else:
        # Se não há termo de busca, retornar todas com estoque
        cursor.execute(f'''
            SELECT p.id, p.nome, p.codigo, p.veiculo_compativel, p.preco, p.fornecedor_id, p.quantidade_estoque, f.nome as fornecedor_nome
            FROM pecas p
            JOIN fornecedores f ON p.fornecedor_id = f.id
            WHERE p.quantidade_estoque > 0 AND p.empresa_id = {placeholder}
            ORDER BY p.nome
        ''', (empresa_id,))
    
    pecas = cursor.fetchall()
    conn.close()
    
    return jsonify(pecas)

@app.route('/manutencao/<int:manutencao_id>/pecas/adicionar', methods=['POST'])
@login_required
def adicionar_peca_manutencao(manutencao_id):
    from empresa_helpers import get_empresa_id
    
    conn = None
    try:
        empresa_id = get_empresa_id()
        data = request.json
        peca_id = data['peca_id']
        quantidade = int(data['quantidade'])
        
        print(f"DEBUG: Adicionando peça {peca_id} (qtd: {quantidade}) à manutenção {manutencao_id}")
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            placeholder = '%s'
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            placeholder = '?'
        
        # Verificar se a manutenção pertence à empresa
        cursor.execute(f'SELECT id FROM manutencoes WHERE id = {placeholder} AND empresa_id = {placeholder}', (manutencao_id, empresa_id))
        if not cursor.fetchone():
            return jsonify({'success': False, 'message': 'Manutenção não encontrada'}), 404
        
        # Verificar se há estoque suficiente (e se a peça pertence à empresa)
        cursor.execute(f'SELECT quantidade_estoque, preco FROM pecas WHERE id = {placeholder} AND empresa_id = {placeholder}', (peca_id, empresa_id))
        peca = cursor.fetchone()
        
        if not peca:
            return jsonify({'success': False, 'message': 'Peça não encontrada!'})
        
        estoque_atual, preco = peca
        if estoque_atual < quantidade:
            return jsonify({'success': False, 'message': f'Estoque insuficiente! Disponível: {estoque_atual}'})
        
        # Adicionar peça à manutenção
        print(f"DEBUG: Inserindo na tabela manutencao_pecas")
        print(f"DEBUG: Valores - manutencao_id={manutencao_id}, peca_id={peca_id}, quantidade={quantidade}, preco={preco}")
        cursor.execute(f'''
            INSERT INTO manutencao_pecas (manutencao_id, peca_id, quantidade, preco_unitario)
            VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder})
        ''', (manutencao_id, peca_id, quantidade, preco))
        print(f"DEBUG: INSERT bem sucedido!")
        
        # Dar baixa no estoque
        print(f"DEBUG: Atualizando estoque")
        cursor.execute(f'''
            UPDATE pecas SET quantidade_estoque = quantidade_estoque - {placeholder}
            WHERE id = {placeholder}
        ''', (quantidade, peca_id))
        
        try:
            # Gerar despesa automática para a peça (opcional - não bloqueia se falhar)
            cursor.execute(f'SELECT nome FROM pecas WHERE id = {placeholder}', (peca_id,))
            nome_peca = cursor.fetchone()[0]
            
            cursor.execute(f'SELECT veiculo_id, data_realizada, data_agendada FROM manutencoes WHERE id = {placeholder}', (manutencao_id,))
            manutencao_info = cursor.fetchone()
            veiculo_id, data_realizada, data_agendada = manutencao_info
            
            # Buscar categoria "Peças e Componentes"
            cursor.execute(f'SELECT id FROM categorias_despesa WHERE nome = {placeholder}', ('Peças e Componentes',))
            categoria_peca = cursor.fetchone()
            categoria_id = categoria_peca[0] if categoria_peca else None
            
            # Criar despesa automática para a peça
            valor_total = preco * quantidade
            data_despesa = data_realizada if data_realizada else data_agendada
            descricao_despesa = f"Peça: {nome_peca} (Qtd: {quantidade})"
            
            cursor.execute(f'''
                INSERT INTO despesas (descricao, valor, data_despesa, categoria_id, veiculo_id, tipo, manutencao_id)
                VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder})
            ''', (descricao_despesa, valor_total, data_despesa, categoria_id, veiculo_id, 'Automática', manutencao_id))
            print(f"DEBUG: Despesa automática criada")
        except Exception as despesa_error:
            print(f"DEBUG: Erro ao criar despesa (não crítico): {despesa_error}")
            # Continua mesmo se falhar ao criar despesa
        
        print(f"DEBUG: Fazendo commit")
        conn.commit()
        print(f"DEBUG: Peça adicionada com sucesso!")
        
        return jsonify({'success': True, 'message': f'Peça adicionada com sucesso! Estoque atualizado.'})
    except Exception as e:
        print(f"DEBUG ERRO: {str(e)}")
        if conn:
            conn.rollback()
        return jsonify({'success': False, 'message': str(e)})
    finally:
        if conn:
            conn.close()

@app.route('/manutencao/<int:manutencao_id>/pecas/<int:item_id>/remover', methods=['DELETE'])
@login_required
def remover_peca_manutencao(manutencao_id, item_id):
    from empresa_helpers import get_empresa_id
    
    try:
        empresa_id = get_empresa_id()
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            placeholder = '%s'
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            placeholder = '?'
        
        # Verificar se a manutenção pertence à empresa
        cursor.execute(f'SELECT id FROM manutencoes WHERE id = {placeholder} AND empresa_id = {placeholder}', (manutencao_id, empresa_id))
        if not cursor.fetchone():
            return jsonify({'success': False, 'message': 'Manutenção não encontrada'}), 404
        
        # Buscar informações da peça para devolver ao estoque
        cursor.execute(f'''
            SELECT peca_id, quantidade
            FROM manutencao_pecas
            WHERE id = {placeholder} AND manutencao_id = {placeholder}
        ''', (item_id, manutencao_id))
        
        item = cursor.fetchone()
        if not item:
            return jsonify({'success': False, 'message': 'Item não encontrado!'})
        
        peca_id, quantidade = item
        
        # Devolver ao estoque
        cursor.execute(f'''
            UPDATE pecas SET quantidade_estoque = quantidade_estoque + {placeholder}
            WHERE id = {placeholder}
        ''', (quantidade, peca_id))
        
        # Remover da manutenção
        cursor.execute(f'''
            DELETE FROM manutencao_pecas
            WHERE id = {placeholder} AND manutencao_id = {placeholder}
        ''', (item_id, manutencao_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Peça removida e estoque atualizado!'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/pecas')
@login_required
def pecas():
    from empresa_helpers import get_empresa_id
    
    try:
        empresa_id = get_empresa_id()
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            placeholder = '%s'
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            placeholder = '?'
        
        # SELECT com colunas na ordem esperada pelo template:
        # [0]=id, [1]=nome, [2]=codigo, [3]=veiculo_compativel, [4]=preco, 
        # [5]=fornecedor_id, [6]=quantidade_estoque, [7]=fornecedor_nome,
        # [8]=fornecedor_telefone, [9]=fornecedor_email, [10]=categoria_id,
        # [11]=categoria_nome, [12]=categoria_cor, [13]=categoria_icone
        cursor.execute(f'''
            SELECT p.id, p.nome, p.codigo, p.veiculo_compativel, p.preco, 
                   p.fornecedor_id, p.quantidade_estoque, f.nome as fornecedor_nome,
                   f.telefone as fornecedor_telefone, f.email as fornecedor_email,
                   p.categoria_id, c.nome as categoria_nome, c.cor as categoria_cor, c.icone as categoria_icone
            FROM pecas p 
            LEFT JOIN fornecedores f ON p.fornecedor_id = f.id
            LEFT JOIN categorias_pecas c ON p.categoria_id = c.id
            WHERE p.empresa_id = {placeholder}
            ORDER BY p.nome
        ''', (empresa_id,))
        pecas_list = cursor.fetchall()
        
        # Buscar fornecedores para os selects (da empresa)
        cursor.execute(f"SELECT id, nome FROM fornecedores WHERE empresa_id = {placeholder}", (empresa_id,))
        fornecedores = cursor.fetchall()
        
        # Buscar categorias para os selects (da empresa)
        ativo_val = 'true' if Config.IS_POSTGRES else '1'
        cursor.execute(f"SELECT id, nome, cor, icone FROM categorias_pecas WHERE empresa_id = {placeholder} AND ativo = {ativo_val} ORDER BY nome", (empresa_id,))
        categorias = cursor.fetchall()
        
        # Calcular valor total do estoque
        cursor.execute(f"SELECT COALESCE(SUM(preco * quantidade_estoque), 0) FROM pecas WHERE empresa_id = {placeholder}", (empresa_id,))
        valor_total_estoque = cursor.fetchone()[0] or 0
        
        conn.close()
        return render_template('pecas.html', 
                             pecas=pecas_list, 
                             fornecedores=fornecedores, 
                             categorias=categorias,
                             valor_total_estoque=valor_total_estoque)
    except Exception as e:
        print(f"ERRO na rota /pecas: {e}")
        import traceback
        traceback.print_exc()
        return render_template('pecas.html', pecas=[], fornecedores=[], categorias=[], valor_total_estoque=0)

# =============================================
# ROTAS - CATEGORIAS DE PEÇAS
# =============================================

@app.route('/api/categorias-pecas', methods=['GET'])
@login_required
def listar_categorias_pecas():
    """Listar categorias de peças da empresa"""
    from empresa_helpers import get_empresa_id
    
    try:
        empresa_id = get_empresa_id()
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            cursor.execute('''
                SELECT id, nome, descricao, cor, icone, ativo 
                FROM categorias_pecas 
                WHERE empresa_id = %s 
                ORDER BY nome
            ''', (empresa_id,))
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            cursor.execute('''
                SELECT id, nome, descricao, cor, icone, ativo 
                FROM categorias_pecas 
                WHERE empresa_id = ? 
                ORDER BY nome
            ''', (empresa_id,))
        
        categorias = cursor.fetchall()
        conn.close()
        
        return jsonify({
            'success': True,
            'categorias': [
                {'id': c[0], 'nome': c[1], 'descricao': c[2], 'cor': c[3], 'icone': c[4], 'ativo': c[5]}
                for c in categorias
            ]
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/categorias-pecas', methods=['POST'])
@csrf.exempt
@login_required
def criar_categoria_peca():
    """Criar nova categoria de peça"""
    from empresa_helpers import get_empresa_id
    
    try:
        empresa_id = get_empresa_id()
        data = request.get_json(force=True)
        
        nome = data.get('nome', '').strip()
        if not nome:
            return jsonify({'success': False, 'message': 'Nome da categoria é obrigatório'})
        
        descricao = data.get('descricao', '')
        cor = data.get('cor', '#6c757d')
        icone = data.get('icone', 'fas fa-tag')
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            
            # Verificar se já existe
            cursor.execute('SELECT id FROM categorias_pecas WHERE empresa_id = %s AND nome = %s', (empresa_id, nome))
            if cursor.fetchone():
                conn.close()
                return jsonify({'success': False, 'message': f'Já existe uma categoria "{nome}"'})
            
            cursor.execute('''
                INSERT INTO categorias_pecas (empresa_id, nome, descricao, cor, icone)
                VALUES (%s, %s, %s, %s, %s) RETURNING id
            ''', (empresa_id, nome, descricao, cor, icone))
            categoria_id = cursor.fetchone()[0]
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            
            cursor.execute('SELECT id FROM categorias_pecas WHERE empresa_id = ? AND nome = ?', (empresa_id, nome))
            if cursor.fetchone():
                conn.close()
                return jsonify({'success': False, 'message': f'Já existe uma categoria "{nome}"'})
            
            cursor.execute('''
                INSERT INTO categorias_pecas (empresa_id, nome, descricao, cor, icone)
                VALUES (?, ?, ?, ?, ?)
            ''', (empresa_id, nome, descricao, cor, icone))
            categoria_id = cursor.lastrowid
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True, 
            'message': 'Categoria criada com sucesso!',
            'categoria': {'id': categoria_id, 'nome': nome, 'cor': cor, 'icone': icone}
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/categorias-pecas/<int:categoria_id>', methods=['PUT'])
@csrf.exempt
@login_required
def editar_categoria_peca(categoria_id):
    """Editar categoria de peça"""
    from empresa_helpers import get_empresa_id
    
    try:
        empresa_id = get_empresa_id()
        data = request.get_json(force=True)
        
        nome = data.get('nome', '').strip()
        descricao = data.get('descricao', '')
        cor = data.get('cor', '#6c757d')
        icone = data.get('icone', 'fas fa-tag')
        ativo = data.get('ativo', True)
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            cursor.execute('''
                UPDATE categorias_pecas 
                SET nome = %s, descricao = %s, cor = %s, icone = %s, ativo = %s, updated_at = CURRENT_TIMESTAMP
                WHERE id = %s AND empresa_id = %s
            ''', (nome, descricao, cor, icone, ativo, categoria_id, empresa_id))
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            cursor.execute('''
                UPDATE categorias_pecas 
                SET nome = ?, descricao = ?, cor = ?, icone = ?, ativo = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ? AND empresa_id = ?
            ''', (nome, descricao, cor, icone, ativo, categoria_id, empresa_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Categoria atualizada!'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/categorias-pecas/<int:categoria_id>', methods=['DELETE'])
@csrf.exempt
@login_required
def excluir_categoria_peca(categoria_id):
    """Excluir categoria de peça"""
    from empresa_helpers import get_empresa_id
    
    try:
        empresa_id = get_empresa_id()
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            # Remover referências nas peças
            cursor.execute('UPDATE pecas SET categoria_id = NULL WHERE categoria_id = %s AND empresa_id = %s', (categoria_id, empresa_id))
            # Excluir categoria
            cursor.execute('DELETE FROM categorias_pecas WHERE id = %s AND empresa_id = %s', (categoria_id, empresa_id))
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            cursor.execute('UPDATE pecas SET categoria_id = NULL WHERE categoria_id = ? AND empresa_id = ?', (categoria_id, empresa_id))
            cursor.execute('DELETE FROM categorias_pecas WHERE id = ? AND empresa_id = ?', (categoria_id, empresa_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Categoria excluída!'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

# Rota para adicionar nova peça
@app.route('/pecas/add', methods=['POST'])
@csrf.exempt
@login_required
def add_peca():
    from empresa_helpers import get_empresa_id
    
    try:
        empresa_id = get_empresa_id()
        nome = request.form['nome']
        codigo = request.form['codigo']
        veiculo_compativel = request.form.get('veiculo_compativel', '')
        preco = float(request.form['preco'])
        quantidade_estoque = int(request.form['quantidade_estoque'])
        fornecedor_id = request.form.get('fornecedor_id')
        fornecedor_id = int(fornecedor_id) if fornecedor_id else None
        categoria_id = request.form.get('categoria_id')
        categoria_id = int(categoria_id) if categoria_id else None
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            
            # Verificar se código já existe para esta empresa
            cursor.execute('SELECT id FROM pecas WHERE codigo = %s AND empresa_id = %s', (codigo, empresa_id))
            if cursor.fetchone():
                conn.close()
                return jsonify({'success': False, 'message': f'Já existe uma peça com o código "{codigo}". Use outro código.'})
            
            cursor.execute('''
                INSERT INTO pecas (empresa_id, nome, codigo, veiculo_compativel, preco, quantidade_estoque, fornecedor_id, categoria_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            ''', (empresa_id, nome, codigo, veiculo_compativel, preco, quantidade_estoque, fornecedor_id, categoria_id))
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            
            # Verificar se código já existe para esta empresa
            cursor.execute('SELECT id FROM pecas WHERE codigo = ? AND empresa_id = ?', (codigo, empresa_id))
            if cursor.fetchone():
                conn.close()
                return jsonify({'success': False, 'message': f'Já existe uma peça com o código "{codigo}". Use outro código.'})
            
            cursor.execute('''
                INSERT INTO pecas (empresa_id, nome, codigo, veiculo_compativel, preco, quantidade_estoque, fornecedor_id, categoria_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (empresa_id, nome, codigo, veiculo_compativel, preco, quantidade_estoque, fornecedor_id, categoria_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Peça adicionada com sucesso!'})
    except Exception as e:
        error_msg = str(e)
        if 'unique_codigo_empresa' in error_msg or 'UNIQUE constraint' in error_msg:
            return jsonify({'success': False, 'message': f'Já existe uma peça com este código. Use outro código.'})
        return jsonify({'success': False, 'message': error_msg})

# Rota para editar peça
@app.route('/pecas/edit/<int:peca_id>', methods=['POST'])
@csrf.exempt
@login_required
def edit_peca(peca_id):
    from empresa_helpers import get_empresa_id
    
    try:
        empresa_id = get_empresa_id()
        nome = request.form['nome']
        codigo = request.form['codigo']
        veiculo_compativel = request.form.get('veiculo_compativel', '')
        preco = float(request.form['preco'])
        quantidade_estoque = int(request.form['quantidade_estoque'])
        fornecedor_id = request.form.get('fornecedor_id')
        fornecedor_id = int(fornecedor_id) if fornecedor_id else None
        categoria_id = request.form.get('categoria_id')
        categoria_id = int(categoria_id) if categoria_id else None
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            cursor.execute('''
                UPDATE pecas 
                SET nome=%s, codigo=%s, veiculo_compativel=%s, preco=%s, quantidade_estoque=%s, fornecedor_id=%s, categoria_id=%s
                WHERE id=%s AND empresa_id=%s
            ''', (nome, codigo, veiculo_compativel, preco, quantidade_estoque, fornecedor_id, categoria_id, peca_id, empresa_id))
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            cursor.execute('''
                UPDATE pecas 
                SET nome=?, codigo=?, veiculo_compativel=?, preco=?, quantidade_estoque=?, fornecedor_id=?, categoria_id=?
                WHERE id=? AND empresa_id=?
            ''', (nome, codigo, veiculo_compativel, preco, quantidade_estoque, fornecedor_id, categoria_id, peca_id, empresa_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Peça atualizada com sucesso!'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

# Rota para deletar peça
@app.route('/pecas/delete/<int:peca_id>', methods=['DELETE'])
@csrf.exempt
@login_required
def delete_peca(peca_id):
    from empresa_helpers import get_empresa_id
    
    try:
        empresa_id = get_empresa_id()
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            cursor.execute('DELETE FROM pecas WHERE id=%s AND empresa_id=%s', (peca_id, empresa_id))
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            cursor.execute('DELETE FROM pecas WHERE id=? AND empresa_id=?', (peca_id, empresa_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Peça excluída com sucesso!'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

# Rota para ajustar estoque
@app.route('/pecas/estoque/<int:peca_id>', methods=['POST'])
@csrf.exempt
@login_required
def ajustar_estoque(peca_id):
    from empresa_helpers import get_empresa_id
    
    try:
        empresa_id = get_empresa_id()
        nova_quantidade = int(request.form['quantidade'])
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            cursor.execute('UPDATE pecas SET quantidade_estoque=%s WHERE id=%s AND empresa_id=%s', (nova_quantidade, peca_id, empresa_id))
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            cursor.execute('UPDATE pecas SET quantidade_estoque=? WHERE id=? AND empresa_id=?', (nova_quantidade, peca_id, empresa_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Estoque atualizado com sucesso!'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

# Rota para obter dados de uma peça específica
@app.route('/pecas/get/<int:peca_id>')
@login_required
def get_peca(peca_id):
    from empresa_helpers import get_empresa_id
    
    try:
        empresa_id = get_empresa_id()
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            cursor.execute('''
                SELECT p.*, f.nome as fornecedor_nome 
                FROM pecas p 
                JOIN fornecedores f ON p.fornecedor_id = f.id
                WHERE p.id=%s AND p.empresa_id=%s
            ''', (peca_id, empresa_id))
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            cursor.execute('''
                SELECT p.*, f.nome as fornecedor_nome 
                FROM pecas p 
                JOIN fornecedores f ON p.fornecedor_id = f.id
                WHERE p.id=? AND p.empresa_id=?
            ''', (peca_id, empresa_id))
        peca = cursor.fetchone()
        conn.close()
        
        if peca:
            return jsonify({
                'success': True,
                'peca': {
                    'id': peca[0],
                    'nome': peca[1],
                    'codigo': peca[2],
                    'veiculo_compativel': peca[3],
                    'preco': peca[4],
                    'fornecedor_id': peca[5],
                    'quantidade_estoque': peca[6],
                    'fornecedor_nome': peca[7],
                    'descricao': ''  # Campo não existe na tabela atual
                }
            })
        else:
            return jsonify({'success': False, 'message': 'Peça não encontrada'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

# =============================================
# ROTAS - GESTÃO DE EMPRESAS (MULTI-EMPRESA)
# =============================================

@app.route('/empresas')
@login_required
@super_admin_required
def empresas():
    """Página de gestão de empresas - Apenas Super Admin"""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM empresas ORDER BY nome")
    empresas = cursor.fetchall()
    conn.close()
    return render_template('empresas.html', empresas=empresas)

@app.route('/api/empresas', methods=['POST'])
@login_required
@super_admin_required
def criar_empresa():
    """Criar nova empresa - Apenas Super Admin"""
    try:
        data = request.json
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO empresas (
                nome, nome_fantasia, cnpj, telefone, email,
                endereco, cidade, estado, cep, plano,
                limite_veiculos, limite_usuarios, ativo
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
        ''', (
            data.get('nome'),
            data.get('nome_fantasia'),
            data.get('cnpj'),
            data.get('telefone'),
            data.get('email'),
            data.get('endereco'),
            data.get('cidade'),
            data.get('estado'),
            data.get('cep'),
            data.get('plano', 'Básico'),
            data.get('limite_veiculos', 10),
            data.get('limite_usuarios', 3)
        ))
        
        empresa_id = cursor.lastrowid
        conn.commit()
        conn.close()
        
        log_action(f'Criou empresa: {data.get("nome")} (ID: {empresa_id})')
        
        return jsonify({
            'success': True,
            'message': 'Empresa criada com sucesso!',
            'empresa_id': empresa_id
        })
        
    except Exception as e:
        print(f"Erro ao criar empresa: {e}")
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@app.route('/api/empresas/<int:empresa_id>', methods=['PUT'])
@login_required
@super_admin_required
def atualizar_empresa(empresa_id):
    """Atualizar empresa - Apenas Super Admin"""
    try:
        data = request.json
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Atualizar apenas o campo ativo ou todos os campos
        if 'ativo' in data and len(data) == 1:
            cursor.execute('''
                UPDATE empresas SET ativo = ?
                WHERE id = ?
            ''', (data['ativo'], empresa_id))
        else:
            cursor.execute('''
                UPDATE empresas SET
                    nome = ?, nome_fantasia = ?, cnpj = ?,
                    telefone = ?, email = ?, endereco = ?,
                    cidade = ?, estado = ?, cep = ?,
                    plano = ?, limite_veiculos = ?, limite_usuarios = ?
                WHERE id = ?
            ''', (
                data.get('nome'),
                data.get('nome_fantasia'),
                data.get('cnpj'),
                data.get('telefone'),
                data.get('email'),
                data.get('endereco'),
                data.get('cidade'),
                data.get('estado'),
                data.get('cep'),
                data.get('plano'),
                data.get('limite_veiculos'),
                data.get('limite_usuarios'),
                empresa_id
            ))
        
        conn.commit()
        conn.close()
        
        log_action(f'Atualizou empresa ID: {empresa_id}')
        
        return jsonify({
            'success': True,
            'message': 'Empresa atualizada com sucesso!'
        })
        
    except Exception as e:
        print(f"Erro ao atualizar empresa: {e}")
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

# =============================================
# ROTAS - FORNECEDORES
# =============================================

@app.route('/fornecedores')
@login_required
def fornecedores():
    from empresa_helpers import get_empresa_id
    
    empresa_id = get_empresa_id()
    print(f"DEBUG /fornecedores: empresa_id={empresa_id}")
    
    if Config.IS_POSTGRES:
        import psycopg2
        conn = psycopg2.connect(Config.DATABASE_URL)
        cursor = conn.cursor()  # Tuplas para compatibilidade com template
        cursor.execute("SELECT id, nome, contato, telefone, email, especialidade FROM fornecedores WHERE empresa_id = %s", (empresa_id,))
    else:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        cursor.execute("SELECT id, nome, contato, telefone, email, especialidade FROM fornecedores WHERE empresa_id = ?", (empresa_id,))
    
    fornecedores_list = cursor.fetchall()
    print(f"DEBUG /fornecedores: resultado={fornecedores_list}")
    conn.close()
    return render_template('fornecedores.html', fornecedores=fornecedores_list)

@app.route('/fornecedores/criar', methods=['POST'])
@csrf.exempt
@login_required
def criar_fornecedor():
    """Criar novo fornecedor"""
    from empresa_helpers import get_empresa_id
    
    try:
        print(f"DEBUG: Content-Type: {request.content_type}")
        print(f"DEBUG: Request data: {request.data}")
        
        empresa_id = get_empresa_id()
        data = request.get_json(force=True)  # Force parse JSON
        
        print(f"DEBUG: Dados recebidos: {data}")
        
        if not data or not data.get('nome'):
            return jsonify({
                'success': False,
                'message': 'Nome é obrigatório'
            }), 400
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            
            cursor.execute('''
                INSERT INTO fornecedores (empresa_id, nome, contato, telefone, email, especialidade, cnpj, endereco)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            ''', (
                empresa_id,
                data.get('nome'),
                data.get('contato'),
                data.get('telefone'),
                data.get('email'),
                data.get('especialidade'),
                data.get('cnpj', ''),
                data.get('endereco', '')
            ))
            
            fornecedor_id = cursor.fetchone()[0]
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            
            cursor.execute('''
                INSERT INTO fornecedores (empresa_id, nome, contato, telefone, email, especialidade, cnpj, endereco)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                empresa_id,
                data.get('nome'),
                data.get('contato'),
                data.get('telefone'),
                data.get('email'),
                data.get('especialidade'),
                data.get('cnpj', ''),
                data.get('endereco', '')
            ))
            
            fornecedor_id = cursor.lastrowid
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Fornecedor cadastrado com sucesso!',
            'id': fornecedor_id
        })
        
    except Exception as e:
        print(f"Erro ao criar fornecedor: {e}")
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Erro ao criar fornecedor: {str(e)}'
        }), 500

@app.route('/fornecedores/detalhes/<int:fornecedor_id>', methods=['GET'])
@login_required
def detalhes_fornecedor(fornecedor_id):
    from empresa_helpers import get_empresa_id
    
    try:
        empresa_id = get_empresa_id()
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id, nome, contato, telefone, email, especialidade
                FROM fornecedores 
                WHERE id = %s AND empresa_id = %s
            """, (fornecedor_id, empresa_id))
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id, nome, contato, telefone, email, especialidade
                FROM fornecedores 
                WHERE id = ? AND empresa_id = ?
            """, (fornecedor_id, empresa_id))
        
        fornecedor_data = cursor.fetchone()
        conn.close()
        
        if fornecedor_data:
            fornecedor = {
                'id': fornecedor_data[0],
                'nome': fornecedor_data[1],
                'contato': fornecedor_data[2],
                'telefone': fornecedor_data[3],
                'email': fornecedor_data[4],
                'especialidade': fornecedor_data[5],
                'cnpj': 'Não informado',
                'inscricao_estadual': 'Não informado',
                'endereco': 'Não informado',
                'site': 'Não informado',
                'prazo_pagamento': 'Não informado',
                'observacoes': 'Nenhuma observação cadastrada'
            }
            return jsonify(fornecedor)
        else:
            return jsonify({'error': 'Fornecedor não encontrado'}), 404
            
    except Exception as e:
        print(f"Erro ao buscar fornecedor: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/fornecedores/editar/<int:fornecedor_id>', methods=['PUT'])
@login_required
def editar_fornecedor(fornecedor_id):
    from empresa_helpers import get_empresa_id
    
    try:
        empresa_id = get_empresa_id()
        dados = request.get_json()
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            
            cursor.execute("""
                UPDATE fornecedores 
                SET nome = %s, contato = %s, telefone = %s, email = %s, especialidade = %s
                WHERE id = %s AND empresa_id = %s
            """, (
                dados.get('nome'),
                dados.get('contato'),
                dados.get('telefone'),
                dados.get('email'),
                dados.get('especialidade'),
                fornecedor_id,
                empresa_id
            ))
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            
            cursor.execute("""
                UPDATE fornecedores 
                SET nome = ?, contato = ?, telefone = ?, email = ?, especialidade = ?
                WHERE id = ? AND empresa_id = ?
            """, (
                dados.get('nome'),
                dados.get('contato'),
                dados.get('telefone'),
                dados.get('email'),
                dados.get('especialidade'),
                fornecedor_id,
                empresa_id
            ))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Fornecedor atualizado com sucesso'})
        
    except Exception as e:
        print(f"Erro ao editar fornecedor: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/fornecedores/excluir/<int:fornecedor_id>', methods=['DELETE'])
@login_required
def excluir_fornecedor(fornecedor_id):
    from empresa_helpers import get_empresa_id
    
    try:
        empresa_id = get_empresa_id()
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            
            # Verifica se o fornecedor existe e pertence à empresa
            cursor.execute("SELECT nome FROM fornecedores WHERE id = %s AND empresa_id = %s", (fornecedor_id, empresa_id))
            fornecedor = cursor.fetchone()
            
            if not fornecedor:
                conn.close()
                return jsonify({'success': False, 'error': 'Fornecedor não encontrado'}), 404
            
            # Exclui o fornecedor
            cursor.execute("DELETE FROM fornecedores WHERE id = %s AND empresa_id = %s", (fornecedor_id, empresa_id))
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            
            # Verifica se o fornecedor existe e pertence à empresa
            cursor.execute("SELECT nome FROM fornecedores WHERE id = ? AND empresa_id = ?", (fornecedor_id, empresa_id))
            fornecedor = cursor.fetchone()
            
            if not fornecedor:
                conn.close()
                return jsonify({'success': False, 'error': 'Fornecedor não encontrado'}), 404
            
            # Exclui o fornecedor
            cursor.execute("DELETE FROM fornecedores WHERE id = ? AND empresa_id = ?", (fornecedor_id, empresa_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Fornecedor excluído com sucesso'})
        
    except Exception as e:
        print(f"Erro ao excluir fornecedor: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

# =============================================
# ROTAS DE CLIENTES (MODO SERVICO)
# =============================================

@app.route('/clientes')
@login_required
def clientes():
    """
    Listagem de clientes - APENAS para empresas SERVICO
    """
    from empresa_helpers import is_servico, get_empresa_id
    
    if not is_servico():
        flash('Este recurso está disponível apenas para empresas de Prestação de Serviços.', 'warning')
        return redirect(url_for('dashboard'))
    
    empresa_id = get_empresa_id()
    clientes_lista = []
    novos_mes = 0
    
    try:
        if Config.IS_POSTGRES:
            import psycopg2
            from psycopg2.extras import RealDictCursor
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            cursor.execute("""
                SELECT id, nome, documento, tipo_documento, telefone, email, 
                       endereco, cidade, estado, cep, observacoes, status, created_at
                FROM clientes 
                WHERE empresa_id = %s
                ORDER BY nome
            """, (empresa_id,))
            clientes_lista = cursor.fetchall()
            
            # Contar novos clientes do mês
            cursor.execute("""
                SELECT COUNT(*) as total FROM clientes 
                WHERE empresa_id = %s 
                AND created_at >= date_trunc('month', CURRENT_DATE)
            """, (empresa_id,))
            novos_mes = cursor.fetchone()['total']
            
            cursor.close()
            conn.close()
        else:
            conn = sqlite3.connect(DATABASE)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT id, nome, documento, tipo_documento, telefone, email, 
                       endereco, cidade, estado, cep, observacoes, status, created_at
                FROM clientes 
                WHERE empresa_id = ?
                ORDER BY nome
            """, (empresa_id,))
            clientes_lista = [dict(row) for row in cursor.fetchall()]
            
            cursor.execute("""
                SELECT COUNT(*) as total FROM clientes 
                WHERE empresa_id = ? 
                AND created_at >= date('now', 'start of month')
            """, (empresa_id,))
            novos_mes = cursor.fetchone()['total']
            
            conn.close()
            
    except Exception as e:
        print(f"Erro ao listar clientes: {e}")
        traceback.print_exc()
        flash('Erro ao carregar lista de clientes.', 'danger')
    
    return render_template('clientes.html', clientes=clientes_lista, novos_mes=novos_mes)


@app.route('/clientes/criar', methods=['POST'])
@login_required
def criar_cliente():
    """Criar novo cliente"""
    from empresa_helpers import is_servico, get_empresa_id, verificar_limite_clientes
    
    if not is_servico():
        return jsonify({'success': False, 'message': 'Recurso não disponível para sua empresa'}), 403
    
    try:
        data = request.json
        empresa_id = get_empresa_id()
        
        # Validação
        if not data.get('nome') or not data.get('nome').strip():
            return jsonify({'success': False, 'message': 'Nome do cliente é obrigatório'}), 400
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            
            # Verificar limite de clientes do plano
            pode_criar, msg_limite = verificar_limite_clientes(cursor, empresa_id)
            if not pode_criar:
                cursor.close()
                conn.close()
                # Notificar sobre bloqueio de limite (ETAPA 12)
                from empresa_helpers import create_notification
                create_notification(empresa_id, 'LIMITE_BLOQUEIO', 
                    'Limite de clientes atingido!', 
                    msg_limite, 
                    link='/minha-empresa')
                return jsonify({'success': False, 'message': msg_limite}), 403
            
            cursor.execute("""
                INSERT INTO clientes (empresa_id, nome, documento, tipo_documento, telefone, email, 
                                      endereco, cidade, estado, cep, observacoes, status)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'ATIVO')
                RETURNING id
            """, (
                empresa_id,
                data.get('nome', '').strip(),
                data.get('documento', '').strip() or None,
                data.get('tipo_documento', '').strip() or None,
                data.get('telefone', '').strip() or None,
                data.get('email', '').strip() or None,
                data.get('endereco', '').strip() or None,
                data.get('cidade', '').strip() or None,
                data.get('estado', '').strip() or None,
                data.get('cep', '').strip() or None,
                data.get('observacoes', '').strip() or None
            ))
            
            cliente_id = cursor.fetchone()[0]
            conn.commit()
            cursor.close()
            conn.close()
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            
            cursor.execute("""
                INSERT INTO clientes (empresa_id, nome, documento, tipo_documento, telefone, email, 
                                      endereco, cidade, estado, cep, observacoes, status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'ATIVO')
            """, (
                empresa_id,
                data.get('nome', '').strip(),
                data.get('documento', '').strip() or None,
                data.get('tipo_documento', '').strip() or None,
                data.get('telefone', '').strip() or None,
                data.get('email', '').strip() or None,
                data.get('endereco', '').strip() or None,
                data.get('cidade', '').strip() or None,
                data.get('estado', '').strip() or None,
                data.get('cep', '').strip() or None,
                data.get('observacoes', '').strip() or None
            ))
            
            cliente_id = cursor.lastrowid
            conn.commit()
            conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Cliente cadastrado com sucesso!',
            'id': cliente_id
        })
        
    except Exception as e:
        print(f"Erro ao criar cliente: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Erro ao criar cliente: {str(e)}'}), 500


@app.route('/clientes/detalhes/<int:cliente_id>', methods=['GET'])
@login_required
def detalhes_cliente(cliente_id):
    """Obter detalhes de um cliente"""
    from empresa_helpers import is_servico, get_empresa_id
    
    if not is_servico():
        return jsonify({'error': 'Recurso não disponível para sua empresa'}), 403
    
    try:
        empresa_id = get_empresa_id()
        
        if Config.IS_POSTGRES:
            import psycopg2
            from psycopg2.extras import RealDictCursor
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            cursor.execute("""
                SELECT id, nome, documento, tipo_documento, telefone, email, 
                       endereco, cidade, estado, cep, observacoes, status, created_at
                FROM clientes 
                WHERE id = %s AND empresa_id = %s
            """, (cliente_id, empresa_id))
            cliente = cursor.fetchone()
            
            cursor.close()
            conn.close()
        else:
            conn = sqlite3.connect(DATABASE)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT id, nome, documento, tipo_documento, telefone, email, 
                       endereco, cidade, estado, cep, observacoes, status, created_at
                FROM clientes 
                WHERE id = ? AND empresa_id = ?
            """, (cliente_id, empresa_id))
            row = cursor.fetchone()
            cliente = dict(row) if row else None
            
            conn.close()
        
        if not cliente:
            return jsonify({'error': 'Cliente não encontrado'}), 404
        
        # Converter datetime para string se necessário
        if cliente.get('created_at'):
            cliente['created_at'] = str(cliente['created_at'])
            
        return jsonify(cliente)
        
    except Exception as e:
        print(f"Erro ao buscar cliente: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/clientes/editar/<int:cliente_id>', methods=['PUT'])
@login_required
def editar_cliente(cliente_id):
    """Editar cliente existente"""
    from empresa_helpers import is_servico, get_empresa_id
    
    if not is_servico():
        return jsonify({'success': False, 'message': 'Recurso não disponível para sua empresa'}), 403
    
    try:
        data = request.json
        empresa_id = get_empresa_id()
        
        # Validação
        if not data.get('nome') or not data.get('nome').strip():
            return jsonify({'success': False, 'message': 'Nome do cliente é obrigatório'}), 400
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            
            # Verificar se pertence à empresa
            cursor.execute("SELECT id FROM clientes WHERE id = %s AND empresa_id = %s", (cliente_id, empresa_id))
            if not cursor.fetchone():
                cursor.close()
                conn.close()
                return jsonify({'success': False, 'message': 'Cliente não encontrado'}), 404
            
            cursor.execute("""
                UPDATE clientes SET
                    nome = %s,
                    documento = %s,
                    tipo_documento = %s,
                    telefone = %s,
                    email = %s,
                    endereco = %s,
                    cidade = %s,
                    estado = %s,
                    cep = %s,
                    observacoes = %s,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = %s AND empresa_id = %s
            """, (
                data.get('nome', '').strip(),
                data.get('documento', '').strip() or None,
                data.get('tipo_documento', '').strip() or None,
                data.get('telefone', '').strip() or None,
                data.get('email', '').strip() or None,
                data.get('endereco', '').strip() or None,
                data.get('cidade', '').strip() or None,
                data.get('estado', '').strip() or None,
                data.get('cep', '').strip() or None,
                data.get('observacoes', '').strip() or None,
                cliente_id,
                empresa_id
            ))
            
            conn.commit()
            cursor.close()
            conn.close()
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            
            # Verificar se pertence à empresa
            cursor.execute("SELECT id FROM clientes WHERE id = ? AND empresa_id = ?", (cliente_id, empresa_id))
            if not cursor.fetchone():
                conn.close()
                return jsonify({'success': False, 'message': 'Cliente não encontrado'}), 404
            
            cursor.execute("""
                UPDATE clientes SET
                    nome = ?,
                    documento = ?,
                    tipo_documento = ?,
                    telefone = ?,
                    email = ?,
                    endereco = ?,
                    cidade = ?,
                    estado = ?,
                    cep = ?,
                    observacoes = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ? AND empresa_id = ?
            """, (
                data.get('nome', '').strip(),
                data.get('documento', '').strip() or None,
                data.get('tipo_documento', '').strip() or None,
                data.get('telefone', '').strip() or None,
                data.get('email', '').strip() or None,
                data.get('endereco', '').strip() or None,
                data.get('cidade', '').strip() or None,
                data.get('estado', '').strip() or None,
                data.get('cep', '').strip() or None,
                data.get('observacoes', '').strip() or None,
                cliente_id,
                empresa_id
            ))
            
            conn.commit()
            conn.close()
        
        return jsonify({'success': True, 'message': 'Cliente atualizado com sucesso!'})
        
    except Exception as e:
        print(f"Erro ao editar cliente: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Erro ao editar cliente: {str(e)}'}), 500


@app.route('/clientes/toggle-status/<int:cliente_id>', methods=['PUT'])
@login_required
def toggle_status_cliente(cliente_id):
    """Ativar/Inativar cliente"""
    from empresa_helpers import is_servico, get_empresa_id
    
    if not is_servico():
        return jsonify({'success': False, 'message': 'Recurso não disponível para sua empresa'}), 403
    
    try:
        data = request.json
        empresa_id = get_empresa_id()
        novo_status = data.get('status', 'INATIVO')
        
        if novo_status not in ['ATIVO', 'INATIVO']:
            return jsonify({'success': False, 'message': 'Status inválido'}), 400
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            
            cursor.execute("""
                UPDATE clientes SET status = %s, updated_at = CURRENT_TIMESTAMP
                WHERE id = %s AND empresa_id = %s
            """, (novo_status, cliente_id, empresa_id))
            
            if cursor.rowcount == 0:
                cursor.close()
                conn.close()
                return jsonify({'success': False, 'message': 'Cliente não encontrado'}), 404
            
            conn.commit()
            cursor.close()
            conn.close()
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            
            cursor.execute("""
                UPDATE clientes SET status = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ? AND empresa_id = ?
            """, (novo_status, cliente_id, empresa_id))
            
            if cursor.rowcount == 0:
                conn.close()
                return jsonify({'success': False, 'message': 'Cliente não encontrado'}), 404
            
            conn.commit()
            conn.close()
        
        acao = 'ativado' if novo_status == 'ATIVO' else 'inativado'
        return jsonify({'success': True, 'message': f'Cliente {acao} com sucesso!'})
        
    except Exception as e:
        print(f"Erro ao alterar status do cliente: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'}), 500


# =============================================
# ROTAS DE TÉCNICOS
# =============================================

@app.route('/tecnicos')
@login_required
def tecnicos():
    from empresa_helpers import get_empresa_id
    
    empresa_id = get_empresa_id()
    
    if Config.IS_POSTGRES:
        import psycopg2
        from psycopg2.extras import RealDictCursor
        conn = psycopg2.connect(Config.DATABASE_URL)
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("SELECT * FROM tecnicos WHERE empresa_id = %s ORDER BY nome", (empresa_id,))
    else:
        conn = sqlite3.connect(DATABASE)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM tecnicos WHERE empresa_id = ? ORDER BY nome", (empresa_id,))
    
    tecnicos = cursor.fetchall()
    conn.close()
    return render_template('tecnicos.html', tecnicos=tecnicos)

@app.route('/api/tecnico/<int:tecnico_id>', methods=['GET'])
@login_required
def buscar_tecnico(tecnico_id):
    from empresa_helpers import get_empresa_id
    
    try:
        empresa_id = get_empresa_id()
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM tecnicos WHERE id=%s AND empresa_id=%s", (tecnico_id, empresa_id))
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM tecnicos WHERE id=? AND empresa_id=?", (tecnico_id, empresa_id))
        
        tecnico = cursor.fetchone()
        conn.close()
        
        if tecnico:
            return jsonify({
                'id': tecnico[0],
                'nome': tecnico[1],
                'cpf': tecnico[2],
                'telefone': tecnico[3],
                'email': tecnico[4],
                'especialidade': tecnico[5],
                'data_admissao': tecnico[6],
                'status': tecnico[7]
            })
        else:
            return jsonify({'success': False, 'message': 'Técnico não encontrado'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/tecnico/<int:tecnico_id>/historico', methods=['GET'])
@login_required
def historico_tecnico(tecnico_id):
    from empresa_helpers import get_empresa_id
    
    try:
        empresa_id = get_empresa_id()
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            
            # Busca manutenções realizadas pelo técnico (filtrado por empresa)
            cursor.execute("""
                SELECT COALESCE(m.data_realizada, m.data_agendada) as data, 
                       v.placa, v.modelo, m.tipo, m.status
                FROM manutencoes m
                LEFT JOIN veiculos v ON m.veiculo_id = v.id
                WHERE m.tecnico = (SELECT nome FROM tecnicos WHERE id = %s AND empresa_id = %s)
                AND m.empresa_id = %s
                ORDER BY data DESC
                LIMIT 10
            """, (tecnico_id, empresa_id, empresa_id))
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            
            # Busca manutenções realizadas pelo técnico (filtrado por empresa)
            cursor.execute("""
                SELECT COALESCE(m.data_realizada, m.data_agendada) as data, 
                       v.placa, v.modelo, m.tipo, m.status
                FROM manutencoes m
                LEFT JOIN veiculos v ON m.veiculo_id = v.id
                WHERE m.tecnico = (SELECT nome FROM tecnicos WHERE id = ? AND empresa_id = ?)
                AND m.empresa_id = ?
                ORDER BY data DESC
                LIMIT 10
            """, (tecnico_id, empresa_id, empresa_id))
        
        manutencoes = cursor.fetchall()
        conn.close()
        
        historico = []
        for m in manutencoes:
            historico.append({
                'data': m[0] if m[0] else 'Não informado',
                'veiculo': f"{m[1]} - {m[2]}" if m[1] else 'Não informado',
                'tipo': m[3],
                'status': m[4]
            })
        
        return jsonify(historico)
        
    except Exception as e:
        print(f"Erro ao buscar histórico: {e}")
        traceback.print_exc()
        return jsonify([]), 500

@app.route('/api/manutencoes/calendario', methods=['GET'])
@login_required
def get_manutencoes_calendario():
    """Retorna manutenções formatadas para o FullCalendar"""
    from empresa_helpers import get_empresa_id
    
    try:
        empresa_id = get_empresa_id()
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            
            # Buscar manutenções da empresa com informações do veículo
            cursor.execute('''
                SELECT 
                    m.id,
                    m.tipo,
                    m.status,
                    m.data_agendada,
                    m.data_realizada,
                    v.placa,
                    v.modelo
                FROM manutencoes m
                LEFT JOIN veiculos v ON m.veiculo_id = v.id
                WHERE m.empresa_id = %s
                ORDER BY m.data_agendada DESC
            ''', (empresa_id,))
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            
            # Buscar manutenções da empresa com informações do veículo
            cursor.execute('''
                SELECT 
                    m.id,
                    m.tipo,
                    m.status,
                    m.data_agendada,
                    m.data_realizada,
                    v.placa,
                    v.modelo
                FROM manutencoes m
                LEFT JOIN veiculos v ON m.veiculo_id = v.id
                WHERE m.empresa_id = ?
                ORDER BY m.data_agendada DESC
            ''', (empresa_id,))
        
        manutencoes = cursor.fetchall()
        conn.close()
        
        # Mapear cores por status
        cores = {
            'Agendada': '#0d6efd',      # Azul (Bootstrap primary)
            'Em Andamento': '#ffc107',  # Amarelo (Bootstrap warning)
            'Concluída': '#198754',     # Verde (Bootstrap success)
            'Cancelada': '#dc3545'      # Vermelho (Bootstrap danger)
        }
        
        # Formatar eventos para o FullCalendar
        eventos = []
        for m in manutencoes:
            # Usar data_realizada se existir, senão data_agendada
            data = m[4] if m[4] else m[3]
            
            if data:  # Apenas se houver data
                eventos.append({
                    'id': m[0],
                    'title': f"{m[5]} - {m[1]}",  # Placa - Tipo
                    'start': data,
                    'backgroundColor': cores.get(m[2], '#6c757d'),
                    'borderColor': cores.get(m[2], '#6c757d'),
                    'extendedProps': {
                        'veiculo': f"{m[5]} - {m[6]}",  # Placa - Modelo
                        'tipo': m[1],
                        'status': m[2]
                    }
                })
        
        return jsonify(eventos)
        
    except Exception as e:
        print(f"Erro ao buscar manutenções para calendário: {e}")
        traceback.print_exc()
        return jsonify([]), 500

@app.route('/api/tecnico', methods=['POST'])
@login_required
def criar_tecnico():
    from empresa_helpers import get_empresa_id
    
    conn = None
    try:
        print("DEBUG: Recebendo requisição POST /api/tecnico")
        
        empresa_id = get_empresa_id()
        data = request.json
        
        # Tratar campos vazios como NULL para PostgreSQL
        data_admissao = data.get('data_admissao', '') or None
        cpf = data.get('cpf', '') or None
        email = data.get('email', '') or None
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            
            cursor.execute('''
                INSERT INTO tecnicos (empresa_id, nome, cpf, telefone, email, especialidade, data_admissao, status)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            ''', (
                empresa_id,
                data['nome'],
                cpf,
                data.get('telefone', ''),
                email,
                data.get('especialidade', ''),
                data_admissao,
                data.get('status', 'Ativo')
            ))
            
            tecnico_id = cursor.fetchone()[0]
        else:
            conn = sqlite3.connect(DATABASE, timeout=30.0)
            cursor = conn.cursor()
            
            cursor.execute('''
                INSERT INTO tecnicos (empresa_id, nome, cpf, telefone, email, especialidade, data_admissao, status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                empresa_id,
                data['nome'],
                cpf,
                data.get('telefone', ''),
                email,
                data.get('especialidade', ''),
                data_admissao,
                data.get('status', 'Ativo')
            ))
            
            tecnico_id = cursor.lastrowid
        
        conn.commit()
        conn.close()
        
        print(f"DEBUG: Técnico criado com ID {tecnico_id}")
        return jsonify({'success': True, 'id': tecnico_id, 'message': 'Técnico cadastrado com sucesso!'})
    except Exception as e:
        print(f"DEBUG ERRO: {str(e)}")
        import traceback
        traceback.print_exc()
        if conn:
            conn.rollback()
        return jsonify({'success': False, 'message': str(e)})
    finally:
        if conn:
            conn.close()

@app.route('/api/tecnico/<int:tecnico_id>', methods=['PUT'])
@login_required
def editar_tecnico(tecnico_id):
    from empresa_helpers import get_empresa_id
    
    try:
        empresa_id = get_empresa_id()
        data = request.json
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            
            cursor.execute('''
                UPDATE tecnicos 
                SET nome=%s, cpf=%s, telefone=%s, email=%s, especialidade=%s, data_admissao=%s, status=%s
                WHERE id=%s AND empresa_id=%s
            ''', (
                data['nome'],
                data.get('cpf', ''),
                data.get('telefone', ''),
                data.get('email', ''),
                data.get('especialidade', ''),
                data.get('data_admissao', ''),
                data.get('status', 'Ativo'),
                tecnico_id,
                empresa_id
            ))
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            
            cursor.execute('''
                UPDATE tecnicos 
                SET nome=?, cpf=?, telefone=?, email=?, especialidade=?, data_admissao=?, status=?
                WHERE id=? AND empresa_id=?
            ''', (
                data['nome'],
                data.get('cpf', ''),
                data.get('telefone', ''),
                data.get('email', ''),
                data.get('especialidade', ''),
                data.get('data_admissao', ''),
                data.get('status', 'Ativo'),
                tecnico_id,
                empresa_id
            ))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Técnico atualizado com sucesso!'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/tecnico/<int:tecnico_id>', methods=['DELETE'])
@login_required
def excluir_tecnico(tecnico_id):
    from empresa_helpers import get_empresa_id
    
    try:
        empresa_id = get_empresa_id()
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            cursor.execute('DELETE FROM tecnicos WHERE id=%s AND empresa_id=%s', (tecnico_id, empresa_id))
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            cursor.execute('DELETE FROM tecnicos WHERE id=? AND empresa_id=?', (tecnico_id, empresa_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Técnico excluído com sucesso!'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/relatorios')
@login_required
def relatorios():
    from empresa_helpers import get_empresa_id
    
    empresa_id = get_empresa_id()
    
    # Obter filtros da query string
    data_inicial = request.args.get('data_inicial', '')
    data_final = request.args.get('data_final', '')
    veiculo_id = request.args.get('veiculo_id', '')
    
    # Valores padrão
    custos_mensais_json = []
    veiculos_mais_manutencoes = []
    tipos_manutencao_json = []
    custo_total_ano = 0
    media_mensal = 0
    total_manutencoes = 0
    manutencoes_emergenciais = 0
    veiculos = []
    
    try:
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            
            # Construir condições WHERE para PostgreSQL
            where_conditions = ["m.data_realizada IS NOT NULL", "v.empresa_id = %s"]
            params = [empresa_id]
            
            if data_inicial:
                where_conditions.append("m.data_realizada >= %s")
                params.append(data_inicial)
            
            if data_final:
                where_conditions.append("m.data_realizada <= %s")
                params.append(data_final)
            
            if veiculo_id:
                where_conditions.append("m.veiculo_id = %s")
                params.append(int(veiculo_id))
            
            where_clause = " AND ".join(where_conditions)
            
            # Custo total de manutenção por mês (PostgreSQL usa TO_CHAR)
            cursor.execute(f'''
                SELECT TO_CHAR(m.data_realizada, 'YYYY-MM') as mes, COALESCE(SUM(m.custo_total), 0) as total
                FROM manutencoes m
                JOIN veiculos v ON m.veiculo_id = v.id
                WHERE {where_clause}
                GROUP BY TO_CHAR(m.data_realizada, 'YYYY-MM')
                ORDER BY mes DESC 
                LIMIT 12
            ''', params)
            custos_mensais = cursor.fetchall()
            custos_mensais_json = [[row[0], float(row[1] or 0)] for row in custos_mensais]
            
            # Veículos com mais manutenções
            cursor.execute(f'''
                SELECT v.placa, v.modelo, COUNT(m.id) as total_manutencoes
                FROM veiculos v 
                LEFT JOIN manutencoes m ON v.id = m.veiculo_id AND m.data_realizada IS NOT NULL
                WHERE v.empresa_id = %s
                GROUP BY v.id, v.placa, v.modelo
                ORDER BY total_manutencoes DESC 
                LIMIT 10
            ''', (empresa_id,))
            veiculos_mais_manutencoes = cursor.fetchall()
            
            # Tipos de manutenção
            cursor.execute(f'''
                SELECT m.tipo, COUNT(*) as quantidade
                FROM manutencoes m
                JOIN veiculos v ON m.veiculo_id = v.id
                WHERE {where_clause}
                GROUP BY m.tipo
                ORDER BY quantidade DESC
            ''', params)
            tipos_manutencao = cursor.fetchall()
            tipos_manutencao_json = [[row[0] or 'Não especificado', int(row[1])] for row in tipos_manutencao]
            
            # Estatísticas - Custo total
            cursor.execute(f'''
                SELECT COALESCE(SUM(m.custo_total), 0) as total
                FROM manutencoes m
                JOIN veiculos v ON m.veiculo_id = v.id
                WHERE {where_clause}
            ''', params)
            custo_total_ano = float(cursor.fetchone()[0] or 0)
            
            # Média mensal
            cursor.execute(f'''
                SELECT COALESCE(AVG(custo_mensal), 0) as media
                FROM (
                    SELECT SUM(m.custo_total) as custo_mensal
                    FROM manutencoes m
                    JOIN veiculos v ON m.veiculo_id = v.id
                    WHERE {where_clause}
                    GROUP BY TO_CHAR(m.data_realizada, 'YYYY-MM')
                ) subq
            ''', params)
            media_mensal = float(cursor.fetchone()[0] or 0)
            
            # Total de manutenções
            cursor.execute(f'''
                SELECT COUNT(*) as total
                FROM manutencoes m
                JOIN veiculos v ON m.veiculo_id = v.id
                WHERE m.status = 'Concluída' AND {where_clause}
            ''', params)
            total_manutencoes = int(cursor.fetchone()[0] or 0)
            
            # Manutenções emergenciais
            cursor.execute(f'''
                SELECT COUNT(*) as total
                FROM manutencoes m
                JOIN veiculos v ON m.veiculo_id = v.id
                WHERE m.tipo = 'Emergencial' AND m.status = 'Concluída' AND {where_clause}
            ''', params)
            manutencoes_emergenciais = int(cursor.fetchone()[0] or 0)
            
            # Veículos para dropdown
            cursor.execute('SELECT id, placa, modelo FROM veiculos WHERE empresa_id = %s ORDER BY placa', (empresa_id,))
            veiculos = cursor.fetchall()
            
            cursor.close()
            conn.close()
        else:
            # SQLite original
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            
            # Construir condições WHERE
            where_conditions = ["data_realizada IS NOT NULL"]
            params = []
            
            if data_inicial:
                where_conditions.append("data_realizada >= ?")
                params.append(data_inicial)
            
            if data_final:
                where_conditions.append("data_realizada <= ?")
                params.append(data_final)
            
            if veiculo_id:
                where_conditions.append("veiculo_id = ?")
                params.append(veiculo_id)
            
            where_clause = " AND ".join(where_conditions)
            
            # Custo total de manutenção por mês
            cursor.execute(f'''
                SELECT strftime('%Y-%m', data_realizada) as mes, SUM(custo) as total
                FROM manutencoes 
                WHERE {where_clause}
                GROUP BY mes 
                ORDER BY mes DESC 
                LIMIT 12
            ''', params)
            custos_mensais = cursor.fetchall()
            custos_mensais_json = [[row[0], row[1]] for row in custos_mensais]
            
            # Veículos com mais manutenções
            cursor.execute('''
                SELECT v.placa, v.modelo, COUNT(m.id) as total_manutencoes
                FROM veiculos v 
                LEFT JOIN manutencoes m ON v.id = m.veiculo_id
                GROUP BY v.id 
                ORDER BY total_manutencoes DESC 
                LIMIT 10
            ''')
            veiculos_mais_manutencoes = cursor.fetchall()
            
            # Tipos de manutenção
            cursor.execute(f'''
                SELECT tipo, COUNT(*) as quantidade
                FROM manutencoes
                WHERE {where_clause}
                GROUP BY tipo
                ORDER BY quantidade DESC
            ''', params)
            tipos_manutencao = cursor.fetchall()
            tipos_manutencao_json = [[row[0], row[1]] for row in tipos_manutencao]
            
            # Custo total
            cursor.execute(f'''
                SELECT COALESCE(SUM(custo), 0) as total
                FROM manutencoes
                WHERE {where_clause}
            ''', params)
            custo_total_ano = cursor.fetchone()[0] or 0
            
            # Média mensal
            cursor.execute(f'''
                SELECT COALESCE(AVG(custo_mensal), 0) as media
                FROM (
                    SELECT SUM(custo) as custo_mensal
                    FROM manutencoes
                    WHERE {where_clause}
                    GROUP BY strftime('%Y-%m', data_realizada)
                )
            ''', params)
            media_mensal = cursor.fetchone()[0] or 0
            
            # Total de manutenções
            cursor.execute(f'''
                SELECT COUNT(*) as total
                FROM manutencoes
                WHERE status = 'Concluída' AND {where_clause}
            ''', params)
            total_manutencoes = cursor.fetchone()[0] or 0
            
            # Manutenções emergenciais
            cursor.execute(f'''
                SELECT COUNT(*) as total
                FROM manutencoes
                WHERE tipo = 'Emergencial' AND status = 'Concluída' AND {where_clause}
            ''', params)
            manutencoes_emergenciais = cursor.fetchone()[0] or 0
            
            # Veículos para dropdown (FILTRADO POR EMPRESA)
            cursor.execute('SELECT id, placa, modelo FROM veiculos WHERE empresa_id = ? ORDER BY placa', (empresa_id,))
            veiculos = cursor.fetchall()
            
            conn.close()
            
    except Exception as e:
        print(f"Erro ao carregar relatórios: {e}")
        import traceback
        traceback.print_exc()
        flash('Erro ao carregar dados dos relatórios.', 'danger')
    
    return render_template('relatorios.html', 
                         custos_mensais=custos_mensais_json,
                         veiculos_mais_manutencoes=veiculos_mais_manutencoes,
                         tipos_manutencao=tipos_manutencao_json,
                         custo_total_ano=custo_total_ano,
                         media_mensal=media_mensal,
                         total_manutencoes=total_manutencoes,
                         manutencoes_emergenciais=manutencoes_emergenciais,
                         veiculos=veiculos,
                         data_inicial=data_inicial,
                         data_final=data_final,
                         veiculo_id=veiculo_id)

# =============================================
# ROTAS DE RELATÓRIOS - APIs
# =============================================

@app.route('/api/relatorios/financeiro')
@login_required
def api_relatorio_financeiro():
    """
    API de relatório financeiro com comportamento dinâmico por tipo_operacao.
    SERVICO: faturamento, serviços, ticket médio, lista por cliente
    FROTA: despesas, manutenções, lista por veículo
    """
    from empresa_helpers import is_servico, get_empresa_id
    
    empresa_id = get_empresa_id()
    if not empresa_id:
        return jsonify({'success': False, 'message': 'Empresa não encontrada'}), 403
    
    # Obter filtros
    data_inicio = request.args.get('data_inicio', '')
    data_fim = request.args.get('data_fim', '')
    
    try:
        if Config.IS_POSTGRES:
            import psycopg2
            from psycopg2.extras import RealDictCursor
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            # Construir condições de filtro
            where_date = ""
            params = [empresa_id]
            
            if is_servico():
                # ========== RELATÓRIO SERVICO ==========
                base_where = "v.empresa_id = %s AND m.status IN ('FINALIZADO', 'FATURADO')"
                
                if data_inicio:
                    base_where += " AND m.financeiro_lancado_em >= %s"
                    params.append(data_inicio)
                if data_fim:
                    base_where += " AND m.financeiro_lancado_em <= %s"
                    params.append(data_fim)
                
                # Faturamento total
                cursor.execute(f"""
                    SELECT COALESCE(SUM(ms.subtotal), 0) as total
                    FROM manutencoes m
                    JOIN veiculos v ON m.veiculo_id = v.id
                    JOIN manutencao_servicos ms ON ms.manutencao_id = m.id
                    WHERE {base_where}
                """, params)
                faturamento_total = float(cursor.fetchone()['total'] or 0)
                
                # Quantidade de serviços
                cursor.execute(f"""
                    SELECT COUNT(DISTINCT m.id) as total
                    FROM manutencoes m
                    JOIN veiculos v ON m.veiculo_id = v.id
                    WHERE {base_where}
                """, params)
                qtd_servicos = cursor.fetchone()['total'] or 0
                
                # Ticket médio
                ticket_medio = faturamento_total / qtd_servicos if qtd_servicos > 0 else 0
                
                # Lista por cliente
                cursor.execute(f"""
                    SELECT c.nome as cliente_nome, 
                           COALESCE(SUM(ms.subtotal), 0) as total_faturado,
                           COUNT(DISTINCT m.id) as qtd_servicos
                    FROM manutencoes m
                    JOIN veiculos v ON m.veiculo_id = v.id
                    JOIN clientes c ON v.cliente_id = c.id
                    JOIN manutencao_servicos ms ON ms.manutencao_id = m.id
                    WHERE {base_where}
                    GROUP BY c.id, c.nome
                    ORDER BY total_faturado DESC
                """, params)
                lista_por_cliente = [dict(row) for row in cursor.fetchall()]
                
                cursor.close()
                conn.close()
                
                return jsonify({
                    'success': True,
                    'tipo_operacao': 'SERVICO',
                    'data_inicio': data_inicio,
                    'data_fim': data_fim,
                    'faturamento_total': faturamento_total,
                    'qtd_servicos': qtd_servicos,
                    'ticket_medio': round(ticket_medio, 2),
                    'lista_por_cliente': lista_por_cliente
                })
                
            else:
                # ========== RELATÓRIO FROTA ==========
                base_where = "v.empresa_id = %s AND m.status = 'Concluída'"
                
                if data_inicio:
                    base_where += " AND m.data_realizada >= %s"
                    params.append(data_inicio)
                if data_fim:
                    base_where += " AND m.data_realizada <= %s"
                    params.append(data_fim)
                
                # Despesas total
                cursor.execute(f"""
                    SELECT COALESCE(SUM(m.custo_total), 0) as total
                    FROM manutencoes m
                    JOIN veiculos v ON m.veiculo_id = v.id
                    WHERE {base_where}
                """, params)
                despesas_total = float(cursor.fetchone()['total'] or 0)
                
                # Quantidade de manutenções
                cursor.execute(f"""
                    SELECT COUNT(*) as total
                    FROM manutencoes m
                    JOIN veiculos v ON m.veiculo_id = v.id
                    WHERE {base_where}
                """, params)
                qtd_manutencoes = cursor.fetchone()['total'] or 0
                
                # Lista por veículo
                cursor.execute(f"""
                    SELECT v.placa, v.modelo, 
                           COALESCE(SUM(m.custo_total), 0) as total_gasto,
                           COUNT(m.id) as qtd_manutencoes
                    FROM manutencoes m
                    JOIN veiculos v ON m.veiculo_id = v.id
                    WHERE {base_where}
                    GROUP BY v.id, v.placa, v.modelo
                    ORDER BY total_gasto DESC
                """, params)
                lista_por_veiculo = [dict(row) for row in cursor.fetchall()]
                
                cursor.close()
                conn.close()
                
                return jsonify({
                    'success': True,
                    'tipo_operacao': 'FROTA',
                    'data_inicio': data_inicio,
                    'data_fim': data_fim,
                    'despesas_total': despesas_total,
                    'qtd_manutencoes': qtd_manutencoes,
                    'lista_por_veiculo': lista_por_veiculo
                })
        else:
            # SQLite fallback
            return jsonify({
                'success': True,
                'tipo_operacao': 'FROTA',
                'despesas_total': 0,
                'qtd_manutencoes': 0,
                'lista_por_veiculo': []
            })
            
    except Exception as e:
        print(f"Erro no relatório financeiro: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/relatorios/manutencoes')
@login_required
def api_relatorio_manutencoes():
    """
    API de relatório de manutenções/serviços com lista detalhada.
    Retorna: data, veículo, cliente (se SERVICO), valor, status
    """
    from empresa_helpers import is_servico, get_empresa_id
    
    empresa_id = get_empresa_id()
    if not empresa_id:
        return jsonify({'success': False, 'message': 'Empresa não encontrada'}), 403
    
    # Obter filtros
    data_inicio = request.args.get('data_inicio', '')
    data_fim = request.args.get('data_fim', '')
    status_filter = request.args.get('status', '')
    
    try:
        if Config.IS_POSTGRES:
            import psycopg2
            from psycopg2.extras import RealDictCursor
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            params = [empresa_id]
            
            if is_servico():
                # Query para SERVICO - inclui cliente
                base_where = "v.empresa_id = %s"
                
                if data_inicio:
                    base_where += " AND COALESCE(m.data_realizada, m.data_agendada) >= %s"
                    params.append(data_inicio)
                if data_fim:
                    base_where += " AND COALESCE(m.data_realizada, m.data_agendada) <= %s"
                    params.append(data_fim)
                if status_filter:
                    base_where += " AND m.status = %s"
                    params.append(status_filter)
                
                cursor.execute(f"""
                    SELECT 
                        m.id,
                        COALESCE(m.data_realizada, m.data_agendada) as data,
                        v.placa,
                        v.modelo,
                        c.nome as cliente_nome,
                        COALESCE(m.valor_total_servicos, 0) as valor,
                        m.status,
                        m.tipo,
                        m.descricao
                    FROM manutencoes m
                    JOIN veiculos v ON m.veiculo_id = v.id
                    LEFT JOIN clientes c ON v.cliente_id = c.id
                    WHERE {base_where}
                    ORDER BY data DESC
                    LIMIT 500
                """, params)
                
            else:
                # Query para FROTA - sem cliente
                base_where = "v.empresa_id = %s"
                
                if data_inicio:
                    base_where += " AND COALESCE(m.data_realizada, m.data_agendada) >= %s"
                    params.append(data_inicio)
                if data_fim:
                    base_where += " AND COALESCE(m.data_realizada, m.data_agendada) <= %s"
                    params.append(data_fim)
                if status_filter:
                    base_where += " AND m.status = %s"
                    params.append(status_filter)
                
                cursor.execute(f"""
                    SELECT 
                        m.id,
                        COALESCE(m.data_realizada, m.data_agendada) as data,
                        v.placa,
                        v.modelo,
                        COALESCE(m.custo_total, 0) as valor,
                        m.status,
                        m.tipo,
                        m.descricao
                    FROM manutencoes m
                    JOIN veiculos v ON m.veiculo_id = v.id
                    WHERE {base_where}
                    ORDER BY data DESC
                    LIMIT 500
                """, params)
            
            registros = [dict(row) for row in cursor.fetchall()]
            
            # Formatar datas
            for reg in registros:
                if reg.get('data'):
                    reg['data'] = str(reg['data'])
            
            cursor.close()
            conn.close()
            
            return jsonify({
                'success': True,
                'tipo_operacao': 'SERVICO' if is_servico() else 'FROTA',
                'data_inicio': data_inicio,
                'data_fim': data_fim,
                'total_registros': len(registros),
                'registros': registros
            })
        else:
            return jsonify({
                'success': True,
                'tipo_operacao': 'FROTA',
                'registros': []
            })
            
    except Exception as e:
        print(f"Erro no relatório de manutenções: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


# =============================================
# ROTAS DE EXPORTAÇÃO DE RELATÓRIOS
# =============================================

@app.route('/exportar/excel')
@login_required
def exportar_excel():
    """Exporta relatório completo em formato XLSX (Excel real) com comportamento dinâmico"""
    from empresa_helpers import is_servico, get_empresa_id
    
    empresa_id = get_empresa_id()
    if not empresa_id:
        return jsonify({'error': 'Empresa não encontrada'}), 403
    
    try:
        # Obter filtros
        data_inicial = request.args.get('data_inicial', '')
        data_final = request.args.get('data_final', '')
        
        if Config.IS_POSTGRES:
            import psycopg2
            from psycopg2.extras import RealDictCursor
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            params = [empresa_id]
            
            if is_servico():
                # Query SERVICO - inclui cliente e valor_total_servicos
                base_where = "v.empresa_id = %s"
                if data_inicial:
                    base_where += " AND COALESCE(m.data_realizada, m.data_agendada) >= %s"
                    params.append(data_inicial)
                if data_final:
                    base_where += " AND COALESCE(m.data_realizada, m.data_agendada) <= %s"
                    params.append(data_final)
                
                cursor.execute(f"""
                    SELECT 
                        m.id,
                        COALESCE(m.data_realizada, m.data_agendada) as data,
                        v.placa,
                        v.modelo,
                        c.nome as cliente,
                        m.tipo,
                        m.descricao,
                        COALESCE(m.valor_total_servicos, 0) as valor,
                        m.status
                    FROM manutencoes m
                    JOIN veiculos v ON m.veiculo_id = v.id
                    LEFT JOIN clientes c ON v.cliente_id = c.id
                    WHERE {base_where}
                    ORDER BY data DESC
                """, params)
                
                registros = cursor.fetchall()
                headers = ['ID', 'Data', 'Placa', 'Modelo', 'Cliente', 'Tipo', 'Descrição', 'Valor (R$)', 'Status']
                
            else:
                # Query FROTA - sem cliente, usa custo_total
                base_where = "v.empresa_id = %s"
                if data_inicial:
                    base_where += " AND COALESCE(m.data_realizada, m.data_agendada) >= %s"
                    params.append(data_inicial)
                if data_final:
                    base_where += " AND COALESCE(m.data_realizada, m.data_agendada) <= %s"
                    params.append(data_final)
                
                cursor.execute(f"""
                    SELECT 
                        m.id,
                        COALESCE(m.data_realizada, m.data_agendada) as data,
                        v.placa,
                        v.modelo,
                        m.tipo,
                        m.descricao,
                        COALESCE(m.custo_total, 0) as valor,
                        m.status,
                        m.tecnico
                    FROM manutencoes m
                    JOIN veiculos v ON m.veiculo_id = v.id
                    WHERE {base_where}
                    ORDER BY data DESC
                """, params)
                
                registros = cursor.fetchall()
                headers = ['ID', 'Data', 'Placa', 'Modelo', 'Tipo', 'Descrição', 'Custo (R$)', 'Status', 'Técnico']
            
            # Calcular total
            total_valor = sum(float(r.get('valor', 0) or 0) for r in registros)
            
            cursor.close()
            conn.close()
        else:
            registros = []
            headers = ['ID', 'Data', 'Placa', 'Modelo', 'Tipo', 'Valor', 'Status']
            total_valor = 0
        
        # Criar Excel usando openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        total_fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        tipo_relatorio = "Faturamento" if is_servico() else "Despesas"
        ws.merge_cells('A1:I1')
        ws['A1'] = f"RELATÓRIO DE {tipo_relatorio.upper()}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Período
        ws.merge_cells('A2:I2')
        periodo = f"Período: {data_inicial or 'Início'} até {data_final or 'Hoje'}"
        ws['A2'] = periodo
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Cabeçalhos (linha 4)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Dados
        for row_num, registro in enumerate(registros, 5):
            if is_servico():
                values = [
                    registro.get('id'),
                    str(registro.get('data', '')) if registro.get('data') else '',
                    registro.get('placa', ''),
                    registro.get('modelo', ''),
                    registro.get('cliente', ''),
                    registro.get('tipo', ''),
                    registro.get('descricao', ''),
                    float(registro.get('valor', 0) or 0),
                    registro.get('status', '')
                ]
            else:
                values = [
                    registro.get('id'),
                    str(registro.get('data', '')) if registro.get('data') else '',
                    registro.get('placa', ''),
                    registro.get('modelo', ''),
                    registro.get('tipo', ''),
                    registro.get('descricao', ''),
                    float(registro.get('valor', 0) or 0),
                    registro.get('status', ''),
                    registro.get('tecnico', '') or ''
                ]
            
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
                if col == (8 if is_servico() else 7):  # Coluna de valor
                    cell.number_format = 'R$ #,##0.00'
        
        # Linha de total
        total_row = len(registros) + 5
        ws.merge_cells(f'A{total_row}:F{total_row}' if not is_servico() else f'A{total_row}:G{total_row}')
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, color="FFFFFF")
        ws.cell(row=total_row, column=1).fill = total_fill
        
        valor_col = 8 if is_servico() else 7
        ws.cell(row=total_row, column=valor_col, value=total_valor)
        ws.cell(row=total_row, column=valor_col).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=total_row, column=valor_col).fill = total_fill
        ws.cell(row=total_row, column=valor_col).number_format = 'R$ #,##0.00'
        
        # Ajustar largura das colunas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 40)
        
        # Salvar em buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"relatorio_{tipo_relatorio.lower()}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename={filename}'
            }
        )
        
    except Exception as e:
        print(f"Erro na exportação Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/exportar/pdf-completo')
@login_required
def exportar_pdf_completo():
    """Exporta relatório COMPLETO com todas as manutenções em PDF - dinâmico por tipo_operacao"""
    from empresa_helpers import is_servico, get_empresa_id
    
    empresa_id = get_empresa_id()
    if not empresa_id:
        return jsonify({'error': 'Empresa não encontrada'}), 403
    
    try:
        from io import BytesIO
        
        # Obter filtros
        data_inicial = request.args.get('data_inicial', '')
        data_final = request.args.get('data_final', '')
        
        if Config.IS_POSTGRES:
            import psycopg2
            from psycopg2.extras import RealDictCursor
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            params = [empresa_id]
            
            if is_servico():
                # Query SERVICO
                base_where = "v.empresa_id = %s"
                if data_inicial:
                    base_where += " AND COALESCE(m.data_realizada, m.data_agendada) >= %s"
                    params.append(data_inicial)
                if data_final:
                    base_where += " AND COALESCE(m.data_realizada, m.data_agendada) <= %s"
                    params.append(data_final)
                
                cursor.execute(f"""
                    SELECT 
                        m.id,
                        COALESCE(m.data_realizada, m.data_agendada) as data,
                        v.placa,
                        v.modelo,
                        c.nome as cliente,
                        m.tipo,
                        m.descricao,
                        COALESCE(m.valor_total_servicos, 0) as valor,
                        m.status
                    FROM manutencoes m
                    JOIN veiculos v ON m.veiculo_id = v.id
                    LEFT JOIN clientes c ON v.cliente_id = c.id
                    WHERE {base_where}
                    ORDER BY data DESC
                """, params)
                
                registros = cursor.fetchall()
                titulo = "RELATÓRIO DE FATURAMENTO - SERVIÇOS"
                col_valor = "Valor"
                headers = ['ID', 'Placa', 'Cliente', 'Tipo', 'Descrição', 'Data', 'Valor', 'Status']
                
            else:
                # Query FROTA
                base_where = "v.empresa_id = %s"
                if data_inicial:
                    base_where += " AND COALESCE(m.data_realizada, m.data_agendada) >= %s"
                    params.append(data_inicial)
                if data_final:
                    base_where += " AND COALESCE(m.data_realizada, m.data_agendada) <= %s"
                    params.append(data_final)
                
                cursor.execute(f"""
                    SELECT 
                        m.id,
                        COALESCE(m.data_realizada, m.data_agendada) as data,
                        v.placa,
                        v.modelo,
                        m.tipo,
                        m.descricao,
                        COALESCE(m.custo_total, 0) as valor,
                        m.status
                    FROM manutencoes m
                    JOIN veiculos v ON m.veiculo_id = v.id
                    WHERE {base_where}
                    ORDER BY data DESC
                """, params)
                
                registros = cursor.fetchall()
                titulo = "RELATÓRIO DE DESPESAS - MANUTENÇÕES"
                col_valor = "Custo"
                headers = ['ID', 'Placa', 'Modelo', 'Tipo', 'Descrição', 'Data', 'Custo', 'Status']
            
            # Calcular totais
            total_valor = sum(float(r.get('valor', 0) or 0) for r in registros)
            media_valor = total_valor / len(registros) if registros else 0
            
            cursor.close()
            conn.close()
        else:
            registros = []
            titulo = "RELATÓRIO"
            headers = ['ID', 'Placa', 'Tipo', 'Data', 'Valor', 'Status']
            total_valor = 0
            media_valor = 0
        
        # Criar PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#0066cc') if not is_servico() else colors.HexColor('#28a745'),
            spaceAfter=20,
            alignment=1
        )
        elements.append(Paragraph(titulo, title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Período
        periodo_texto = f"Período: {data_inicial or 'Início'} até {data_final or 'Hoje'}"
        elements.append(Paragraph(periodo_texto, styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))
        
        # Resumo
        resumo_style = ParagraphStyle(
            'Resumo',
            parent=styles['Normal'],
            fontSize=11,
            textColor=colors.HexColor('#28a745') if is_servico() else colors.HexColor('#dc3545'),
            spaceAfter=5
        )
        
        total_formatado = f'R$ {total_valor:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
        media_formatada = f'R$ {media_valor:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
        label_total = "Total Faturado" if is_servico() else "Total Gasto"
        
        elements.append(Paragraph(f'<b>{label_total}:</b> {total_formatado} | <b>Média:</b> {media_formatada} | <b>Quantidade:</b> {len(registros)}', resumo_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Tabela
        data_table = [headers]
        
        for r in registros:
            valor_formatado = f'R$ {float(r.get("valor", 0) or 0):,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
            descricao = r.get('descricao', '') or '-'
            if len(descricao) > 25:
                descricao = descricao[:25] + '...'
            
            if is_servico():
                data_table.append([
                    str(r.get('id', '')),
                    r.get('placa', '-'),
                    (r.get('cliente', '') or '-')[:15],
                    r.get('tipo', '-'),
                    descricao,
                    str(r.get('data', '-')),
                    valor_formatado,
                    r.get('status', '-')
                ])
            else:
                data_table.append([
                    str(r.get('id', '')),
                    r.get('placa', '-'),
                    r.get('modelo', '-'),
                    r.get('tipo', '-'),
                    descricao,
                    str(r.get('data', '-')),
                    valor_formatado,
                    r.get('status', '-')
                ])
        
        # Criar tabela com larguras ajustadas
        col_widths = [0.4*inch, 0.8*inch, 0.9*inch, 0.8*inch, 1.5*inch, 0.9*inch, 0.9*inch, 0.8*inch]
        table = Table(data_table, colWidths=col_widths, repeatRows=1)
        
        header_color = colors.HexColor('#28a745') if is_servico() else colors.HexColor('#0066cc')
        
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), header_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('ALIGN', (6, 1), (6, -1), 'RIGHT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Rodapé
        elements.append(Paragraph(
            f'Relatório gerado em: {datetime.now().strftime("%d/%m/%Y às %H:%M")} | Total de registros: {len(registros)}',
            styles['Italic']
        ))
        
        doc.build(elements)
        buffer.seek(0)
        
        tipo_arquivo = "faturamento" if is_servico() else "despesas"
        filename = f"relatorio_{tipo_arquivo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
        
    except Exception as e:
        print(f"Erro na exportação PDF completo: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/exportar/pdf')
@login_required
def exportar_pdf():
    """Exporta resumo executivo em PDF - dinâmico por tipo_operacao"""
    from empresa_helpers import is_servico, get_empresa_id
    
    empresa_id = get_empresa_id()
    if not empresa_id:
        return jsonify({'error': 'Empresa não encontrada'}), 403
    
    try:
        from io import BytesIO
        
        # Obter filtros
        data_inicial = request.args.get('data_inicial', '')
        data_final = request.args.get('data_final', '')
        
        if Config.IS_POSTGRES:
            import psycopg2
            from psycopg2.extras import RealDictCursor
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            params = [empresa_id]
            
            if is_servico():
                # Estatísticas SERVICO
                base_where = "v.empresa_id = %s AND m.status IN ('FINALIZADO', 'FATURADO')"
                if data_inicial:
                    base_where += " AND m.financeiro_lancado_em >= %s"
                    params.append(data_inicial)
                if data_final:
                    base_where += " AND m.financeiro_lancado_em <= %s"
                    params.append(data_final)
                
                # Total faturado
                cursor.execute(f"""
                    SELECT 
                        COALESCE(SUM(ms.subtotal), 0) as total,
                        COUNT(DISTINCT m.id) as quantidade
                    FROM manutencoes m
                    JOIN veiculos v ON m.veiculo_id = v.id
                    JOIN manutencao_servicos ms ON ms.manutencao_id = m.id
                    WHERE {base_where}
                """, params)
                stats = cursor.fetchone()
                total = float(stats['total'] or 0)
                quantidade = stats['quantidade'] or 0
                media = total / quantidade if quantidade > 0 else 0
                
                # Top 5 clientes
                cursor.execute(f"""
                    SELECT c.nome, COUNT(DISTINCT m.id) as total_servicos, 
                           COALESCE(SUM(ms.subtotal), 0) as faturamento
                    FROM manutencoes m
                    JOIN veiculos v ON m.veiculo_id = v.id
                    JOIN clientes c ON v.cliente_id = c.id
                    JOIN manutencao_servicos ms ON ms.manutencao_id = m.id
                    WHERE {base_where}
                    GROUP BY c.id, c.nome
                    ORDER BY faturamento DESC
                    LIMIT 5
                """, params)
                top_lista = cursor.fetchall()
                
                titulo = "RESUMO EXECUTIVO - FATURAMENTO"
                label_total = "Total Faturado"
                label_top = "TOP 5 CLIENTES"
                col_headers = ['Cliente', 'Serviços', 'Faturamento']
                
            else:
                # Estatísticas FROTA
                base_where = "v.empresa_id = %s AND m.status = 'Concluída'"
                if data_inicial:
                    base_where += " AND m.data_realizada >= %s"
                    params.append(data_inicial)
                if data_final:
                    base_where += " AND m.data_realizada <= %s"
                    params.append(data_final)
                
                cursor.execute(f"""
                    SELECT 
                        COALESCE(SUM(m.custo_total), 0) as total,
                        COALESCE(AVG(m.custo_total), 0) as media,
                        COUNT(*) as quantidade
                    FROM manutencoes m
                    JOIN veiculos v ON m.veiculo_id = v.id
                    WHERE {base_where}
                """, params)
                stats = cursor.fetchone()
                total = float(stats['total'] or 0)
                media = float(stats['media'] or 0)
                quantidade = stats['quantidade'] or 0
                
                # Top 5 veículos
                cursor.execute(f"""
                    SELECT v.placa, v.modelo, COUNT(m.id) as total_manut, 
                           COALESCE(SUM(m.custo_total), 0) as custo_total
                    FROM veiculos v
                    JOIN manutencoes m ON v.id = m.veiculo_id
                    WHERE {base_where}
                    GROUP BY v.id, v.placa, v.modelo
                    ORDER BY custo_total DESC
                    LIMIT 5
                """, params)
                top_lista = cursor.fetchall()
                
                titulo = "RESUMO EXECUTIVO - DESPESAS"
                label_total = "Total Gasto"
                label_top = "TOP 5 VEÍCULOS"
                col_headers = ['Placa', 'Manutenções', 'Custo Total']
            
            cursor.close()
            conn.close()
        else:
            total = 0
            media = 0
            quantidade = 0
            top_lista = []
            titulo = "RESUMO EXECUTIVO"
            label_total = "Total"
            label_top = "TOP 5"
            col_headers = ['Item', 'Qtd', 'Valor']
        
        # Criar PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=22,
            textColor=colors.HexColor('#28a745') if is_servico() else colors.HexColor('#0066cc'),
            spaceAfter=30,
            alignment=1
        )
        elements.append(Paragraph(titulo, title_style))
        elements.append(Spacer(1, 0.3*inch))
        
        # Período
        periodo_texto = f"Período: {data_inicial or 'Início'} até {data_final or 'Hoje'}"
        elements.append(Paragraph(periodo_texto, styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))
        
        # Estatísticas
        data_stats = [
            ['ESTATÍSTICAS GERAIS', ''],
            [label_total, f'R$ {total:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')],
            ['Ticket Médio' if is_servico() else 'Custo Médio', f'R$ {media:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')],
            ['Quantidade de Serviços' if is_servico() else 'Quantidade de Manutenções', str(quantidade)]
        ]
        
        table_stats = Table(data_stats, colWidths=[3*inch, 2*inch])
        header_color = colors.HexColor('#28a745') if is_servico() else colors.HexColor('#0066cc')
        table_stats.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), header_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table_stats)
        elements.append(Spacer(1, 0.5*inch))
        
        # Top lista
        elements.append(Paragraph(label_top, styles['Heading2']))
        elements.append(Spacer(1, 0.2*inch))
        
        data_top = [col_headers]
        for item in top_lista:
            if is_servico():
                valor_formatado = f'R$ {float(item.get("faturamento", 0)):,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
                data_top.append([item.get('nome', '-'), str(item.get('total_servicos', 0)), valor_formatado])
            else:
                valor_formatado = f'R$ {float(item.get("custo_total", 0)):,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
                data_top.append([f"{item.get('placa', '-')} - {item.get('modelo', '')}", str(item.get('total_manut', 0)), valor_formatado])
        
        table_top = Table(data_top, colWidths=[2.5*inch, 1.2*inch, 1.5*inch])
        table_top.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), header_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table_top)
        elements.append(Spacer(1, 0.5*inch))
        
        # Rodapé
        elements.append(Paragraph(
            f'Relatório gerado em: {datetime.now().strftime("%d/%m/%Y às %H:%M")}',
            styles['Italic']
        ))
        
        doc.build(elements)
        buffer.seek(0)
        
        tipo_arquivo = "faturamento" if is_servico() else "despesas"
        filename = f"resumo_{tipo_arquivo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
        
    except Exception as e:
        print(f"Erro na exportação PDF: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/exportar/csv')
@login_required
def exportar_csv():
    """Exporta dados brutos em CSV - dinâmico por tipo_operacao"""
    from empresa_helpers import is_servico, get_empresa_id
    
    empresa_id = get_empresa_id()
    if not empresa_id:
        return jsonify({'error': 'Empresa não encontrada'}), 403
    
    try:
        # Obter filtros
        data_inicial = request.args.get('data_inicial', '')
        data_final = request.args.get('data_final', '')
        
        if Config.IS_POSTGRES:
            import psycopg2
            from psycopg2.extras import RealDictCursor
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            params = [empresa_id]
            
            if is_servico():
                base_where = "v.empresa_id = %s"
                if data_inicial:
                    base_where += " AND COALESCE(m.data_realizada, m.data_agendada) >= %s"
                    params.append(data_inicial)
                if data_final:
                    base_where += " AND COALESCE(m.data_realizada, m.data_agendada) <= %s"
                    params.append(data_final)
                
                cursor.execute(f"""
                    SELECT 
                        m.id,
                        COALESCE(m.data_realizada, m.data_agendada) as data,
                        v.placa,
                        v.modelo,
                        c.nome as cliente,
                        m.tipo,
                        m.descricao,
                        COALESCE(m.valor_total_servicos, 0) as valor,
                        m.status
                    FROM manutencoes m
                    JOIN veiculos v ON m.veiculo_id = v.id
                    LEFT JOIN clientes c ON v.cliente_id = c.id
                    WHERE {base_where}
                    ORDER BY data DESC
                """, params)
                
                registros = cursor.fetchall()
                headers = ['id', 'data', 'placa', 'modelo', 'cliente', 'tipo', 'descricao', 'valor', 'status']
                
            else:
                base_where = "v.empresa_id = %s"
                if data_inicial:
                    base_where += " AND COALESCE(m.data_realizada, m.data_agendada) >= %s"
                    params.append(data_inicial)
                if data_final:
                    base_where += " AND COALESCE(m.data_realizada, m.data_agendada) <= %s"
                    params.append(data_final)
                
                cursor.execute(f"""
                    SELECT 
                        m.id,
                        COALESCE(m.data_realizada, m.data_agendada) as data,
                        v.placa,
                        v.modelo,
                        m.tipo,
                        m.descricao,
                        COALESCE(m.custo_total, 0) as custo,
                        m.status,
                        m.tecnico
                    FROM manutencoes m
                    JOIN veiculos v ON m.veiculo_id = v.id
                    WHERE {base_where}
                    ORDER BY data DESC
                """, params)
                
                registros = cursor.fetchall()
                headers = ['id', 'data', 'placa', 'modelo', 'tipo', 'descricao', 'custo', 'status', 'tecnico']
            
            cursor.close()
            conn.close()
        else:
            registros = []
            headers = ['id', 'data', 'placa', 'modelo', 'tipo', 'valor', 'status']
        
        # Criar CSV com UTF-8 BOM
        output = BytesIO()
        output.write('\ufeff'.encode('utf-8'))
        
        wrapper = codecs.getwriter('utf-8')(output)
        writer = csv.writer(wrapper, lineterminator='\n')
        
        # Cabeçalho
        writer.writerow(headers)
        
        # Dados
        for registro in registros:
            row = []
            for header in headers:
                value = registro.get(header, '')
                if value is None:
                    row.append('')
                elif isinstance(value, float):
                    row.append(f"{value:.2f}")
                else:
                    row.append(str(value))
            writer.writerow(row)
        
        output.seek(0)
        
        tipo_arquivo = "servicos" if is_servico() else "manutencoes"
        filename = f"dados_{tipo_arquivo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        return Response(
            output.getvalue(),
            mimetype='text/csv; charset=utf-8',
            headers={
                'Content-Disposition': f'attachment; filename={filename}',
                'Content-Type': 'text/csv; charset=utf-8'
            }
        )
        
    except Exception as e:
        print(f"Erro na exportação CSV: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


def gerar_catalogo_pdf(empresa_id):
    """Gera um PDF detalhado com todas as peças do catálogo da empresa"""
    # Buscar todas as peças do banco (filtrado por empresa)
    if Config.IS_POSTGRES:
        import psycopg2
        conn = psycopg2.connect(Config.DATABASE_URL)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT p.nome, p.codigo, p.veiculo_compativel, p.preco, p.quantidade_estoque, f.nome as fornecedor
            FROM pecas p 
            LEFT JOIN fornecedores f ON p.fornecedor_id = f.id 
            WHERE p.empresa_id = %s
            ORDER BY p.nome
        ''', (empresa_id,))
    else:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT p.nome, p.codigo, p.veiculo_compativel, p.preco, p.quantidade_estoque, f.nome as fornecedor
            FROM pecas p 
            LEFT JOIN fornecedores f ON p.fornecedor_id = f.id 
            WHERE p.empresa_id = ?
            ORDER BY p.nome
        ''', (empresa_id,))
    pecas = cursor.fetchall()
    conn.close()
    
    # Criar diretório para PDFs se não existir
    pdf_dir = os.path.join(os.path.dirname(__file__), 'static', 'pdfs')
    os.makedirs(pdf_dir, exist_ok=True)
    
    # Caminho do arquivo PDF
    pdf_path = os.path.join(pdf_dir, 'catalogo_pecas_detalhado.pdf')
    
    # Criar o documento PDF
    doc = SimpleDocTemplate(pdf_path, pagesize=A4)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Centralizado
    )
    
    # Título
    title = Paragraph("CATÁLOGO DETALHADO DE PEÇAS", title_style)
    elements.append(title)
    elements.append(Spacer(1, 20))
    
    # Data de geração
    data_atual = datetime.now().strftime("%d/%m/%Y às %H:%M")
    data_para = Paragraph(f"<b>Gerado em:</b> {data_atual}", styles['Normal'])
    elements.append(data_para)
    elements.append(Spacer(1, 20))
    
    # Criar tabela com dados das peças
    data = [['#', 'Nome da Peça', 'Código', 'Veículo Compatível', 'Preço (R$)', 'Estoque', 'Fornecedor']]
    
    for i, (nome, codigo, veiculo, preco, estoque, fornecedor) in enumerate(pecas, 1):
        data.append([
            str(i),
            nome,
            codigo or 'N/A',
            veiculo or 'Universal',
            f'{preco:.2f}' if preco else '0.00',
            str(estoque) if estoque else '0',
            fornecedor or 'Não informado'
        ])
    
    # Criar e estilizar a tabela
    table = Table(data, colWidths=[0.5*inch, 1.8*inch, 1*inch, 1.5*inch, 0.8*inch, 0.6*inch, 1.3*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 20))
    
    # Resumo estatístico
    total_pecas = len(pecas)
    valor_total = sum(preco * estoque for _, _, _, preco, estoque, _ in pecas if preco and estoque)
    estoque_total = sum(estoque for _, _, _, _, estoque, _ in pecas if estoque)
    
    resumo = f"""
    <b>RESUMO ESTATÍSTICO:</b><br/>
    • Total de peças catalogadas: {total_pecas}<br/>
    • Total de unidades em estoque: {estoque_total}<br/>
    • Valor total do estoque: R$ {valor_total:,.2f}<br/>
    • Média de preço por peça: R$ {(sum(preco for _, _, _, preco, _, _ in pecas if preco) / total_pecas):,.2f}
    """
    
    resumo_para = Paragraph(resumo, styles['Normal'])
    elements.append(resumo_para)
    
    # Gerar o PDF
    doc.build(elements)
    return pdf_path

# Rota para servir catálogos PDF
@app.route('/catalogo-pdf')
@login_required
def catalogo_pdf():
    from empresa_helpers import get_empresa_id
    
    try:
        empresa_id = get_empresa_id()
        # Gerar PDF atualizado
        pdf_path = gerar_catalogo_pdf(empresa_id)
        return send_file(pdf_path, as_attachment=False, download_name='catalogo_pecas_detalhado.pdf')
    except Exception as e:
        return f"Erro ao gerar PDF: {str(e)}", 500

# Rota para página de acesso ao catálogo
@app.route('/catalogo')
@login_required
def catalogo_acesso():
    return '''
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Catálogo de Peças - Sistema de Manutenção</title>
        <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
        <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0/css/all.min.css">
        <style>
            body { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); min-height: 100vh; }
            .container { margin-top: 50px; }
            .card { border-radius: 15px; box-shadow: 0 10px 30px rgba(0,0,0,0.3); }
            .btn-pdf { background: linear-gradient(45deg, #FF6B6B, #4ECDC4); border: none; padding: 15px 30px; font-size: 18px; border-radius: 10px; }
            .btn-pdf:hover { transform: translateY(-2px); box-shadow: 0 8px 25px rgba(0,0,0,0.3); }
        </style>
    </head>
    <body>
        <div class="container">
            <div class="row justify-content-center">
                <div class="col-md-8">
                    <div class="card">
                        <div class="card-body text-center p-5">
                            <h1 class="card-title mb-4">
                                <i class="fas fa-file-pdf text-danger me-2"></i>
                                Catálogo de Peças
                            </h1>
                            <p class="lead mb-4">Acesse o catálogo completo com todas as peças, preços e especificações técnicas.</p>
                            
                            <div class="row text-start mb-4">
                                <div class="col-md-6">
                                    <h5><i class="fas fa-check-circle text-success me-2"></i>Inclui:</h5>
                                    <ul class="list-unstyled">
                                        <li><i class="fas fa-cog me-2"></i>Todas as peças em estoque</li>
                                        <li><i class="fas fa-tag me-2"></i>Preços atualizados</li>
                                        <li><i class="fas fa-warehouse me-2"></i>Quantidades disponíveis</li>
                                    </ul>
                                </div>
                                <div class="col-md-6">
                                    <h5><i class="fas fa-star text-warning me-2"></i>Recursos:</h5>
                                    <ul class="list-unstyled">
                                        <li><i class="fas fa-print me-2"></i>Layout para impressão</li>
                                        <li><i class="fas fa-sync me-2"></i>Atualizado em tempo real</li>
                                        <li><i class="fas fa-table me-2"></i>Tabela organizada</li>
                                    </ul>
                                </div>
                            </div>
                            
                            <a href="/catalogo-pdf" class="btn btn-pdf text-white btn-lg">
                                <i class="fas fa-download me-2"></i>
                                Abrir Catálogo PDF
                            </a>
                            
                            <hr class="my-4">
                            <a href="/" class="btn btn-outline-primary">
                                <i class="fas fa-arrow-left me-2"></i>
                                Voltar ao Sistema
                            </a>
                        </div>
                    </div>
                </div>
            </div>
        </div>
    </body>
    </html>
    '''

# Integração com Chatbot
@app.route('/api/chatbot', methods=['POST'])
@login_required
def chatbot():
    from empresa_helpers import get_empresa_id
    
    empresa_id = get_empresa_id()
    mensagem = request.json.get('mensagem', '').lower().strip()
    
    # Reconhecer números das opções do menu e atalhos
    if mensagem == '1' or 'próxima' in mensagem or 'agendada' in mensagem or 'manutenção' in mensagem:
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            cursor.execute('''
                SELECT v.placa, m.tipo, m.data_agendada 
                FROM manutencoes m 
                JOIN veiculos v ON m.veiculo_id = v.id 
                WHERE m.status = 'Agendada' AND m.empresa_id = %s
                ORDER BY m.data_agendada 
                LIMIT 3
            ''', (empresa_id,))
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            cursor.execute('''
                SELECT v.placa, m.tipo, m.data_agendada 
                FROM manutencoes m 
                JOIN veiculos v ON m.veiculo_id = v.id 
                WHERE m.status = 'Agendada' AND m.empresa_id = ?
                ORDER BY m.data_agendada 
                LIMIT 3
            ''', (empresa_id,))
        proximas = cursor.fetchall()
        conn.close()
        
        resposta = "📅 PRÓXIMAS MANUTENÇÕES\n"
        resposta += "━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        for i, (placa, tipo, data) in enumerate(proximas, 1):
            resposta += f"{i}. {placa} - {tipo}\n"
            resposta += f"   Data: {data}\n\n"
        
        return jsonify({'resposta': resposta})
    
    elif mensagem == '4' or 'estoque' in mensagem or 'baixo' in mensagem:
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            cursor.execute("SELECT nome, quantidade_estoque FROM pecas WHERE quantidade_estoque < 10 AND empresa_id = %s", (empresa_id,))
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            cursor.execute("SELECT nome, quantidade_estoque FROM pecas WHERE quantidade_estoque < 10 AND empresa_id = ?", (empresa_id,))
        alertas = cursor.fetchall()
        conn.close()
        
        resposta = "⚠️ ESTOQUE BAIXO\n"
        resposta += "━━━━━━━━━━━━━━━━━\n\n"
        for i, (nome, quantidade) in enumerate(alertas, 1):
            resposta += f"{i}. {nome} - {quantidade} unidades\n"
        
        return jsonify({'resposta': resposta})
    
    elif mensagem == '2' or ('catálogo' in mensagem and 'pdf' not in mensagem) or ('catalogo' in mensagem and 'pdf' not in mensagem) or ('peças' in mensagem and 'pdf' not in mensagem) or ('pecas' in mensagem and 'pdf' not in mensagem):
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM pecas WHERE empresa_id = %s", (empresa_id,))
            total_pecas = cursor.fetchone()[0]
            
            cursor.execute("SELECT SUM(quantidade_estoque) FROM pecas WHERE empresa_id = %s", (empresa_id,))
            total_estoque = cursor.fetchone()[0] or 0
            
            cursor.execute("SELECT SUM(preco * quantidade_estoque) FROM pecas WHERE preco IS NOT NULL AND quantidade_estoque IS NOT NULL AND empresa_id = %s", (empresa_id,))
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM pecas WHERE empresa_id = ?", (empresa_id,))
            total_pecas = cursor.fetchone()[0]
            
            cursor.execute("SELECT SUM(quantidade_estoque) FROM pecas WHERE empresa_id = ?", (empresa_id,))
            total_estoque = cursor.fetchone()[0] or 0
            
            cursor.execute("SELECT SUM(preco * quantidade_estoque) FROM pecas WHERE preco IS NOT NULL AND quantidade_estoque IS NOT NULL AND empresa_id = ?", (empresa_id,))
        valor_total = cursor.fetchone()[0] or 0
        conn.close()
        
        resposta = "� RESUMO DO ESTOQUE\n"
        resposta += "━━━━━━━━━━━━━━━━━━━━\n\n"
        resposta += f"🔢 Total de peças: {total_pecas}\n"
        resposta += f"📦 Unidades em estoque: {total_estoque:,}\n"
        resposta += f"💰 Valor total: R$ {valor_total:,.2f}\n\n"
        resposta += "Para ver lista completa digite: 3"
        
        return jsonify({'resposta': resposta})
    
    elif mensagem == '3' or 'pdf' in mensagem:
        resposta = "� CATÁLOGO PDF\n"
        resposta += "━━━━━━━━━━━━━━━━\n\n"
        resposta += "Clique no botão abaixo:\n\n"
        
        resposta += '<a href="http://127.0.0.1:5000/catalogo" target="_blank" style="background: linear-gradient(45deg, #007bff, #0056b3); color: white; padding: 12px 24px; border-radius: 8px; text-decoration: none; font-weight: bold; font-size: 16px; display: inline-block; margin: 10px 0; box-shadow: 0 4px 15px rgba(0,123,255,0.3); transition: all 0.3s ease;">📄 ABRIR CATÁLOGO</a>\n\n'
        
        resposta += "Lista completa com todas as peças,\npreços e quantidades em estoque."
        
        return jsonify({'resposta': resposta})
    
    elif mensagem == '5' or 'fornecedor' in mensagem or 'contato' in mensagem:
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            cursor.execute("SELECT nome, telefone, especialidade FROM fornecedores WHERE empresa_id = %s LIMIT 3", (empresa_id,))
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            cursor.execute("SELECT nome, telefone, especialidade FROM fornecedores WHERE empresa_id = ? LIMIT 3", (empresa_id,))
        fornecedores = cursor.fetchall()
        conn.close()
        
        resposta = "📞 FORNECEDORES\n"
        resposta += "━━━━━━━━━━━━━━━━\n\n"
        for i, (nome, telefone, especialidade) in enumerate(fornecedores, 1):
            resposta += f"{i}. {nome}\n"
            resposta += f"   {especialidade}\n"
            resposta += f"   📱 {telefone}\n\n"
        
        return jsonify({'resposta': resposta})
    
    else:
        resposta = "🤖 ASSISTENTE DE MANUTENÇÃO\n"
        resposta += "━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        resposta += "Digite o número da opção:\n\n"
        
        resposta += "1 📅 Ver próximas manutenções\n"
        resposta += "2 � Resumo das peças\n"
        resposta += "3 � Catálogo PDF completo\n"
        resposta += "4 ⚠️  Ver estoque baixo\n"
        resposta += "5 📞 Contatos fornecedores\n\n"
        
        resposta += "Exemplo: Digite apenas \"1\" ou \"2\"\n\n"
        resposta += "❓ O que você precisa?"
        
        return jsonify({'resposta': resposta})

# =============================================
# ROTAS DE MÉTRICAS DO DASHBOARD
# =============================================

@app.route('/api/dashboard/metrics')
@login_required
def dashboard_metrics():
    """
    Retorna métricas do dashboard baseadas no tipo_operacao da empresa.
    FROTA: despesas, manutenções concluídas, top veículos por custo
    SERVICO: faturamento, serviços finalizados, top clientes, ticket médio
    """
    from empresa_helpers import is_servico, is_frota, get_empresa_id
    from datetime import datetime, timedelta
    
    empresa_id = get_empresa_id()
    if not empresa_id:
        return jsonify({'success': False, 'message': 'Empresa não encontrada'}), 403
    
    try:
        if Config.IS_POSTGRES:
            import psycopg2
            from psycopg2.extras import RealDictCursor
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            # Primeiro dia do mês atual
            hoje = datetime.now()
            primeiro_dia_mes = hoje.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            
            if is_servico():
                # ========== MÉTRICAS PARA SERVICO ==========
                
                # Faturamento do mês atual (serviços finalizados/faturados)
                cursor.execute("""
                    SELECT COALESCE(SUM(ms.subtotal), 0) as total
                    FROM manutencoes m
                    JOIN manutencao_servicos ms ON ms.manutencao_id = m.id
                    JOIN veiculos v ON m.veiculo_id = v.id
                    WHERE v.empresa_id = %s 
                    AND m.status IN ('FINALIZADO', 'FATURADO')
                    AND m.financeiro_lancado_em >= %s
                """, (empresa_id, primeiro_dia_mes))
                faturamento_mes = float(cursor.fetchone()['total'] or 0)
                
                # Faturamento últimos 30 dias
                cursor.execute("""
                    SELECT COALESCE(SUM(ms.subtotal), 0) as total
                    FROM manutencoes m
                    JOIN manutencao_servicos ms ON ms.manutencao_id = m.id
                    JOIN veiculos v ON m.veiculo_id = v.id
                    WHERE v.empresa_id = %s 
                    AND m.status IN ('FINALIZADO', 'FATURADO')
                    AND m.financeiro_lancado_em >= CURRENT_DATE - INTERVAL '30 days'
                """, (empresa_id,))
                faturamento_30_dias = float(cursor.fetchone()['total'] or 0)
                
                # Serviços finalizados no mês
                cursor.execute("""
                    SELECT COUNT(*) as total
                    FROM manutencoes m
                    JOIN veiculos v ON m.veiculo_id = v.id
                    WHERE v.empresa_id = %s 
                    AND m.status IN ('FINALIZADO', 'FATURADO')
                    AND m.data_realizada >= %s
                """, (empresa_id, primeiro_dia_mes.date()))
                servicos_finalizados_mes = cursor.fetchone()['total'] or 0
                
                # Ticket médio
                ticket_medio = faturamento_mes / servicos_finalizados_mes if servicos_finalizados_mes > 0 else 0
                
                # Top 5 clientes por faturamento (mês atual)
                cursor.execute("""
                    SELECT c.nome, COALESCE(SUM(ms.subtotal), 0) as faturamento
                    FROM manutencoes m
                    JOIN veiculos v ON m.veiculo_id = v.id
                    JOIN clientes c ON v.cliente_id = c.id
                    JOIN manutencao_servicos ms ON ms.manutencao_id = m.id
                    WHERE v.empresa_id = %s 
                    AND m.status IN ('FINALIZADO', 'FATURADO')
                    AND m.financeiro_lancado_em >= %s
                    GROUP BY c.id, c.nome
                    ORDER BY faturamento DESC
                    LIMIT 5
                """, (empresa_id, primeiro_dia_mes))
                top_clientes = [dict(row) for row in cursor.fetchall()]
                
                # Serviços pendentes
                cursor.execute("""
                    SELECT COUNT(*) FROM manutencoes m
                    JOIN veiculos v ON m.veiculo_id = v.id
                    WHERE v.empresa_id = %s AND m.status IN ('RASCUNHO', 'ORCAMENTO', 'APROVADO', 'EM_EXECUCAO')
                """, (empresa_id,))
                servicos_pendentes = cursor.fetchone()['count'] or 0
                
                cursor.close()
                conn.close()
                
                return jsonify({
                    'success': True,
                    'tipo_operacao': 'SERVICO',
                    'faturamento_mes_atual': faturamento_mes,
                    'faturamento_30_dias': faturamento_30_dias,
                    'servicos_finalizados_mes': servicos_finalizados_mes,
                    'servicos_pendentes': servicos_pendentes,
                    'ticket_medio': round(ticket_medio, 2),
                    'top_clientes': top_clientes
                })
                
            else:
                # ========== MÉTRICAS PARA FROTA ==========
                
                # Despesas do mês atual (custo_total de manutenções concluídas)
                cursor.execute("""
                    SELECT COALESCE(SUM(m.custo_total), 0) as total
                    FROM manutencoes m
                    JOIN veiculos v ON m.veiculo_id = v.id
                    WHERE v.empresa_id = %s 
                    AND m.status = 'Concluída'
                    AND m.data_realizada >= %s
                """, (empresa_id, primeiro_dia_mes.date()))
                despesas_mes = float(cursor.fetchone()['total'] or 0)
                
                # Despesas últimos 30 dias
                cursor.execute("""
                    SELECT COALESCE(SUM(m.custo_total), 0) as total
                    FROM manutencoes m
                    JOIN veiculos v ON m.veiculo_id = v.id
                    WHERE v.empresa_id = %s 
                    AND m.status = 'Concluída'
                    AND m.data_realizada >= CURRENT_DATE - INTERVAL '30 days'
                """, (empresa_id,))
                despesas_30_dias = float(cursor.fetchone()['total'] or 0)
                
                # Manutenções concluídas no mês
                cursor.execute("""
                    SELECT COUNT(*) as total
                    FROM manutencoes m
                    JOIN veiculos v ON m.veiculo_id = v.id
                    WHERE v.empresa_id = %s 
                    AND m.status = 'Concluída'
                    AND m.data_realizada >= %s
                """, (empresa_id, primeiro_dia_mes.date()))
                manutencoes_mes = cursor.fetchone()['total'] or 0
                
                # Manutenções pendentes
                cursor.execute("""
                    SELECT COUNT(*) FROM manutencoes m
                    JOIN veiculos v ON m.veiculo_id = v.id
                    WHERE v.empresa_id = %s AND m.status IN ('Agendada', 'Em Andamento')
                """, (empresa_id,))
                manutencoes_pendentes = cursor.fetchone()['count'] or 0
                
                # Top 5 veículos por custo (mês atual)
                cursor.execute("""
                    SELECT v.placa, v.modelo, COALESCE(SUM(m.custo_total), 0) as custo_total
                    FROM manutencoes m
                    JOIN veiculos v ON m.veiculo_id = v.id
                    WHERE v.empresa_id = %s 
                    AND m.status = 'Concluída'
                    AND m.data_realizada >= %s
                    GROUP BY v.id, v.placa, v.modelo
                    ORDER BY custo_total DESC
                    LIMIT 5
                """, (empresa_id, primeiro_dia_mes.date()))
                top_veiculos = [dict(row) for row in cursor.fetchall()]
                
                cursor.close()
                conn.close()
                
                return jsonify({
                    'success': True,
                    'tipo_operacao': 'FROTA',
                    'despesas_mes_atual': despesas_mes,
                    'despesas_30_dias': despesas_30_dias,
                    'manutencoes_concluidas_mes': manutencoes_mes,
                    'manutencoes_pendentes': manutencoes_pendentes,
                    'top_veiculos': top_veiculos
                })
        else:
            # SQLite - métricas básicas para desenvolvimento local
            return jsonify({
                'success': True,
                'tipo_operacao': 'FROTA',
                'despesas_mes_atual': 0,
                'despesas_30_dias': 0,
                'manutencoes_concluidas_mes': 0,
                'manutencoes_pendentes': 0,
                'top_veiculos': []
            })
            
    except Exception as e:
        print(f"Erro ao buscar métricas do dashboard: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/dashboard/timeseries')
@login_required
def dashboard_timeseries():
    """
    Retorna série temporal dos últimos 14 dias para gráfico.
    FROTA: soma de despesas por dia
    SERVICO: soma de faturamento por dia
    """
    from empresa_helpers import is_servico, get_empresa_id
    from datetime import datetime, timedelta
    
    empresa_id = get_empresa_id()
    if not empresa_id:
        return jsonify({'success': False, 'message': 'Empresa não encontrada'}), 403
    
    try:
        if Config.IS_POSTGRES:
            import psycopg2
            from psycopg2.extras import RealDictCursor
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            # Gerar últimos 14 dias
            hoje = datetime.now().date()
            data_inicio = hoje - timedelta(days=13)
            
            labels = []
            for i in range(14):
                d = data_inicio + timedelta(days=i)
                labels.append(d.strftime('%d/%m'))
            
            if is_servico():
                # Faturamento por dia (soma dos serviços finalizados)
                cursor.execute("""
                    SELECT DATE(m.financeiro_lancado_em) as dia, 
                           COALESCE(SUM(ms.subtotal), 0) as valor
                    FROM manutencoes m
                    JOIN veiculos v ON m.veiculo_id = v.id
                    JOIN manutencao_servicos ms ON ms.manutencao_id = m.id
                    WHERE v.empresa_id = %s 
                    AND m.status IN ('FINALIZADO', 'FATURADO')
                    AND m.financeiro_lancado_em >= %s
                    GROUP BY DATE(m.financeiro_lancado_em)
                    ORDER BY dia
                """, (empresa_id, data_inicio))
                
                dados_raw = {str(row['dia']): float(row['valor']) for row in cursor.fetchall()}
                tipo = 'faturamento'
                label = 'Faturamento (R$)'
                cor = 'rgb(25, 135, 84)'  # verde
                
            else:
                # Despesas por dia (custo das manutenções concluídas)
                cursor.execute("""
                    SELECT m.data_realizada as dia, 
                           COALESCE(SUM(m.custo_total), 0) as valor
                    FROM manutencoes m
                    JOIN veiculos v ON m.veiculo_id = v.id
                    WHERE v.empresa_id = %s 
                    AND m.status = 'Concluída'
                    AND m.data_realizada >= %s
                    GROUP BY m.data_realizada
                    ORDER BY dia
                """, (empresa_id, data_inicio))
                
                dados_raw = {str(row['dia']): float(row['valor']) for row in cursor.fetchall()}
                tipo = 'despesas'
                label = 'Despesas (R$)'
                cor = 'rgb(220, 53, 69)'  # vermelho
            
            # Preencher valores para todos os dias (0 se não houver dados)
            values = []
            for i in range(14):
                d = data_inicio + timedelta(days=i)
                values.append(dados_raw.get(str(d), 0))
            
            cursor.close()
            conn.close()
            
            return jsonify({
                'success': True,
                'tipo': tipo,
                'labels': labels,
                'values': values,
                'dataset_label': label,
                'cor': cor
            })
        else:
            # SQLite - dados vazios para desenvolvimento local
            return jsonify({
                'success': True,
                'tipo': 'despesas',
                'labels': [],
                'values': [],
                'dataset_label': 'Dados',
                'cor': 'rgb(100, 100, 100)'
            })
            
    except Exception as e:
        print(f"Erro ao buscar timeseries do dashboard: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


# =============================================
# ROTAS DO MÓDULO FINANCEIRO
# =============================================

@app.route('/financeiro')
@login_required
def financeiro():
    return render_template('financeiro.html')

# Rotas para categorias
@app.route('/api/financeiro/categorias/entrada', methods=['GET'])
@login_required
def get_categorias_entrada():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT id, nome, descricao FROM categorias_entrada ORDER BY nome')
        categorias = [{'id': row[0], 'nome': row[1], 'descricao': row[2]} for row in cursor.fetchall()]
        conn.close()
        return jsonify(categorias)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/financeiro/categorias/despesa', methods=['GET'])
@login_required
def get_categorias_despesa():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT id, nome, descricao FROM categorias_despesa ORDER BY nome')
        categorias = [{'id': row[0], 'nome': row[1], 'descricao': row[2]} for row in cursor.fetchall()]
        conn.close()
        return jsonify(categorias)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

# Rotas alternativas com hífen (para compatibilidade)
@app.route('/api/financeiro/categorias-entrada', methods=['GET'])
@login_required
def get_categorias_entrada_alt():
    return get_categorias_entrada()

@app.route('/api/financeiro/categorias-despesa', methods=['GET'])
@login_required
def get_categorias_despesa_alt():
    return get_categorias_despesa()

# Rotas para entradas
@app.route('/api/financeiro/entradas', methods=['GET'])
@login_required
def get_entradas():
    from empresa_helpers import get_empresa_id
    
    try:
        empresa_id = get_empresa_id()
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            cursor.execute('''
                SELECT e.id, e.descricao, e.valor, e.data_entrada, e.categoria_id, 
                       e.veiculo_id, e.tipo, e.observacoes, e.data_criacao,
                       c.nome as categoria_nome, v.placa, v.modelo
                FROM entradas e
                LEFT JOIN categorias_entrada c ON e.categoria_id = c.id
                LEFT JOIN veiculos v ON e.veiculo_id = v.id
                WHERE e.empresa_id = %s
                ORDER BY e.data_entrada DESC
            ''', (empresa_id,))
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            cursor.execute('''
                SELECT e.id, e.descricao, e.valor, e.data_entrada, e.categoria_id, 
                       e.veiculo_id, e.tipo, e.observacoes, e.data_criacao,
                       c.nome as categoria_nome, v.placa, v.modelo
                FROM entradas e
                LEFT JOIN categorias_entrada c ON e.categoria_id = c.id
                LEFT JOIN veiculos v ON e.veiculo_id = v.id
                WHERE e.empresa_id = ?
                ORDER BY e.data_entrada DESC
            ''', (empresa_id,))
        
        entradas = []
        for row in cursor.fetchall():
            entradas.append({
                'id': row[0],
                'descricao': row[1],
                'valor': row[2],
                'data_entrada': row[3],
                'categoria_id': row[4],
                'veiculo_id': row[5],
                'tipo': row[6],
                'observacoes': row[7],
                'data_criacao': row[8],
                'categoria_nome': row[9],
                'veiculo_placa': row[10],
                'veiculo_modelo': row[11]
            })
        
        conn.close()
        return jsonify(entradas)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/financeiro/entrada', methods=['POST'])
@login_required
def create_entrada():
    from empresa_helpers import get_empresa_id
    
    try:
        empresa_id = get_empresa_id()
        data = request.json
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            
            cursor.execute('''
                INSERT INTO entradas (empresa_id, descricao, valor, data_entrada, categoria_id, veiculo_id, tipo, observacoes)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            ''', (
                empresa_id,
                data['descricao'],
                data['valor'],
                data['data_entrada'],
                data['categoria_id'],
                data.get('veiculo_id'),
                data.get('tipo', 'Manual'),
                data.get('observacoes', '')
            ))
            
            entrada_id = cursor.fetchone()[0]
        else:
            conn = sqlite3.connect(DATABASE, timeout=30.0)
            cursor = conn.cursor()
            
            cursor.execute('''
                INSERT INTO entradas (empresa_id, descricao, valor, data_entrada, categoria_id, veiculo_id, tipo, observacoes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                empresa_id,
                data['descricao'],
                data['valor'],
                data['data_entrada'],
                data['categoria_id'],
                data.get('veiculo_id'),
                data.get('tipo', 'Manual'),
                data.get('observacoes', '')
            ))
            
            entrada_id = cursor.lastrowid
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'id': entrada_id, 'message': 'Entrada cadastrada com sucesso!'})
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({'success': False, 'message': str(e)})
    finally:
        if conn:
            conn.close()

# Rotas para despesas
@app.route('/api/financeiro/despesas', methods=['GET'])
@login_required
def get_despesas():
    from empresa_helpers import get_empresa_id
    
    try:
        empresa_id = get_empresa_id()
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            cursor.execute('''
                SELECT d.id, d.descricao, d.valor, d.data_despesa, d.categoria_id, 
                       d.veiculo_id, d.tipo, d.manutencao_id, d.observacoes, d.data_criacao,
                       c.nome as categoria_nome, v.placa, v.modelo
                FROM despesas d
                LEFT JOIN categorias_despesa c ON d.categoria_id = c.id
                LEFT JOIN veiculos v ON d.veiculo_id = v.id
                WHERE d.empresa_id = %s
                ORDER BY d.data_despesa DESC
            ''', (empresa_id,))
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            cursor.execute('''
                SELECT d.id, d.descricao, d.valor, d.data_despesa, d.categoria_id, 
                       d.veiculo_id, d.tipo, d.manutencao_id, d.observacoes, d.data_criacao,
                       c.nome as categoria_nome, v.placa, v.modelo
                FROM despesas d
                LEFT JOIN categorias_despesa c ON d.categoria_id = c.id
                LEFT JOIN veiculos v ON d.veiculo_id = v.id
                WHERE d.empresa_id = ?
                ORDER BY d.data_despesa DESC
            ''', (empresa_id,))
        
        despesas = []
        for row in cursor.fetchall():
            despesas.append({
                'id': row[0],
                'descricao': row[1],
                'valor': row[2],
                'data_despesa': row[3],
                'categoria_id': row[4],
                'veiculo_id': row[5],
                'tipo': row[6],
                'manutencao_id': row[7],
                'observacoes': row[8],
                'data_criacao': row[9],
                'categoria_nome': row[10],
                'veiculo_placa': row[11],
                'veiculo_modelo': row[12]
            })
        
        conn.close()
        return jsonify(despesas)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/financeiro/despesa', methods=['POST'])
@login_required
def create_despesa():
    from empresa_helpers import get_empresa_id
    
    try:
        empresa_id = get_empresa_id()
        data = request.json
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            
            cursor.execute('''
                INSERT INTO despesas (empresa_id, descricao, valor, data_despesa, categoria_id, veiculo_id, tipo, observacoes, manutencao_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            ''', (
                empresa_id,
                data['descricao'],
                data['valor'],
                data['data_despesa'],
                data['categoria_id'],
                data.get('veiculo_id'),
                data.get('tipo', 'Manual'),
                data.get('observacoes', ''),
                data.get('manutencao_id')
            ))
            
            despesa_id = cursor.fetchone()[0]
        else:
            conn = sqlite3.connect(DATABASE, timeout=30.0)
            cursor = conn.cursor()
            
            cursor.execute('''
                INSERT INTO despesas (empresa_id, descricao, valor, data_despesa, categoria_id, veiculo_id, tipo, observacoes, manutencao_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                empresa_id,
                data['descricao'],
                data['valor'],
                data['data_despesa'],
                data['categoria_id'],
                data.get('veiculo_id'),
                data.get('tipo', 'Manual'),
                data.get('observacoes', ''),
                data.get('manutencao_id')
            ))
            
            despesa_id = cursor.lastrowid
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'id': despesa_id, 'message': 'Despesa cadastrada com sucesso!'})
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({'success': False, 'message': str(e)})
    finally:
        if conn:
            conn.close()

# Rota para relatórios financeiros
@app.route('/api/financeiro/resumo', methods=['GET'])
@login_required
def get_resumo_financeiro():
    try:
        from empresa_helpers import get_empresa_id
        empresa_id = get_empresa_id()
        
        periodo = request.args.get('periodo', '30')  # últimos 30 dias por padrão
        
        # Validar período (segurança)
        try:
            periodo_int = int(periodo)
            if periodo_int < 1 or periodo_int > 3650:  # Máximo 10 anos
                periodo_int = 30
        except ValueError:
            periodo_int = 30
        
        if Config.IS_POSTGRES:
            import psycopg2
            conn = psycopg2.connect(Config.DATABASE_URL)
            cursor = conn.cursor()
            
            # Total de entradas (PostgreSQL usa INTERVAL)
            cursor.execute('''
                SELECT COALESCE(SUM(e.valor), 0) 
                FROM entradas e
                JOIN veiculos v ON e.veiculo_id = v.id
                WHERE v.empresa_id = %s AND e.data_entrada >= CURRENT_DATE - INTERVAL '%s days'
            ''', (empresa_id, periodo_int))
            total_entradas = float(cursor.fetchone()[0] or 0)
            
            # Total de despesas
            cursor.execute('''
                SELECT COALESCE(SUM(d.valor), 0) 
                FROM despesas d
                JOIN veiculos v ON d.veiculo_id = v.id
                WHERE v.empresa_id = %s AND d.data_despesa >= CURRENT_DATE - INTERVAL '%s days'
            ''', (empresa_id, periodo_int))
            total_despesas = float(cursor.fetchone()[0] or 0)
            
            # Entradas por categoria
            cursor.execute('''
                SELECT c.nome, COALESCE(SUM(e.valor), 0)
                FROM categorias_entrada c
                LEFT JOIN entradas e ON c.id = e.categoria_id 
                    AND e.data_entrada >= CURRENT_DATE - INTERVAL '%s days'
                    AND EXISTS (SELECT 1 FROM veiculos v WHERE v.id = e.veiculo_id AND v.empresa_id = %s)
                GROUP BY c.id, c.nome
                ORDER BY SUM(e.valor) DESC NULLS LAST
            ''', (periodo_int, empresa_id))
            entradas_categoria = cursor.fetchall()
            
            # Despesas por categoria
            cursor.execute('''
                SELECT c.nome, COALESCE(SUM(d.valor), 0)
                FROM categorias_despesa c
                LEFT JOIN despesas d ON c.id = d.categoria_id 
                    AND d.data_despesa >= CURRENT_DATE - INTERVAL '%s days'
                    AND EXISTS (SELECT 1 FROM veiculos v WHERE v.id = d.veiculo_id AND v.empresa_id = %s)
                GROUP BY c.id, c.nome
                ORDER BY SUM(d.valor) DESC NULLS LAST
            ''', (periodo_int, empresa_id))
            despesas_categoria = cursor.fetchall()
            
            # Resultado por veículo
            cursor.execute('''
                SELECT v.placa, v.modelo,
                       COALESCE((SELECT SUM(e.valor) FROM entradas e WHERE e.veiculo_id = v.id AND e.data_entrada >= CURRENT_DATE - INTERVAL '%s days'), 0) as entradas,
                       COALESCE((SELECT SUM(d.valor) FROM despesas d WHERE d.veiculo_id = v.id AND d.data_despesa >= CURRENT_DATE - INTERVAL '%s days'), 0) as despesas
                FROM veiculos v
                WHERE v.empresa_id = %s
                ORDER BY (entradas - despesas) DESC
            ''', (periodo_int, periodo_int, empresa_id))
            resultado_veiculo = [(row[0], row[1], float(row[2] or 0), float(row[3] or 0), float(row[2] or 0) - float(row[3] or 0)) for row in cursor.fetchall()]
            
            cursor.close()
            conn.close()
        else:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            
            # Total de entradas (com filtro empresa_id)
            cursor.execute('SELECT COALESCE(SUM(valor), 0) FROM entradas WHERE empresa_id = ? AND data_entrada >= date("now", "-" || ? || " days")', (empresa_id, periodo_int))
            total_entradas = cursor.fetchone()[0]
            
            # Total de despesas (com filtro empresa_id)
            cursor.execute('SELECT COALESCE(SUM(valor), 0) FROM despesas WHERE empresa_id = ? AND data_despesa >= date("now", "-" || ? || " days")', (empresa_id, periodo_int))
            total_despesas = cursor.fetchone()[0]
            
            # Entradas por categoria (com filtro empresa_id)
            cursor.execute('''
                SELECT c.nome, COALESCE(SUM(e.valor), 0)
                FROM categorias_entrada c
                LEFT JOIN entradas e ON c.id = e.categoria_id AND e.empresa_id = ? AND e.data_entrada >= date("now", "-" || ? || " days")
                GROUP BY c.id, c.nome
                ORDER BY SUM(e.valor) DESC
            ''', (empresa_id, periodo_int))
            entradas_categoria = cursor.fetchall()
            
            # Despesas por categoria (com filtro empresa_id)
            cursor.execute('''
                SELECT c.nome, COALESCE(SUM(d.valor), 0)
                FROM categorias_despesa c
                LEFT JOIN despesas d ON c.id = d.categoria_id AND d.empresa_id = ? AND d.data_despesa >= date("now", "-" || ? || " days")
                GROUP BY c.id, c.nome
                ORDER BY SUM(d.valor) DESC
            ''', (empresa_id, periodo_int))
            despesas_categoria = cursor.fetchall()
            
            # Resultado por veículo (com filtro empresa_id)
            cursor.execute('''
                SELECT v.placa, v.modelo,
                       COALESCE(SUM(e.valor), 0) as entradas,
                       COALESCE(SUM(d.valor), 0) as despesas,
                       (COALESCE(SUM(e.valor), 0) - COALESCE(SUM(d.valor), 0)) as saldo
                FROM veiculos v
                LEFT JOIN entradas e ON v.id = e.veiculo_id AND e.data_entrada >= date("now", "-" || ? || " days")
                LEFT JOIN despesas d ON v.id = d.veiculo_id AND d.data_despesa >= date("now", "-" || ? || " days")
                WHERE v.empresa_id = ?
                GROUP BY v.id, v.placa, v.modelo
                HAVING entradas > 0 OR despesas > 0
                ORDER BY saldo DESC
            ''', (periodo_int, periodo_int, empresa_id))
            resultado_veiculo = cursor.fetchall()
            
            conn.close()
        
        return jsonify({
            'success': True,
            'total_entradas': total_entradas,
            'total_despesas': total_despesas,
            'saldo': total_entradas - total_despesas,
            'entradas_categoria': [{'categoria': row[0], 'valor': float(row[1] or 0)} for row in entradas_categoria],
            'despesas_categoria': [{'categoria': row[0], 'valor': float(row[1] or 0)} for row in despesas_categoria],
            'resultado_veiculo': [{'placa': row[0], 'modelo': row[1], 'entradas': float(row[2] or 0), 'despesas': float(row[3] or 0), 'saldo': float(row[4] if len(row) > 4 else (row[2] or 0) - (row[3] or 0))} for row in resultado_veiculo]
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})

# =============================================
# ROTAS DE ADMINISTRAÇÃO E GERENCIAMENTO
# =============================================

@app.route('/admin/backup')
@login_required
@admin_required
def admin_backup():
    """Criar backup do banco de dados"""
    try:
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"frota_backup_{timestamp}.db"
        backup_path = os.path.join('backups', backup_filename)
        
        # Criar diretório de backups se não existir
        os.makedirs('backups', exist_ok=True)
        
        success = db_manager.backup_database(backup_path)
        
        if success:
            log_action(current_user.id, 'BACKUP', None, None, 
                      f'Criou backup: {backup_filename}', request.remote_addr)
            return send_file(backup_path, as_attachment=True, download_name=backup_filename)
        else:
            flash('Erro ao criar backup!', 'danger')
            return redirect(url_for('dashboard'))
    except Exception as e:
        app.logger.error(f'Erro ao criar backup: {e}')
        flash(f'Erro ao criar backup: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))

@app.route('/perfil')
@login_required
def perfil():
    """Perfil do usuário logado"""
    return render_template('perfil.html', user=current_user)

@app.route('/alterar-senha', methods=['POST'])
@login_required
def alterar_senha():
    """Alterar senha do usuário logado"""
    from auth import change_password
    
    senha_atual = request.form.get('senha_atual')
    senha_nova = request.form.get('senha_nova')
    senha_confirma = request.form.get('senha_confirma')
    
    if not all([senha_atual, senha_nova, senha_confirma]):
        flash('Todos os campos são obrigatórios!', 'danger')
        return redirect(url_for('perfil'))
    
    if senha_nova != senha_confirma:
        flash('As senhas não coincidem!', 'danger')
        return redirect(url_for('perfil'))
    
    if len(senha_nova) < 6:
        flash('A senha deve ter no mínimo 6 caracteres!', 'danger')
        return redirect(url_for('perfil'))
    
    success, message = change_password(current_user.id, senha_atual, senha_nova)
    
    if success:
        log_action(current_user.id, 'CHANGE_PASSWORD', 'usuarios', current_user.id, 
                  'Alterou sua própria senha', request.remote_addr)
        flash(message, 'success')
    else:
        flash(message, 'danger')
    
    return redirect(url_for('perfil'))

if __name__ == '__main__':
    init_db()
    # Configuração para produção
    import os
    port = int(os.environ.get('PORT', 5000))
    debug_mode = os.environ.get('FLASK_ENV', 'development') == 'development'
    app.run(host='0.0.0.0', port=port, debug=debug_mode)